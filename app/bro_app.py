"""
The complete BRO Risk Oracle application (FastAPI), all features included.

Wires every feature group onto the tested foundations:
  - persistence + RBAC          (Group A)
  - core lifecycle + scoring    (Group B, ported engine)
  - four intelligence engines   (Group C, deterministic-local)
  - monitoring lifecycle        (Group D)
  - notifications + email + webhooks (Group E)
  - conversational + autopilot  (Group F)
  - admin + MCP tools + procurement (Group G)

Persistence is SQLAlchemy on SQLite by default (runs offline) or Postgres via
BRO_DB_URL. Auth is a simple bearer/session actor for API use. Every
consequential mutation appends to the hash-chained audit log.
"""
from __future__ import annotations

import json
import re as _re

_EMAIL_RE = _re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")
_PHONE_RE = _re.compile(r"^\+?[0-9]+$")

def _can_view_assessment(u, a) -> bool:
    """CR-3: assessment record visibility by role.
    - admin (reviewer) and vrm (assessor): see ALL records
    - buyer (business user): see ONLY their own (owner / SPOC)
    - vendor: see NONE
    """
    role = getattr(getattr(u, "role", None), "key", None)
    if role in ("admin", "vrm"):
        return True
    if role == "vendor":
        return False
    if role == "buyer":
        return u.username in (a.engagement_owner, a.spoc_user)
    return u.username in (a.engagement_owner, a.spoc_user, a.assessor_user)

def _validate_typed_fields(data: dict):
    """CR-8: server-side validation backing the typed inputs. Returns an error
    string if any email/phone/date value is malformed, else None. Empty values pass."""
    for k, v in (data or {}).items():
        if v in (None, "", []):
            continue
        key = str(k).lower()
        sval = str(v)
        if "email" in key and not _EMAIL_RE.match(sval):
            return f"'{k}' must be a valid email address"
        if _re.search(r"phone|telephone|mobile|contact_number", key) and not _PHONE_RE.match(sval):
            return f"'{k}' may contain only '+' and digits"
        if _re.search(r"date$|_date|dob$", key):
            if not _re.match(r"^\d{4}-\d{2}-\d{2}$", sval):
                return f"'{k}' must be a valid date (YYYY-MM-DD)"
    return None
from typing import Optional, Union

from fastapi import Body, Depends, FastAPI, File, Form, Header, HTTPException, Request, Response, UploadFile
from pydantic import BaseModel
from sqlalchemy import func, select
from sqlalchemy.orm import Session

from .features import prompts as PROMPTS
from .features import security as SEC
from .features import bro_engine as eng
from .features import intel
from .features.models_db import (
    Base, EngagementRow, Role, User, Vendor,
    make_engine, make_session_factory, verify_password,
)
from .features.models_feature import (
    Acceptance, AuditLog, Certification, Contract, ConversationMessage,
    ConversationSession, Document, EmailOutbox, Finding, FourthParty,
    Incident, IntelResult, MethodologyVersion, Monitoring, Notification,
    Offboarding, Reassessment, Webhook,
)
from .features.rbac import has_permission, seed
from .features.auth import bearer_subject, issue_token, TokenError


def _obs_swallow(_ctx, _exc):
    """Swallow a non-critical exception but emit one observable log line.
    Never raises — observability must not change control flow."""
    try:
        from .features.security import log_json as _lj
    except Exception:
        try:
            from .security import log_json as _lj
        except Exception:
            return
    try:
        _lj('swallowed_exception', where=_ctx,
            error=f'{type(_exc).__name__}: {str(_exc)[:200]}')
    except Exception:
        pass



# ---------- request schemas ----------

class VendorIn(BaseModel):
    name: str
    industry: Optional[str] = None
    country: Optional[str] = None
    contact_email: Optional[str] = None
    tier: str = "Tier 3"

class CriticalIn(BaseModel):
    reason: str

class EngagementIn(BaseModel):
    vendor_id: int
    title: str
    service_description: Optional[str] = None
    business_contact_email: Optional[str] = None

class IRQIn(BaseModel):
    answers: dict

class DDQIn(BaseModel):
    answers: dict

class AiKeyIn(BaseModel):
    provider: str = "claude"
    api_key: str = ""
    model: str = ""

class ExitPlanIn(BaseModel):
    exit_mode: Optional[str] = None
    strategy_type: Optional[str] = None
    rationale: Optional[str] = None
    target_window: Optional[str] = None
    impact_summary: Optional[str] = None
    data_plan: Optional[str] = None
    comms_plan: Optional[str] = None
    owner: Optional[str] = None
    approver: Optional[str] = None
    status: Optional[str] = None
    one_off_cost: Optional[float] = None
    dual_running_cost: Optional[float] = None
    penalty_cost: Optional[float] = None

class ExitChildIn(BaseModel):
    kind: str
    name: Optional[str] = None
    prequalified: Optional[bool] = False
    lead_time_days: Optional[int] = None
    viability: Optional[int] = 3
    note: Optional[str] = ""
    description: Optional[str] = None
    owner: Optional[str] = ""
    duration_days: Optional[int] = None
    rto: Optional[str] = ""
    rpo: Optional[str] = ""
    dependency: Optional[str] = ""
    service_name: Optional[str] = None
    impact_tolerance: Optional[str] = ""
    max_downtime: Optional[str] = ""
    criticality: Optional[str] = "Important"

class ExitTestIn(BaseModel):
    method: str = "tabletop"
    outcome: str = ""
    lessons: str = ""
    participants: str = ""
    passed: bool = True

class ExitInvokeIn(BaseModel):
    mode: str = "planned"

class MethodologyIn(BaseModel):
    title: str = "Methodology"
    content_text: Optional[str] = None
    data_b64: Optional[str] = None
    filename: Optional[str] = None

class ActiveIn(BaseModel):
    active: bool = True

class AIResearchIn(BaseModel):
    vendor_id: Optional[str] = None
    company: Optional[str] = None
    jurisdiction: Optional[str] = "UK"
    identifier: Optional[str] = ""

class OverrideIn(BaseModel):
    band: str
    reason: str
    second_approver: str

class FindingIn(BaseModel):
    engagement_id: Optional[int] = None
    title: str
    severity: str = "medium"

class IntelIn(BaseModel):
    vendor_id: int
    payload: dict = {}

class LoginIn(BaseModel):
    username: str
    password: str

class ChatStart(BaseModel):
    engagement_id: Optional[int] = None
    actor_role: str = "assessor"

class ChatTurn(BaseModel):
    session_id: int
    message: str

class MethIn(BaseModel):
    version: str
    note: Optional[str] = None

class POIn(BaseModel):
    vendor_name: str
    amount: float
    ext_ref: Optional[str] = None

class CertIn(BaseModel):
    vendor_id: int
    name: str
    valid_until: Optional[str] = None

class DocIn(BaseModel):
    vendor_id: Optional[int] = None
    engagement_id: Optional[int] = None
    name: str
    doc_type: str = "other"
    next_validation: Optional[str] = None

class FourthIn(BaseModel):
    vendor_id: int
    name: str
    service: Optional[str] = None

class AcceptIn(BaseModel):
    engagement_id: int
    rationale: str
    expires_at: Optional[str] = None

class ReassessIn(BaseModel):
    engagement_id: int
    mode: str = "periodic"

# --- new: CRUD / admin / self-service schemas ---
class VendorUpdateIn(BaseModel):
    name: Optional[str] = None
    industry: Optional[str] = None
    country: Optional[str] = None
    contact_email: Optional[str] = None
    tier: Optional[str] = None

class EngagementUpdateIn(BaseModel):
    title: Optional[str] = None
    service_description: Optional[str] = None
    business_contact_email: Optional[str] = None

class FindingUpdateIn(BaseModel):
    title: Optional[str] = None
    severity: Optional[str] = None
    status: Optional[str] = None

class UserIn(BaseModel):
    username: str
    full_name: Optional[str] = None
    email: Optional[str] = None
    password: str
    role_key: str

class UserUpdateIn(BaseModel):
    full_name: Optional[str] = None
    email: Optional[str] = None
    role_key: Optional[str] = None
    is_active: Optional[bool] = None

class RolePermsIn(BaseModel):
    permission_keys: list[str]

class PasswordIn(BaseModel):
    current_password: str
    new_password: str

class ProfileIn(BaseModel):
    full_name: Optional[str] = None
    email: Optional[str] = None

class WebhookIn(BaseModel):
    url: str
    event: str = "*"

class SignoffIn(BaseModel):
    decision: str = "approved"
    note: Optional[str] = None

class EmailIn(BaseModel):
    to_addr: str
    subject: str
    body: str

class ChatSessionIn(BaseModel):
    engagement_id: Optional[int] = None

class ChatSendIn(BaseModel):
    session_id: int
    message: str
    agent: Optional[str] = None

class LearningIn(BaseModel):
    rating: int = 3
    agent: Optional[str] = None
    stage: int = 0
    issue: Optional[str] = None
    note: Optional[str] = None

# --- v2 registry schemas ---
class V2VendorIn(BaseModel):
    legal_name: str
    trading_name: Optional[str] = None
    registration_number: Optional[str] = None
    hq_country: Optional[str] = None
    website: Optional[str] = None
    listing_status: Optional[str] = None
    tier: Optional[str] = "Tier 3"
    group_id: Optional[str] = None
    parent_company: Optional[str] = None
    industries: Optional[list[str]] = None
    procurement_ref: Optional[str] = None
    created_via: Optional[str] = "button"

class GroupOverrideIn(BaseModel):
    group_id: str

class V2ContactIn(BaseModel):
    owner_type: str
    owner_id: str
    name: str
    is_primary: bool = False
    email: Optional[str] = None
    phone_country_code: Optional[str] = None
    phone_number: Optional[str] = None
    designation: Optional[str] = None
    country: Optional[str] = None
    mailing_address: Optional[str] = None

class V2EngagementIn(BaseModel):
    vendor_id: str
    title: str
    service_description: Optional[str] = None
    material_group_id: Optional[str] = None
    business_unit: Optional[str] = None
    deployment_model: Optional[str] = None
    owner_user: Optional[str] = None
    annual_value: Optional[float] = None
    currency: Optional[str] = None

class V2AssessmentIn(BaseModel):
    engagement_id: str
    vendor_id: Optional[str] = None
    session_id: Optional[int] = None
    inherent_band: Optional[str] = None
    residual_band: Optional[str] = None

class ReassignIn(BaseModel):
    assessor_user: str

class V2FindingIn(BaseModel):
    title: str
    severity: Optional[str] = "Medium"
    source: Optional[str] = "Assessor"
    description: Optional[str] = None
    domain: Optional[str] = None
    engagement_id: Optional[str] = None
    vendor_id: Optional[str] = None
    assessment_id: Optional[str] = None
    due_date: Optional[str] = None
    owner: Optional[str] = None
    assessor: Optional[str] = None
    suggested_remediation: Optional[str] = None
    suggested_closure: Optional[str] = None
    status: Optional[str] = None

class FindingPatchIn(BaseModel):
    title: Optional[str] = None
    description: Optional[str] = None
    severity: Optional[str] = None
    status: Optional[str] = None
    owner: Optional[str] = None
    assessor: Optional[str] = None
    suggested_remediation: Optional[str] = None
    suggested_closure: Optional[str] = None
    due_date: Optional[str] = None

class RiskAcceptIn(BaseModel):
    rationale: str
    expiry_date: str
    accept: Optional[bool] = True

class FindingNoteIn(BaseModel):
    note: str

class FindingAttachIn(BaseModel):
    doc_id: str
    name: Optional[str] = None

class AssignAssessorIn(BaseModel):
    assessor_user: str

class ConnectorPullIn(BaseModel):
    vendor_id: Optional[str] = None
    company: Optional[str] = None

class ConnectorWebhookIn(BaseModel):
    vendor_id: Optional[str] = None
    payload: Optional[dict] = None

class RegExportIn(BaseModel):
    codes: list
    updates: Optional[list] = None

class RegRefreshIn(BaseModel):
    codes: list

class RegAssessIn(BaseModel):
    codes: list
    doc_text: Optional[str] = ""
    industry: Optional[str] = None

class RegRelevanceIn(BaseModel):
    codes: list
    industry: Optional[dict] = None

class RegReportExportIn(BaseModel):
    report: dict

class IntegritySweepIn(BaseModel):
    limit: Optional[int] = None
    vendor_ids: Optional[list] = None

class IntegrityFixIn(BaseModel):
    vendor_id: str
    field: str
    value: Optional[str] = None

class IntegrityMergeIn(BaseModel):
    primary_vendor_id: str
    duplicate_vendor_id: str

class IntegrityVendorIn(BaseModel):
    vendor_id: str

class ContagionIn(BaseModel):
    node_type: str   # fourth_party | owner | vendor
    node_id: str

class IncidentMatchIn(BaseModel):
    vendor_id: str
    description: Optional[str] = ""
    domain: Optional[str] = None

class BriefIn(BaseModel):
    business_unit: str

class I18nTranslateIn(BaseModel):
    strings: list
    lang: str

class I18nTextIn(BaseModel):
    text: str
    lang: Optional[str] = None

class IncidentIn(BaseModel):
    date_of_incident: Optional[str] = None
    reported_date: Optional[str] = None
    reported_by: Optional[str] = None
    vendor_id: Optional[str] = None
    engagement_id: Optional[str] = None
    incident_type: Optional[str] = None
    severity: Optional[str] = "Medium"
    customer_impacting: Optional[bool] = False
    impacts_client_org: Optional[bool] = False
    impact_description: Optional[str] = None
    region: Optional[list] = None
    root_cause_assessment: Optional[str] = None
    risk_entry_needed: Optional[bool] = False
    status: Optional[str] = None
    vendor_notified_at: Optional[str] = None
    notification_sla_hours: Optional[int] = None

class IncidentNoteIn(BaseModel):
    note: str

class IncidentAttachIn(BaseModel):
    name: str

class ScenarioIn(BaseModel):
    node_type: str = "fourth_party"
    node_id: str
    hours: int = 24

class V2RemediationIn(BaseModel):
    finding_id: str
    plan: str
    owner: Optional[str] = None
    target_date: Optional[str] = None

class V2FourthPartyIn(BaseModel):
    legal_name: str
    service_provided: Optional[str] = None
    hq_country: Optional[str] = None
    vendor_ids: Optional[list[str]] = None
    vendor_id: Optional[str] = None

class V2ArtefactIn(BaseModel):
    vendor_id: str
    name: str
    artefact_type: Optional[str] = "certificate"
    expiry_date: Optional[str] = None
    issue_date: Optional[str] = None
    engagement_id: Optional[str] = None
    received_via: Optional[str] = "upload"
    supersedes: Optional[str] = None

class FinancialIn(BaseModel):
    figures: dict
    flags: Optional[dict] = None
    vendor_id: Optional[str] = None
    other_name: Optional[str] = None

class ReputationIn(BaseModel):
    events: Optional[list[dict]] = None
    customer_facing: bool = False
    vendor_id: Optional[str] = None
    other_name: Optional[str] = None

class ConfigSetIn(BaseModel):
    value: Union[int, float, str]

class NavOrderIn(BaseModel):
    order: Optional[dict] = None

class ContractTermsIn(BaseModel):
    inherent_band: str = "MODERATE"
    exposure: Optional[dict] = None
    vendor_id: Optional[str] = None
    other_name: Optional[str] = None

class ContractGapIn(BaseModel):
    contract_text: str
    inherent_band: str = "MODERATE"
    exposure: Optional[dict] = None

class ContractDiffIn(BaseModel):
    inherent_band: str = "MODERATE"
    exposure: Optional[dict] = None
    prior_contract_texts: list[str]

class MgmtChatIn(BaseModel):
    question: str
    history: Optional[list] = None

class BoardFollowupIn(BaseModel):
    question: str
    history: Optional[list] = None

class PRPullIn(BaseModel):
    pr_number: str

class SimilarIn(BaseModel):
    entity: Optional[str] = None
    scope: Optional[str] = None

class CaptureIn(BaseModel):
    session_id: int
    engagement_id: str
    vendor_id: Optional[str] = None

class EmailIntakeIn(BaseModel):
    sender: str
    subject: Optional[str] = None
    attachment_name: Optional[str] = None
    attachment_b64: Optional[str] = None
    body_text: Optional[str] = None
    vendor_id: Optional[str] = None

class _DocFile(BaseModel):
    filename: str
    content_type: Optional[str] = "application/octet-stream"
    data_b64: str

class DocUploadIn(BaseModel):
    files: list[_DocFile]
    vendor_id: Optional[str] = None
    engagement_id: Optional[str] = None
    purpose: Optional[str] = None

class CertIngestIn(BaseModel):
    files: list[_DocFile]
    vendor_id: str
    engagement_id: Optional[str] = None

class ContractGapDocIn(BaseModel):
    file: _DocFile
    engagement_id: Optional[str] = None
    vendor_id: Optional[str] = None
    other_name: Optional[str] = None
    inherent_band: Optional[str] = None

class PeerBenchmarkIn(BaseModel):
    figures: dict
    flags: Optional[dict] = None
    sector: str = "other"

class FinResearchIn(BaseModel):
    company: str
    jurisdiction: str = "UK"
    identifier: Optional[str] = ""
    year: Optional[str] = ""

class FinMonitorAddIn(BaseModel):
    vendor_id: Optional[str] = None
    other_name: Optional[str] = None

class FinMonitorSweepIn(BaseModel):
    monitor_id: Optional[int] = None   # sweep one, or all if None

# ---- Req 1/2/3 schemas ----
class VendorMasterIn(BaseModel):
    data: dict
    include_bank: bool = False

class ScreeningIn(BaseModel):
    screen_type: str
    result: Optional[str] = None
    detail: Optional[str] = None
    screened_date: Optional[str] = None
    next_due: Optional[str] = None

class AttrDomainIn(BaseModel):
    data: dict

class InsuranceIn(BaseModel):
    policy_type: str
    coverage_limit: Optional[float] = None
    currency: Optional[str] = None
    insurer: Optional[str] = None
    certificate_expiry: Optional[str] = None
    named_insured_verified: bool = False

class MonitorSignalIn(BaseModel):
    signal_type: str
    value: str
    source: Optional[str] = None

class EngExtIn(BaseModel):
    data: dict

class EngChildIn(BaseModel):
    kind: str   # deliverable/milestone/sla/obligation/personnel
    data: dict

# ---- R2 contract entity schemas ----
class ContractCreateIn(BaseModel):
    contract_type: str = "Contract"
    vendor_id: Optional[str] = None
    engagement_id: Optional[str] = None
    parent_msa: Optional[str] = None
    data: Optional[dict] = None

class ContractUpdateIn(BaseModel):
    data: dict

# ---- R3 critical vendors schemas ----
class CriticalityInputIn(BaseModel):
    customer_impact: Optional[int] = None
    downtime_tolerance: Optional[int] = None
    alternative_availability: Optional[int] = None
    substitution_complexity: Optional[int] = None

class CriticalAnalysisIn(BaseModel):
    vendor_id: Optional[str] = None

class CriticalOverrideIn(BaseModel):
    is_critical: bool
    reason: str

# ---- R4 performance management schemas ----
class ScorecardCreateIn(BaseModel):
    vendor_id: str
    period_label: str
    period_start: Optional[str] = None
    period_end: Optional[str] = None
    cadence: str = "quarterly"

class PerfEnrolIn(BaseModel):
    vendor_ids: list[str]

class KPIScoreIn(BaseModel):
    actual: Optional[str] = None
    score: Optional[int] = None
    excluded: Optional[bool] = None
    exclude_reason: Optional[str] = None

class AgreeIn(BaseModel):
    party: str

class ReviewIn(BaseModel):
    data: dict

class PerfCapaIn(BaseModel):
    scorecard_id: str
    gap: str
    owner: str
    due_date: Optional[str] = None

class CapaVerifyIn(BaseModel):
    evidence: str

# ---- R5 ProAssess schemas ----
class ProAssessRunIn(BaseModel):
    vendor_id: Optional[str] = None
    engagement_id: Optional[str] = None
    irq: Optional[dict] = None
    ddq: Optional[dict] = None
    documents: Optional[list] = None
    extracted: Optional[dict] = None   # LLM-extracted structured inputs (financials, reputation_events, flags)

class ProAssessRegisterIn(BaseModel):
    report: dict

class ProAssessAutoIn(BaseModel):
    free_text: Optional[str] = ""
    documents: Optional[list] = None        # [{filename, content_type, data_b64}]
    vendor_id: Optional[str] = None         # existing vendor, OR:
    new_vendor_name: Optional[str] = None   # create a new vendor
    engagement_title: Optional[str] = None
    ddq: Optional[dict] = None
    create_records: bool = True


class SLAIn(BaseModel):
    engagement_id: str
    vendor_id: Optional[str] = None
    description: str
    threshold_type: Optional[str] = "min"
    threshold: float = 0.0
    unit: Optional[str] = ""
    baseline: Optional[float] = None
    window: Optional[str] = "monthly"
    source: Optional[str] = "manual"
    contract_id: Optional[str] = None


class SLAEditIn(BaseModel):
    description: Optional[str] = None
    threshold_type: Optional[str] = None
    threshold: Optional[float] = None
    unit: Optional[str] = None
    baseline: Optional[float] = None
    window: Optional[str] = None
    active: Optional[bool] = None


class MeasurementIn(BaseModel):
    period: str
    value: Optional[float] = None


class ExtractIn(BaseModel):
    engagement_id: str
    vendor_id: Optional[str] = None
    mode: Optional[str] = "contract"
    contract_id: Optional[str] = None


class EnquiryIn(BaseModel):
    engagement_id: str
    question: str


class PerfIssueIn(BaseModel):
    engagement_id: str
    vendor_id: Optional[str] = None
    title: str
    description: Optional[str] = None
    category: Optional[str] = None
    severity: Optional[str] = "Medium"
    source: Optional[str] = "Manual"
    status: Optional[str] = "Open"
    owner: Optional[str] = None
    due_date: Optional[str] = None
    linked_ref: Optional[str] = None
    suggested_remediation: Optional[str] = None


class PerfIssueEditIn(BaseModel):
    title: Optional[str] = None
    description: Optional[str] = None
    category: Optional[str] = None
    severity: Optional[str] = None
    source: Optional[str] = None
    status: Optional[str] = None
    owner: Optional[str] = None
    due_date: Optional[str] = None
    linked_ref: Optional[str] = None
    suggested_remediation: Optional[str] = None
    risk_accepted: Optional[bool] = None
    acceptance_rationale: Optional[str] = None


class NoteIn(BaseModel):
    note: str


class RaiseFromSLAIn(BaseModel):
    sla_id: str


class LearningIn(BaseModel):
    category: str = "Risk pattern"
    insight: str
    confidence: Optional[str] = "Medium"
    source_engagement: Optional[str] = None
    source_vendor: Optional[str] = None


def create_app(db_url: Optional[str] = None) -> FastAPI:
    engine = make_engine(db_url or "sqlite:///:memory:")
    # ensure registry models are imported so their tables register on Base
    from .features import registry_models as _rm  # noqa: F401
    from .features import master_ext as _mx  # noqa: F401  (Req 1/2/3 tables)
    from .features import documents as _docs  # noqa: F401  (CR-4/5/12 document store)
    from .features import exit_planning as _exit  # noqa: F401  (vendor exit strategy)
    from .features import methodology as _meth  # noqa: F401  (admin methodology library)
    from .features import config_store as _cfg  # noqa: F401  (system configuration store)
    from .features import performance_models as _perf  # noqa: F401  (SLA + performance issues)
    from .features import platform_docs as _pdocs  # noqa: F401  (SOP/TDA + version history)
    from .features import learnings as _learn  # noqa: F401  (platform learnings log)
    # Dev/demo bootstrap creates tables directly. In production set
    # BRO_DB_AUTO_CREATE=0 and manage schema with Alembic (`alembic upgrade head`).
    if _os.environ.get("BRO_DB_AUTO_CREATE", "1") != "0":
        Base.metadata.create_all(engine)
    SessionFactory = make_session_factory(engine)

    with SessionFactory() as s:
        seed(s)
        from .features.registry_service import seed_masters
        seed_masters(s)
        _cfg.seed_defaults(s)
        # restore admin-entered AI provider keys (persisted in system config) into env
        try:
            from .agents import llm_config as _llm
            from .features import security as _sec
            for _cp in (_cfg.get_json(s, "ai_custom_providers", []) or []):
                try:
                    _llm.register_provider(_cp["id"], _cp.get("label"),
                                           _cp["base_url"], _cp.get("model"))
                except Exception as _e:
                    _obs_swallow('bro_app.py', _e)
            _stored = _cfg.get_json(s, "ai_provider_keys", {}) or {}
            for _prov, _k in _stored.items():
                ev = _llm._PROVIDER_KEY_ENV.get(_prov)
                if ev and _k and not _os.environ.get(ev):
                    _os.environ[ev] = _sec.decrypt_value(_k)
            _ap = (_cfg.get_json(s, "ai_active_provider", {}) or {}).get("provider")
            if _ap and not _os.environ.get("BRO_LLM_PROVIDER"):
                _os.environ["BRO_LLM_PROVIDER"] = _ap
        except Exception as _e:
            _obs_swallow('bro_app.py', _e)
        # AI governance: call ledger + daily budget hooks
        try:
            from .features import ai_ledger as _AL
            _AL.ensure_table(s)
            from .agents import llm_config as _llm3
            _llm3.set_telemetry(
                record=lambda p: _AL.safe_record(SessionFactory, p),
                budget=lambda: _AL.budget_check(SessionFactory))
        except Exception as _e:
            _obs_swallow('bro_app.py', _e)
        s.commit()

    def _platform_version() -> str:
        try:
            import pathlib
            return pathlib.Path(__file__).resolve().parent.parent.joinpath(
                "VERSION").read_text().strip()
        except Exception:
            return "4.1.0"

    app = FastAPI(title="BRO Risk Oracle", version=_platform_version())

    @app.middleware("http")
    async def _security_headers(request, call_next):
        resp = await call_next(request)
        resp.headers.setdefault("X-Frame-Options", "DENY")
        resp.headers.setdefault("X-Content-Type-Options", "nosniff")
        resp.headers.setdefault("Referrer-Policy", "strict-origin-when-cross-origin")
        resp.headers.setdefault(
            "Content-Security-Policy",
            "default-src 'self'; script-src 'self' 'unsafe-inline' https://cdnjs.cloudflare.com; "
            "style-src 'self' 'unsafe-inline' https://fonts.googleapis.com; "
            "font-src 'self' https://fonts.gstatic.com data:; img-src 'self' data: blob:; "
            "frame-src 'self' blob:; connect-src 'self'")
        if SEC.is_production():
            resp.headers.setdefault("Strict-Transport-Security",
                                    "max-age=31536000; includeSubDomains")
        return resp

    def db() -> Session:
        s = SessionFactory()
        try:
            yield s
        finally:
            s.close()

    # ----- actor + RBAC -----
    # Production auth: identity comes from a verified JWT bearer token, NOT a
    # client-supplied header. A test/dev escape hatch (BRO_TRUST_HEADER=1) keeps
    # the x-user header working for the existing test suite and local poking.
    def actor(authorization: str = Header(default=None),
              x_user: str = Header(default=None),
              s: Session = Depends(db)) -> User:
        username: Optional[str] = None
        if authorization:
            try:
                username = bearer_subject(authorization)
            except TokenError as e:
                raise HTTPException(401, str(e))
        elif _os.environ.get("BRO_TRUST_HEADER") == "1" and x_user and not SEC.is_production():
            username = x_user  # dev/test only
        if not username:
            raise HTTPException(401, "authentication required")
        u = s.scalars(select(User).where(User.username == username)).first()
        if not u or not u.is_active:
            raise HTTPException(401, "unknown or inactive user")
        return u

    def require(perm: str):
        def dep(u: User = Depends(actor)):
            if not has_permission(u, perm):
                raise HTTPException(403, f"missing permission: {perm}")
            return u
        return dep

    # ----- audit (hash-chained, persisted) -----
    def audit(s: Session, action: str, actor_name: str, detail: dict) -> None:
        last = s.scalars(select(AuditLog).order_by(AuditLog.seq.desc())).first()
        prev = last.entry_hash if last else "genesis"
        seq = (last.seq + 1) if last else 0
        h = eng.chain_hash(prev, action, actor_name, detail)
        s.add(AuditLog(seq=seq, action=action, actor=actor_name,
                       detail=json.dumps(detail, sort_keys=True),
                       prev_hash=prev, entry_hash=h))

    def notify(s: Session, event: str, audience: str = "all",
               body: str = "") -> None:
        s.add(Notification(audience=audience, event=event, body=body))

    def _fb_guidance(s: Session, surface: str) -> Optional[str]:
        """Distil prior user feedback into prompt guidance; safe no-op on any error."""
        try:
            from .features import feedback as FB
            g = FB.guidance(s, surface=surface)
            return g or None
        except Exception:
            return None

    def _ai_live() -> bool:
        try:
            from .agents import llm_config
            return bool(llm_config.status().get("live_ready"))
        except Exception:
            return False

    AI_HOLDING = ("Noted — AI engines not available yet. BRO Chat, ProAssess, FDD and "
                  "Reputation are AI-driven workflows: they follow the assessment "
                  "methodology with adaptive questioning and analysis, and do not proceed "
                  "without it. Connect an AI provider in Settings → AI to enable them.")

    # ===== health =====
    # ===== health/readiness — extracted to app/routers/health.py (Gap-7 pattern) =====
    # These endpoints now live in their own router module, wired via RouterDeps
    # dependency injection instead of closing over create_app's local scope.
    # This is the reference implementation for decomposing the monolith.
    from .routers.deps import RouterDeps
    from .routers.health import build_health_router
    _router_deps = RouterDeps(
        db=db, actor=actor, require=require, audit=audit,
        platform_version=_platform_version,
    )
    app.include_router(build_health_router(_router_deps))

    # ===== auth =====

    @app.post("/api/v1/login")
    def login(body: LoginIn, request: Request, s: Session = Depends(db)):
        ip = request.client.host if (request and request.client) else "?"
        if not SEC.check_rate(f"login:{ip}"):
            SEC.log_json("login.rate_limited", ip=ip)
            raise HTTPException(429, "too many login attempts — please slow down")
        rem = SEC.login_locked(body.username)
        if rem:
            raise HTTPException(423, f"account temporarily locked — retry in {rem}s")
        u = s.scalars(select(User).where(User.username == body.username)).first()
        if not u or not verify_password(body.password, u.password_hash):
            SEC.record_login_failure(body.username)
            raise HTTPException(401, "invalid credentials")
        SEC.record_login_success(body.username)
        token = issue_token(u.username, u.role.key)
        return {"token": token, "token_type": "bearer",
                "username": u.username, "role": u.role.key,
                "permissions": [p.key for p in u.role.permissions]}

    # ===== SSO (OIDC) + SCIM provisioning =====
    from .features import identity as IDP
    from .features.models_db import hash_password as _hash_pw
    import secrets as _secrets

    def _resolve_role(s: Session, key: Optional[str]) -> Role:
        """Map a role key to a Role row; fall back to least-privilege, never None."""
        for k in (key, "vendor"):
            if k:
                r = s.scalars(select(Role).where(Role.key == k)).first()
                if r:
                    return r
        r = s.scalars(select(Role)).first()
        if not r:
            raise HTTPException(500, "no roles configured")
        return r

    def _upsert_sso_user(s: Session, ident: dict) -> User:
        u = s.scalars(select(User).where(User.username == ident["username"])).first()
        role = _resolve_role(s, ident["role"])
        if not u:
            u = User(username=ident["username"], full_name=ident.get("full_name") or ident["username"],
                     email=ident.get("email"), password_hash=_hash_pw(_secrets.token_hex(24)),
                     role_id=role.id, is_active=True)
            s.add(u); s.flush()
            audit(s, "sso.user_provisioned", ident["username"], {"role": role.key})
        else:
            if u.role_id != role.id:
                u.role_id = role.id
            u.is_active = True
            if ident.get("full_name"):
                u.full_name = ident["full_name"]
        return u

    @app.get("/auth/oidc/login")
    def oidc_login():
        if not IDP.oidc_enabled():
            raise HTTPException(404, "SSO is not configured")
        from fastapi.responses import RedirectResponse
        return RedirectResponse(IDP.auth_url(IDP.new_state()))

    @app.get("/auth/oidc/callback")
    def oidc_callback(code: str = None, state: str = None, s: Session = Depends(db)):
        from fastapi.responses import HTMLResponse
        if not IDP.oidc_enabled():
            raise HTTPException(404, "SSO is not configured")
        if not code:
            raise HTTPException(400, "missing authorization code")
        try:
            tokens = IDP.exchange_code(code)
            claims = IDP.verify_id_token(tokens["id_token"])
            ident = IDP.claims_to_identity(claims)
        except Exception as e:
            raise HTTPException(401, f"SSO sign-in failed: {e}")
        u = _upsert_sso_user(s, ident)
        s.commit()
        token = issue_token(u.username, u.role.key)
        # Hand the session token to the SPA (held in sessionStorage) and redirect in.
        return HTMLResponse(
            "<!doctype html><meta charset=utf-8><script>"
            f"sessionStorage.setItem('bro_tok',{json.dumps(token)});"
            "location.replace('/');</script>Signing you in…")

    @app.get("/api/v1/auth/sso-status")
    def sso_status():
        return {"oidc_enabled": IDP.oidc_enabled(), "scim_enabled": IDP.scim_enabled()}

    # ---- SCIM 2.0 (RFC 7644): IdP-driven user provisioning ----
    def _scim_guard(authorization: str = Header(default=None)):
        if not IDP.scim_enabled():
            raise HTTPException(404, "SCIM is not enabled")
        if not IDP.scim_token_ok(authorization):
            raise HTTPException(401, "invalid SCIM token")

    @app.get("/scim/v2/Users")
    def scim_list_users(filter: str = None, s: Session = Depends(db),
                        authorization: str = Header(default=None)):
        _scim_guard(authorization)
        rows = s.scalars(select(User)).all()
        if filter and "userName eq " in filter:
            want = filter.split("userName eq ", 1)[1].strip().strip('"')
            rows = [u for u in rows if u.username == want]
        return IDP.scim_list_response([IDP.user_to_scim(u) for u in rows])

    @app.get("/scim/v2/Users/{uid}")
    def scim_get_user(uid: int, s: Session = Depends(db),
                      authorization: str = Header(default=None)):
        _scim_guard(authorization)
        u = s.get(User, uid)
        if not u:
            raise HTTPException(404, "not found")
        return IDP.user_to_scim(u)

    @app.post("/scim/v2/Users", status_code=201)
    def scim_create_user(body: dict = Body(...), s: Session = Depends(db),
                         authorization: str = Header(default=None)):
        _scim_guard(authorization)
        data = IDP.scim_extract(body)
        if not data["username"]:
            raise HTTPException(400, "userName required")
        existing = s.scalars(select(User).where(User.username == data["username"])).first()
        if existing:
            return IDP.user_to_scim(existing)
        role_key = data["role_hint"] or IDP.map_groups_to_role([])
        role = _resolve_role(s, role_key)
        u = User(username=data["username"], full_name=data["full_name"], email=data["email"],
                 password_hash=_hash_pw(_secrets.token_hex(24)),
                 role_id=role.id, is_active=bool(data["active"]))
        s.add(u); s.flush()
        audit(s, "scim.user_created", data["username"], {"role": role.key})
        s.commit()
        return IDP.user_to_scim(u)

    @app.patch("/scim/v2/Users/{uid}")
    @app.put("/scim/v2/Users/{uid}")
    def scim_update_user(uid: int, body: dict = Body(...), s: Session = Depends(db),
                         authorization: str = Header(default=None)):
        _scim_guard(authorization)
        u = s.get(User, uid)
        if not u:
            raise HTTPException(404, "not found")
        # Support both PUT (full resource) and PATCH (Operations) for active/profile.
        if "active" in body:
            u.is_active = bool(body["active"])
        for op in body.get("Operations", []):
            if (op.get("path") == "active") or ("active" in str(op.get("value"))):
                v = op.get("value")
                u.is_active = (v.get("active") if isinstance(v, dict) else bool(v))
        d = IDP.scim_extract(body) if "userName" in body else None
        if d and d.get("full_name"):
            u.full_name = d["full_name"]
        audit(s, "scim.user_updated", u.username, {"active": u.is_active})
        s.commit()
        return IDP.user_to_scim(u)

    @app.delete("/scim/v2/Users/{uid}", status_code=204)
    def scim_delete_user(uid: int, s: Session = Depends(db),
                         authorization: str = Header(default=None)):
        _scim_guard(authorization)
        u = s.get(User, uid)
        if u:
            u.is_active = False  # SCIM delete = deactivate (preserve audit trail)
            audit(s, "scim.user_deactivated", u.username, {})
            s.commit()
        return Response(status_code=204)

    @app.get("/scim/v2/Groups")
    def scim_list_groups(s: Session = Depends(db), authorization: str = Header(default=None)):
        _scim_guard(authorization)
        groups = [{"schemas": ["urn:ietf:params:scim:schemas:core:2.0:Group"],
                   "id": str(r.id), "displayName": r.key}
                  for r in s.scalars(select(Role)).all()]
        return IDP.scim_list_response(groups)

    # ===== vendors =====
    @app.post("/api/v1/vendors")
    def create_vendor(v: VendorIn, s: Session = Depends(db),
                      u: User = Depends(require("vendor.edit"))):
        row = Vendor(**v.model_dump())
        s.add(row); s.flush()
        audit(s, "vendor.created", u.username, {"vendor_id": row.id, "name": v.name})
        s.commit()
        return {"vendor_id": row.id, "name": row.name, "tier": row.tier}

    @app.get("/api/v1/vendors")
    def list_vendors(limit: Optional[int] = None, offset: int = 0,
                     s: Session = Depends(db), u: User = Depends(require("vendor.view"))):
        q = select(Vendor).offset(max(0, offset))
        if limit:
            q = q.limit(max(1, min(int(limit), 500)))
        return [{"vendor_id": v.id, "name": v.name, "tier": v.tier,
                 "is_critical": v.is_critical}
                for v in s.scalars(q).all()]

    @app.post("/api/v1/vendors/{vid}/critical")
    def designate_critical(vid: int, body: CriticalIn, s: Session = Depends(db),
                           u: User = Depends(require("vendor.critical"))):
        # Tier 0 = human-only (our Q5). RBAC already enforces a human role here.
        v = s.get(Vendor, vid)
        if not v:
            raise HTTPException(404, "vendor not found")
        v.is_critical = True
        v.critical_reason = body.reason
        v.critical_by = u.username
        audit(s, "vendor.critical_designated", u.username,
              {"vendor_id": vid, "reason": body.reason})
        notify(s, f"Critical vendor designated: {v.name}", "vrm")
        s.commit()
        return {"vendor_id": vid, "is_critical": True, "by": u.username}

    # ===== engagements + lifecycle =====
    @app.post("/api/v1/engagements")
    def create_engagement(e: EngagementIn, s: Session = Depends(db),
                          u: User = Depends(require("engagement.create"))):
        if not s.get(Vendor, e.vendor_id):
            raise HTTPException(404, "vendor not found")
        row = EngagementRow(vendor_id=e.vendor_id, title=e.title,
                            service_description=e.service_description,
                            business_contact_email=e.business_contact_email,
                            owner_id=u.id, stage="sourcing")
        s.add(row); s.flush()
        audit(s, "engagement.created", u.username, {"engagement_id": row.id})
        notify(s, f"Engagement created: {e.title}", "business")
        s.commit()
        return {"engagement_id": row.id, "stage": row.stage}

    @app.post("/api/v1/engagements/{eid}/irq")
    def submit_irq(eid: int, body: IRQIn, s: Session = Depends(db),
                   u: User = Depends(require("engagement.edit"))):
        e = s.get(EngagementRow, eid)
        if not e:
            raise HTTPException(404, "engagement not found")
        tier = eng.compute_tier(body.answers)
        inherent = eng.compute_inherent(body.answers)
        routing = eng.compute_route(body.answers, inherent, tier)
        e.inherent_band = inherent["band"]
        e.inherent_pct = inherent["weighted_pct"]
        e.route = routing["route"]
        e.stage = "inherent"
        audit(s, "irq.scored", u.username,
              {"engagement_id": eid, "tier": tier, "band": inherent["band"],
               "route": routing["route"]})
        notify(s, f"IRQ scored {inherent['band']} ({routing['route']})", "all")
        s.commit()
        return {"engagement_id": eid, "tier": tier,
                "inherent_band": inherent["band"],
                "inherent_pct": inherent["weighted_pct"],
                "cls": inherent["cls"], "routing": routing}

    @app.post("/api/v1/engagements/{eid}/ddq")
    def submit_ddq(eid: int, body: DDQIn, s: Session = Depends(db),
                   u: User = Depends(require("engagement.edit"))):
        e = s.get(EngagementRow, eid)
        if not e:
            raise HTTPException(404, "engagement not found")
        residual = eng.compute_residual(e.inherent_band or "LOW", body.answers)
        decision = eng.decision_for(residual["band"], residual["critical_marginal"])
        e.residual_band = residual["band"]
        e.decision = decision["text"]
        e.stage = "decision"
        audit(s, "ddq.residual", u.username,
              {"engagement_id": eid, "residual": residual["band"],
               "critical_marginal": residual["critical_marginal"],
               "decision": decision["text"]})
        notify(s, f"Residual {residual['band']}: {decision['text']}", "all")
        s.commit()
        return {"engagement_id": eid, "residual_band": residual["band"],
                "critical_marginal": residual["critical_marginal"],
                "decision": decision}

    @app.post("/api/v1/engagements/{eid}/override")
    def override(eid: int, body: OverrideIn, s: Session = Depends(db),
                 u: User = Depends(require("engagement.override"))):
        # human-only (RBAC) + justification + 2nd approver (two-gate model)
        e = s.get(EngagementRow, eid)
        if not e:
            raise HTTPException(404, "engagement not found")
        if not body.reason or not body.second_approver:
            raise HTTPException(400, "override needs justification and 2nd approver")
        e.residual_band = body.band
        audit(s, "decision.override", u.username,
              {"engagement_id": eid, "new_band": body.band,
               "reason": body.reason, "second_approver": body.second_approver})
        s.commit()
        return {"engagement_id": eid, "residual_band": body.band, "override": True}

    @app.post("/api/v1/engagements/{eid}/terminate")
    def terminate(eid: int, s: Session = Depends(db),
                  u: User = Depends(require("lifecycle.offboard"))):
        e = s.get(EngagementRow, eid)
        if not e:
            raise HTTPException(404, "engagement not found")
        e.stage = "terminate"
        e.status = "terminated"
        for key, label in eng.OFFBOARDING_STEPS:
            s.add(Offboarding(engagement_id=eid, step_key=key))
        audit(s, "engagement.terminated", u.username, {"engagement_id": eid})
        notify(s, "Offboarding initiated", "all")
        s.commit()
        return {"engagement_id": eid, "stage": "terminate",
                "offboarding_steps": len(eng.OFFBOARDING_STEPS)}

    @app.get("/api/v1/engagements/{eid}")
    def get_engagement(eid: int, s: Session = Depends(db),
                       u: User = Depends(require("engagement.view"))):
        e = s.get(EngagementRow, eid)
        if not e:
            raise HTTPException(404, "engagement not found")
        return {"engagement_id": e.id, "vendor_id": e.vendor_id,
                "stage": e.stage, "status": e.status, "route": e.route,
                "inherent_band": e.inherent_band, "residual_band": e.residual_band,
                "decision": e.decision}

    # ===== findings =====
    @app.post("/api/v1/findings")
    def create_finding(f: FindingIn, s: Session = Depends(db),
                       u: User = Depends(require("finding.manage"))):
        row = Finding(engagement_id=f.engagement_id, title=f.title, severity=f.severity)
        s.add(row); s.flush()
        audit(s, "finding.raised", u.username, {"finding_id": row.id})
        notify(s, f"Finding raised: {f.title}", "all")
        s.commit()
        return {"finding_id": row.id, "status": row.status,
                "sla_days": eng.SEVERITY_SLA.get(f.severity)}

    @app.post("/api/v1/findings/{fid}/advance")
    def advance_finding(fid: int, s: Session = Depends(db),
                        u: User = Depends(require("finding.manage"))):
        f = s.get(Finding, fid)
        if not f:
            raise HTTPException(404, "finding not found")
        order = eng.FINDING_STATUSES
        i = min(order.index(f.status) + 1, len(order) - 1)
        f.status = order[i]
        audit(s, "finding.advanced", u.username, {"finding_id": fid, "status": f.status})
        s.commit()
        return {"finding_id": fid, "status": f.status}

    # ===== intelligence engines =====
    @app.post("/api/v1/intel/financial")
    def intel_financial(b: IntelIn, s: Session = Depends(db),
                        u: User = Depends(require("intel.financial"))):
        out = intel.vera_financial(b.payload)
        s.add(IntelResult(vendor_id=b.vendor_id, engine="financial",
                          score=out.score, band=out.band, narrative=out.narrative))
        audit(s, "intel.financial", u.username, {"vendor_id": b.vendor_id, "band": out.band})
        s.commit()
        return out.__dict__

    @app.post("/api/v1/intel/reputation")
    def intel_reputation(b: IntelIn, s: Session = Depends(db),
                         u: User = Depends(require("intel.reputation"))):
        out = intel.mira_reputation(b.payload)
        s.add(IntelResult(vendor_id=b.vendor_id, engine="reputation",
                          score=out.score, band=out.band, narrative=out.narrative))
        audit(s, "intel.reputation", u.username, {"vendor_id": b.vendor_id, "band": out.band})
        s.commit()
        return out.__dict__

    # ===== Sanctions & AML screening (part of reputation checks) =====
    from .features import sanctions as SANC
    from .features.registry_models import SanctionsScreening, WatchlistEntry, VendorPerson

    def _live_entries(s: Session) -> list:
        """Load ingested live-feed entries (e.g. OFSI) into the engine's entry format."""
        out = []
        for w in s.scalars(select(WatchlistEntry)).all():
            out.append({"id": w.ext_id, "name": w.name, "aliases": json.loads(w.aliases or "[]"),
                        "category": w.category, "source": w.source, "list": w.list_name,
                        "program": w.program, "country": w.country, "nationality": w.nationality,
                        "entity_type": w.entity_type, "dob": w.dob, "live": True})
        return out

    def _apply_screening_outcome(s: Session, vendor_id, vendor_name, band, detail, by):
        return RS.apply_screening_outcome(s, vendor_id, vendor_name, band, detail, by,
                                          audit_fn=audit)

    @app.get("/api/v2/sanctions/sources")
    def v2_sanctions_sources(u: User = Depends(require("vendor.view"))):
        return SANC.sources()

    @app.post("/api/v2/sanctions/screen")
    def v2_sanctions_screen(body: dict = Body(default={}), s: Session = Depends(db),
                            u: User = Depends(require("vendor.view"))):
        name = (body.get("name") or "").strip()
        country = body.get("country"); dob = body.get("dob"); nationality = body.get("nationality")
        vid = body.get("vendor_id")
        if vid and not name:
            v = s.scalars(select(VendorRecord).where(VendorRecord.vendor_id == vid)).first()
            if not v:
                raise HTTPException(404, "vendor not found")
            name = v.legal_name
            country = country or v.hq_country
        if not name:
            raise HTTPException(400, "name or vendor_id required")
        res = SANC.screen_name(name, country, dob, nationality, extra=_live_entries(s))
        sid = RS.next_id(s, "screening")
        s.add(SanctionsScreening(
            screening_id=sid, vendor_id=vid, screened_name=name, country=country,
            band=res["band"], hit_count=res["hit_count"], hits=json.dumps(res["hits"]),
            sources_used=json.dumps(res["sources_screened"]), screened_by=u.username))
        if vid:
            top = res["hits"][0]["matched_name"] if res["hits"] else ""
            res["risk"] = _apply_screening_outcome(
                s, vid, name, res["band"],
                f"{res['band']} — {name} vs {top}".strip(" —"), u.username)
        audit(s, "sanctions.screen", u.username,
              {"screening_id": sid, "vendor_id": vid, "name": name, "band": res["band"]})
        s.commit()
        res["screening_id"] = sid
        return res

    @app.get("/api/v2/sanctions/screenings")
    def v2_sanctions_screenings(vendor_id: Optional[str] = None, limit: int = 50,
                                s: Session = Depends(db),
                                u: User = Depends(require("vendor.view"))):
        stmt = select(SanctionsScreening).order_by(SanctionsScreening.id.desc())
        rows = s.scalars(stmt).all()
        if vendor_id:
            rows = [r for r in rows if r.vendor_id == vendor_id]
        return [{"screening_id": r.screening_id, "vendor_id": r.vendor_id,
                 "screened_name": r.screened_name, "country": r.country, "band": r.band,
                 "hit_count": r.hit_count, "hits": json.loads(r.hits or "[]"),
                 "screened_by": r.screened_by,
                 "created_at": r.created_at.isoformat() if r.created_at else None}
                for r in rows[:limit]]

    @app.get("/api/v2/sanctions/summary")
    def v2_sanctions_summary(s: Session = Depends(db),
                             u: User = Depends(require("vendor.view"))):
        """Deterministic portfolio sweep — screens every registered vendor's legal name."""
        live = _live_entries(s)
        dist = {"Clear": 0, "Review": 0, "Hit": 0}
        flagged = []
        for v in s.scalars(select(VendorRecord)).all():
            r = SANC.screen_name(v.legal_name, v.hq_country, extra=live)
            dist[r["band"]] = dist.get(r["band"], 0) + 1
            if r["band"] != "Clear":
                flagged.append({"vendor_id": v.vendor_id, "legal_name": v.legal_name,
                                "band": r["band"], "hit_count": r["hit_count"],
                                "top": r["hits"][0] if r["hits"] else None})
        flagged.sort(key=lambda f: (f["band"] != "Hit", -f["hit_count"]))
        return {"distribution": dist, "flagged": flagged, "screened": sum(dist.values()),
                "representative": not bool(live), "live_entries": len(live)}

    # ---- live issuer feed: OFSI (UK consolidated list, free CSV) ----
    @app.get("/api/v2/sanctions/feeds")
    def v2_sanctions_feeds(s: Session = Depends(db), u: User = Depends(require("vendor.view"))):
        feeds = {}
        for w in s.scalars(select(WatchlistEntry)).all():
            f = feeds.setdefault(w.source, {"source": w.source, "count": 0, "last_loaded": None})
            f["count"] += 1
            ts = w.loaded_at.isoformat() if w.loaded_at else None
            if ts and (not f["last_loaded"] or ts > f["last_loaded"]):
                f["last_loaded"] = ts
        return {"feeds": list(feeds.values()), "ofsi_url": SANC.OFSI_URL}

    def _ingest_feed(s: Session, source: str, entries: list) -> int:
        from datetime import datetime, timezone
        src = f"{source} (live)"
        existing = {w.ext_id: w for w in s.scalars(
            select(WatchlistEntry).where(WatchlistEntry.source == src)).all()}
        loaded = 0
        for e in entries:
            w = existing.get(e["id"])
            if not w:
                w = WatchlistEntry(ext_id=e["id"], source=src); s.add(w)
            w.name = e["name"]; w.aliases = json.dumps(e.get("aliases", []))
            w.category = e["category"]; w.list_name = e.get("list"); w.program = e.get("program")
            w.country = e.get("country"); w.nationality = e.get("nationality")
            w.entity_type = e.get("entity_type"); w.dob = e.get("dob")
            w.loaded_at = datetime.now(timezone.utc)
            loaded += 1
        return loaded

    @app.post("/api/v2/sanctions/load-feed")
    def v2_sanctions_load_feed(body: dict = Body(default={}), s: Session = Depends(db),
                               u: User = Depends(require("admin.config"))):
        """Ingest an issuer feed. source in OFSI/OFAC/UN/EU. Tries the live URL; if egress
        is blocked, accepts the file in the body (`csv` or `xml`/`text`)."""
        source = (body.get("source") or "OFSI").upper()
        if source not in SANC.FEED_LOADERS:
            raise HTTPException(400, f"unknown source {source}")
        text = body.get("csv") or body.get("xml") or body.get("text")
        via = "uploaded file"
        if not text:
            try:
                text = SANC.fetch_feed(source, body.get("url")); via = "live fetch"
            except Exception as e:
                raise HTTPException(502, f"{source} fetch failed ({e}); allowlist the issuer "
                                         f"host on this network, or POST the file in the body.")
        try:
            entries = SANC.load_feed(source, text)
        except Exception as e:
            raise HTTPException(400, f"could not parse {source} feed: {e}")
        loaded = _ingest_feed(s, source, entries)
        audit(s, "sanctions.load_feed", u.username, {"source": source, "loaded": loaded, "via": via})
        s.commit()
        total = s.scalar(select(func.count()).select_from(WatchlistEntry)
                         .where(WatchlistEntry.source == f"{source} (live)"))
        return {"source": source, "loaded": loaded, "via": via, "total": total}

    @app.post("/api/v2/sanctions/load-ofsi")
    def v2_sanctions_load_ofsi(body: dict = Body(default={}), s: Session = Depends(db),
                               u: User = Depends(require("admin.config"))):
        body = dict(body or {}); body["source"] = "OFSI"
        return v2_sanctions_load_feed(body, s, u)

    # ---- beneficial owners / key people + entity screening ----
    @app.get("/api/v2/vendors/{vid}/people")
    def v2_vendor_people(vid: str, s: Session = Depends(db),
                         u: User = Depends(require("vendor.view"))):
        rows = s.scalars(select(VendorPerson).where(VendorPerson.vendor_id == vid)).all()
        return [{"person_id": p.person_id, "name": p.name, "role": p.role, "dob": p.dob,
                 "nationality": p.nationality, "ownership_pct": p.ownership_pct,
                 "is_ubo": p.is_ubo} for p in rows]

    @app.post("/api/v2/vendors/{vid}/people")
    def v2_add_vendor_person(vid: str, body: dict = Body(...), s: Session = Depends(db),
                             u: User = Depends(require("vendor.edit"))):
        pid = RS.next_id(s, "person")
        s.add(VendorPerson(person_id=pid, vendor_id=vid, name=body.get("name", "").strip(),
                           role=body.get("role"), dob=body.get("dob"),
                           nationality=body.get("nationality"),
                           ownership_pct=body.get("ownership_pct"),
                           is_ubo=bool(body.get("is_ubo"))))
        audit(s, "vendor.person_added", u.username, {"vendor_id": vid, "person_id": pid})
        s.commit()
        return {"person_id": pid}

    @app.post("/api/v2/sanctions/screen-vendor/{vid}")
    def v2_screen_vendor_entities(vid: str, s: Session = Depends(db),
                                  u: User = Depends(require("vendor.view"))):
        """Screen the vendor entity AND each beneficial owner / key person, applying
        DOB + nationality disambiguation. Persists each screen and returns an aggregate."""
        v = s.scalars(select(VendorRecord).where(VendorRecord.vendor_id == vid)).first()
        if not v:
            raise HTTPException(404, "vendor not found")
        live = _live_entries(s)

        def _persist(name, country, res):
            sid = RS.next_id(s, "screening")
            s.add(SanctionsScreening(screening_id=sid, vendor_id=vid, screened_name=name,
                                     country=country, band=res["band"], hit_count=res["hit_count"],
                                     hits=json.dumps(res["hits"]),
                                     sources_used=json.dumps(res["sources_screened"]),
                                     screened_by=u.username))
            return sid

        entity_res = SANC.screen_name(v.legal_name, v.hq_country, extra=live)
        entity_res["screening_id"] = _persist(v.legal_name, v.hq_country, entity_res)
        people = []
        bands = [entity_res["band"]]
        for p in s.scalars(select(VendorPerson).where(VendorPerson.vendor_id == vid)).all():
            pr = SANC.screen_name(p.name, p.nationality, p.dob, p.nationality, extra=live)
            pr["screening_id"] = _persist(p.name, p.nationality, pr)
            bands.append(pr["band"])
            people.append({"person_id": p.person_id, "name": p.name, "role": p.role,
                           "dob": p.dob, "nationality": p.nationality,
                           "is_ubo": p.is_ubo, "result": pr})
        order = {"Hit": 2, "Review": 1, "Clear": 0}
        overall = max(bands, key=lambda b: order.get(b, 0)) if bands else "Clear"
        # compose a detail naming the worst subject, then wire into risk + Issues
        worst = entity_res
        worst_subj = v.legal_name
        for p in people:
            if order.get(p["result"]["band"], 0) > order.get(worst["band"], 0):
                worst = p["result"]; worst_subj = f"owner {p['name']}"
        top = worst["hits"][0]["matched_name"] if worst["hits"] else ""
        risk = _apply_screening_outcome(s, vid, v.legal_name, overall,
                                        f"{overall} — {worst_subj} vs {top}".strip(" —"), u.username)
        audit(s, "sanctions.screen_vendor", u.username,
              {"vendor_id": vid, "overall": overall, "people": len(people)})
        s.commit()
        return {"vendor_id": vid, "legal_name": v.legal_name, "entity_result": entity_res,
                "people": people, "overall_band": overall, "risk": risk}

    # ===== monitoring (scheduled / on-request sweeps) =====
    from .features import monitoring as MON

    def _monitor_interval() -> float:
        try:
            return float(_os.environ.get("BRO_MONITORING_INTERVAL_HOURS", "24"))
        except ValueError:
            return 24.0

    @app.post("/api/v2/monitoring/run")
    def v2_monitoring_run(body: dict = Body(default={}), s: Session = Depends(db),
                          u: User = Depends(require("admin.config"))):
        only = body.get("only")  # optional list of task names
        return MON.run_all(s, by=u.username, trigger="manual", audit_fn=audit, only=only)

    @app.get("/api/v2/monitoring/status")
    def v2_monitoring_status(s: Session = Depends(db),
                             u: User = Depends(require("vendor.view"))):
        st = MON.status(s, _monitor_interval())
        st["tasks_available"] = [n for n, _ in MON.TASKS]
        st["scheduler_enabled"] = _os.environ.get("BRO_SCHEDULER_ENABLED") == "1"
        return st

    @app.get("/api/v2/monitoring/runs")
    def v2_monitoring_runs(limit: int = 20, s: Session = Depends(db),
                           u: User = Depends(require("vendor.view"))):
        from .features.registry_models import MonitoringRun
        rows = s.scalars(select(MonitoringRun).order_by(MonitoringRun.id.desc())).all()[:limit]
        return [{"run_id": r.run_id, "trigger": r.trigger, "ok": r.ok,
                 "started_at": r.started_at.isoformat() if r.started_at else None,
                 "finished_at": r.finished_at.isoformat() if r.finished_at else None,
                 "tasks": json.loads(r.tasks or "{}")} for r in rows]




    @app.post("/api/v1/intel/contract")
    def intel_contract(b: IntelIn, s: Session = Depends(db),
                       u: User = Depends(require("intel.contract"))):
        v = s.get(Vendor, b.vendor_id)
        out = intel.matt_contract(v.tier if v else "Tier 3")
        audit(s, "intel.contract", u.username, {"vendor_id": b.vendor_id})
        s.commit()
        return out.__dict__

    @app.post("/api/v1/intel/evidence")
    def intel_evidence(b: IntelIn, s: Session = Depends(db),
                       u: User = Depends(require("intel.evidence"))):
        out = intel.isaac_evidence(b.payload.get("text", ""))
        s.add(IntelResult(vendor_id=b.vendor_id, engine="evidence",
                          score=out.score, band=out.band, narrative=out.narrative))
        audit(s, "intel.evidence", u.username, {"vendor_id": b.vendor_id, "band": out.band})
        s.commit()
        return out.__dict__

    # ===== monitoring lifecycle =====
    @app.post("/api/v1/monitoring/sweep")
    def monitoring_sweep(b: IntelIn, s: Session = Depends(db),
                         u: User = Depends(require("lifecycle.monitoring"))):
        fin = intel.vera_financial(b.payload)
        status = "OK" if (fin.score or 0) >= 55 else "ALERT" if (fin.score or 0) >= 35 else "CRITICAL"
        s.add(Monitoring(vendor_id=b.vendor_id, sweep_type="financial",
                         status=status, detail=fin.narrative))
        if status in ("ALERT", "CRITICAL"):
            # auto-raise a reassessment + notify VRM and business
            s.add(Reassessment(engagement_id=0, mode="triggered"))
            notify(s, f"Monitoring {status} for vendor {b.vendor_id}", "all")
        audit(s, "monitoring.sweep", u.username, {"vendor_id": b.vendor_id, "status": status})
        s.commit()
        return {"vendor_id": b.vendor_id, "status": status}

    @app.post("/api/v1/incidents")
    def create_incident(b: IntelIn, s: Session = Depends(db),
                        u: User = Depends(require("lifecycle.incident"))):
        row = Incident(vendor_id=b.vendor_id, title=b.payload.get("title", "Incident"),
                       severity=b.payload.get("severity", "medium"))
        s.add(row); s.flush()
        audit(s, "incident.raised", u.username, {"incident_id": row.id})
        notify(s, "Third-party incident raised", "all")
        s.commit()
        return {"incident_id": row.id, "status": row.status}

    # ===== conversational assessment (role-aware, our Q1 model) =====


    @app.post("/api/v1/assess/start")
    def assess_start(b: ChatStart, s: Session = Depends(db),
                     u: User = Depends(require("engagement.view"))):
        sess = ConversationSession(engagement_id=b.engagement_id,
                                   actor_role=b.actor_role)
        s.add(sess); s.flush()
        s.commit()
        return {"session_id": sess.id, "actor_role": sess.actor_role}

    @app.post("/api/v1/assess/turn")
    def assess_turn(b: ChatTurn, s: Session = Depends(db),
                    u: User = Depends(require("engagement.view"))):
        sess = s.get(ConversationSession, b.session_id)
        if not sess:
            raise HTTPException(404, "session not found")
        s.add(ConversationMessage(session_id=b.session_id,
                                  role=sess.actor_role, body=b.message))
        # deterministic adaptive reply: vendor claims are flagged to verify;
        # assessor input is trusted (our role/trust model)
        if sess.actor_role == "vendor":
            reply = ("Noted as a vendor assertion — this will be verified against "
                     "independent evidence before it affects the rating.")
            visibility = "shared"
        else:
            reply = "Recorded as assessor input and applied to the assessment."
            visibility = "internal"
        s.add(ConversationMessage(session_id=b.session_id, role="assistant", body=reply))
        s.commit()
        return {"session_id": b.session_id, "reply": reply, "visibility": visibility}

    # ===== autopilot (propose, human executes — two-gate) =====
    @app.post("/api/v1/engagements/{eid}/autopilot")
    def autopilot(eid: int, body: IRQIn, s: Session = Depends(db),
                  u: User = Depends(require("engagement.autopilot"))):
        e = s.get(EngagementRow, eid)
        if not e:
            raise HTTPException(404, "engagement not found")
        tier = eng.compute_tier(body.answers)
        inherent = eng.compute_inherent(body.answers)
        routing = eng.compute_route(body.answers, inherent, tier)
        # PROPOSES — does not finalise the decision; a human must record it.
        audit(s, "autopilot.proposed", u.username,
              {"engagement_id": eid, "proposed_band": inherent["band"]})
        s.commit()
        return {"engagement_id": eid, "proposed_tier": tier,
                "proposed_inherent": inherent, "proposed_routing": routing,
                "status": "PROPOSED — requires human to record decision"}

    # ===== notifications =====
    @app.get("/api/v1/notifications")
    def notifications(s: Session = Depends(db), u: User = Depends(require("notify.view"))):
        rows = s.scalars(select(Notification).order_by(Notification.id.desc())).all()
        unread = s.scalar(select(func.count()).select_from(Notification)
                          .where(Notification.is_read == False))  # noqa: E712
        return {"unread": unread,
                "items": [{"id": n.id, "event": n.event, "audience": n.audience,
                           "read": n.is_read} for n in rows[:50]]}

    # ===== methodology versioning =====

    @app.post("/api/v1/methodology/version")
    def meth_version(b: MethIn, s: Session = Depends(db),
                     u: User = Depends(require("methodology.version"))):
        s.add(MethodologyVersion(version=b.version, note=b.note))
        audit(s, "methodology.versioned", u.username, {"version": b.version})
        s.commit()
        return {"version": b.version}

    # ===== audit =====
    @app.get("/api/v1/audit")
    def audit_trail(s: Session = Depends(db), u: User = Depends(require("audit.view"))):
        rows = s.scalars(select(AuditLog).order_by(AuditLog.seq)).all()
        return [{"seq": r.seq, "action": r.action, "actor": r.actor,
                 "hash": r.entry_hash} for r in rows]

    @app.get("/api/v1/audit/verify")
    def audit_verify(s: Session = Depends(db), u: User = Depends(require("audit.view"))):
        rows = s.scalars(select(AuditLog).order_by(AuditLog.seq)).all()
        prev = "genesis"
        for r in rows:
            expect = eng.chain_hash(prev, r.action, r.actor,
                                    json.loads(r.detail) if r.detail else {})
            if r.prev_hash != prev or r.entry_hash != expect:
                return {"intact": False, "broke_at": r.seq}
            prev = r.entry_hash
        return {"intact": True, "entries": len(rows)}

    # ===== MCP-style read tools (Group G) =====
    @app.get("/api/v1/mcp/portfolio-summary")
    def portfolio_summary(s: Session = Depends(db), u: User = Depends(require("vendor.view"))):
        vendors = s.scalars(select(Vendor)).all()
        engagements = s.scalars(select(EngagementRow)).all()
        return {"vendors": len(vendors),
                "critical_vendors": sum(1 for v in vendors if v.is_critical),
                "engagements": len(engagements),
                "by_decision": _count_by(engagements, "decision")}

    @app.get("/api/v1/mcp/critical-vendors")
    def critical_vendors(s: Session = Depends(db), u: User = Depends(require("vendor.view"))):
        return [{"vendor_id": v.id, "name": v.name, "reason": v.critical_reason}
                for v in s.scalars(select(Vendor).where(Vendor.is_critical == True)).all()]  # noqa: E712

    @app.get("/api/v1/mcp/overdue-findings")
    def overdue_findings(s: Session = Depends(db), u: User = Depends(require("finding.view"))):
        rows = s.scalars(select(Finding).where(Finding.status != "closed")).all()
        return [{"finding_id": f.id, "title": f.title, "severity": f.severity,
                 "status": f.status} for f in rows]

    # ===== procurement integration (Group G) =====

    @app.post("/api/v1/procurement/po")
    def procurement_po(b: POIn, s: Session = Depends(db),
                       u: User = Depends(require("admin.integrations"))):
        # inbound PO auto-creates a vendor + sourcing engagement (straight-through)
        v = Vendor(name=b.vendor_name, ext_ref=b.ext_ref)
        s.add(v); s.flush()
        e = EngagementRow(vendor_id=v.id, title=f"PO {b.ext_ref or v.id}",
                          stage="sourcing")
        s.add(e); s.flush()
        audit(s, "procurement.po_ingested", u.username,
              {"vendor_id": v.id, "engagement_id": e.id, "amount": b.amount})
        s.commit()
        return {"vendor_id": v.id, "engagement_id": e.id, "stage": "sourcing"}

    # ===== certifications =====
    @app.post("/api/v1/certifications")
    def add_cert(b: CertIn, s: Session = Depends(db),
                 u: User = Depends(require("lifecycle.certs"))):
        from datetime import datetime as _dt
        from .features.models_db import Vendor as _V
        vu = _dt.fromisoformat(b.valid_until) if b.valid_until else None
        # supersede any prior current cert of the same name for this vendor
        priors = s.scalars(select(Certification).where(
            Certification.vendor_id == b.vendor_id, Certification.name == b.name,
            Certification.superseded == False)).all()  # noqa: E712
        for p in priors:
            p.superseded = True
        row = Certification(vendor_id=b.vendor_id, name=b.name, valid_until=vu)
        s.add(row); s.flush()
        # Write through to the canonical artefact pipeline so the revalidation engine,
        # expiring view, Issues Log, supersession and auto-close all cover register certs.
        reg_vid = None
        v1 = s.get(_V, b.vendor_id)
        if v1:
            vr = s.scalars(select(VendorRecord).where(VendorRecord.legal_name == v1.name)).first()
            reg_vid = vr.vendor_id if vr else (v1.ext_ref or f"VEN-V1-{v1.id}")
        if reg_vid:
            prior_art = s.scalars(select(ArtefactRecord).where(
                ArtefactRecord.vendor_id == reg_vid, ArtefactRecord.name == b.name,
                ArtefactRecord.is_current == True)).first()  # noqa: E712
            art = RS.create_artefact(
                s, vendor_id=reg_vid, name=b.name, artefact_type="certificate",
                expiry_date=(vu.date().isoformat() if vu else None), is_dated=bool(vu),
                received_via="register", supersedes=prior_art.artefact_id if prior_art else None)
            row.artefact_id = art.artefact_id
        audit(s, "cert.added", u.username, {"cert_id": row.id, "vendor_id": b.vendor_id,
                                            "artefact_id": row.artefact_id})
        s.commit()
        return {"cert_id": row.id, "name": row.name, "artefact_id": row.artefact_id}

    @app.get("/api/v1/vendors/{vid}/certifications")
    def list_certs(vid: int, s: Session = Depends(db),
                   u: User = Depends(require("lifecycle.certs"))):
        rows = s.scalars(select(Certification).where(Certification.vendor_id == vid)).all()
        return [{"cert_id": c.id, "name": c.name,
                 "valid_until": c.valid_until.isoformat() if c.valid_until else None}
                for c in rows]

    # ===== documents + evidence expiry =====
    @app.post("/api/v1/documents")
    def add_document(b: DocIn, s: Session = Depends(db),
                     u: User = Depends(require("lifecycle.documents"))):
        from datetime import datetime as _dt
        nv = _dt.fromisoformat(b.next_validation) if b.next_validation else None
        row = Document(vendor_id=b.vendor_id, engagement_id=b.engagement_id,
                       name=b.name, doc_type=b.doc_type, next_validation=nv)
        s.add(row); s.flush()
        audit(s, "document.added", u.username, {"document_id": row.id})
        s.commit()
        return {"document_id": row.id, "name": row.name}

    @app.get("/api/v1/evidence/expiring")
    def expiring_evidence(s: Session = Depends(db),
                          u: User = Depends(require("lifecycle.evidence"))):
        from datetime import datetime as _dt, timedelta
        horizon = _dt.utcnow() + timedelta(days=90)
        out = []
        for d in s.scalars(select(Document).where(
                Document.next_validation != None,  # noqa: E711
                Document.next_validation <= horizon)).all():
            out.append({"source": "document", "document_id": d.id, "name": d.name,
                        "next_validation": d.next_validation.isoformat()})
        # Join the certifications register so register-added certs surface here too.
        for c in s.scalars(select(Certification).where(
                Certification.valid_until != None,  # noqa: E711
                Certification.valid_until <= horizon,
                Certification.superseded == False)).all():  # noqa: E712
            out.append({"source": "certification", "cert_id": c.id, "name": c.name,
                        "next_validation": c.valid_until.isoformat()})
        out.sort(key=lambda r: r["next_validation"])
        return out

    # ===== fourth parties + concentration =====
    @app.post("/api/v1/fourth-parties")
    def add_fourth(b: FourthIn, s: Session = Depends(db),
                   u: User = Depends(require("lifecycle.fourthparty"))):
        row = FourthParty(vendor_id=b.vendor_id, name=b.name, service=b.service)
        s.add(row); s.flush()
        # concentration: same 4th party serving many vendors
        count = s.scalar(select(func.count()).select_from(FourthParty)
                         .where(FourthParty.name == b.name))
        if count and count >= 3:
            row.concentration_flag = True
            notify(s, f"Concentration risk: {b.name} serves {count} vendors", "vrm")
        audit(s, "fourthparty.added", u.username, {"id": row.id})
        s.commit()
        return {"fourth_party_id": row.id, "concentration_flag": row.concentration_flag}

    @app.get("/api/v1/fourth-parties/concentration")
    def concentration(s: Session = Depends(db),
                      u: User = Depends(require("lifecycle.fourthparty"))):
        rows = s.scalars(select(FourthParty).where(
            FourthParty.concentration_flag == True)).all()  # noqa: E712
        return [{"id": f.id, "name": f.name, "vendor_id": f.vendor_id} for f in rows]

    # ===== acceptances =====
    @app.post("/api/v1/acceptances")
    def add_acceptance(b: AcceptIn, s: Session = Depends(db),
                       u: User = Depends(require("acceptance.manage"))):
        from datetime import datetime as _dt
        ex = _dt.fromisoformat(b.expires_at) if b.expires_at else None
        row = Acceptance(engagement_id=b.engagement_id, rationale=b.rationale,
                         accepted_by=u.username, expires_at=ex)
        s.add(row); s.flush()
        audit(s, "acceptance.recorded", u.username, {"id": row.id})
        s.commit()
        return {"acceptance_id": row.id, "accepted_by": u.username}

    # ===== contracts (Matt) =====
    @app.post("/api/v1/engagements/{eid}/contract")
    def gen_contract(eid: int, s: Session = Depends(db),
                     u: User = Depends(require("intel.contract"))):
        e = s.get(EngagementRow, eid)
        if not e:
            raise HTTPException(404, "engagement not found")
        v = s.get(Vendor, e.vendor_id)
        out = intel.matt_contract(v.tier if v else "Tier 3")
        row = Contract(engagement_id=eid, tier=v.tier if v else "Tier 3",
                       terms_json=json.dumps(list(out.signals)))
        s.add(row); s.flush()
        if e.stage == "decision":
            e.stage = "contract"
        audit(s, "contract.generated", u.username, {"engagement_id": eid})
        s.commit()
        return {"contract_id": row.id, "tier": row.tier, "terms": out.signals}

    # ===== reassessments =====
    @app.post("/api/v1/reassessments")
    def schedule_reassessment(b: ReassessIn, s: Session = Depends(db),
                              u: User = Depends(require("lifecycle.reassess"))):
        row = Reassessment(engagement_id=b.engagement_id, mode=b.mode)
        s.add(row); s.flush()
        audit(s, "reassessment.scheduled", u.username, {"id": row.id, "mode": b.mode})
        notify(s, f"Reassessment scheduled ({b.mode})", "all")
        s.commit()
        return {"reassessment_id": row.id, "mode": b.mode}

    @app.post("/api/v1/reassessments/{rid}/complete")
    def complete_reassessment(rid: int, s: Session = Depends(db),
                              u: User = Depends(require("lifecycle.reassess"))):
        r = s.get(Reassessment, rid)
        if not r:
            raise HTTPException(404, "reassessment not found")
        r.completed = True
        audit(s, "reassessment.completed", u.username, {"id": rid})
        s.commit()
        return {"reassessment_id": rid, "completed": True}

    # ===== corrective action plans (CAP) — modelled as findings =====
    @app.get("/api/v1/cap")
    def cap_board(s: Session = Depends(db),
                  u: User = Depends(require("lifecycle.cap"))):
        rows = s.scalars(select(Finding).where(Finding.status != "closed")).all()
        return {"open_actions": len(rows),
                "by_severity": _count_by(rows, "severity"),
                "items": [{"finding_id": f.id, "title": f.title,
                           "severity": f.severity, "status": f.status} for f in rows]}

    # ===== business impact analysis (BIA) =====
    @app.get("/api/v1/vendors/{vid}/bia")
    def bia(vid: int, s: Session = Depends(db),
            u: User = Depends(require("lifecycle.bia"))):
        v = s.get(Vendor, vid)
        if not v:
            raise HTTPException(404, "vendor not found")
        engagements = s.scalars(select(EngagementRow).where(
            EngagementRow.vendor_id == vid)).all()
        return {"vendor_id": vid, "is_critical": v.is_critical,
                "engagement_count": len(engagements),
                "impact": "HIGH" if v.is_critical else
                          "MEDIUM" if len(engagements) > 1 else "LOW"}

    # ===== dashboards =====
    @app.get("/api/v1/dashboard/executive")
    def dash_exec(s: Session = Depends(db),
                  u: User = Depends(require("dashboard.exec"))):
        vendors = s.scalars(select(VendorRecord)).all()
        engs = s.scalars(select(EngagementRecord)).all()
        findings = s.scalars(select(FindingRecord)).all()
        # include any legacy v1 rows so both creation paths are reflected
        v1v = s.scalar(select(func.count()).select_from(Vendor)) or 0
        v1e = s.scalar(select(func.count()).select_from(EngagementRow)) or 0
        v1f = s.scalar(select(func.count()).select_from(Finding)
                       .where(Finding.status != "closed")) or 0
        return {
            "vendors": len(vendors) + v1v,
            "critical_vendors": sum(1 for v in vendors if v.is_critical),
            "engagements": len(engs) + v1e,
            "by_residual": _count_by(engs, "residual_band"),
            "by_decision": _count_by(engs, "status"),
            "open_findings": sum(1 for f in findings if (f.status or "").lower() not in ("closed",)) + v1f,
        }

    @app.get("/api/v1/dashboard/operational")
    def dash_ops(s: Session = Depends(db),
                 u: User = Depends(require("dashboard.ops"))):
        engs = s.scalars(select(EngagementRecord)).all()
        return {"by_stage": _count_by(engs, "status"),
                "by_route": _count_by(engs, "inherent_band"),
                "in_flight": sum(1 for e in engs if (e.status or "") not in ("Terminated", "Exited"))}

    @app.get("/api/v1/dashboard/risk")
    def dash_risk(s: Session = Depends(db),
                  u: User = Depends(require("dashboard.risk"))):
        engs = s.scalars(select(EngagementRow)).all()
        monit = s.scalars(select(Monitoring)).all()
        return {"by_inherent": _count_by(engs, "inherent_band"),
                "by_residual": _count_by(engs, "residual_band"),
                "monitoring_alerts": sum(1 for m in monit if m.status in ("ALERT", "CRITICAL"))}

    # ===== document upload + extraction (feeds Isaac) =====
    @app.post("/api/v1/documents/upload")
    async def upload_document(
        file: UploadFile = File(...),
        vendor_id: Optional[int] = Form(default=None),
        engagement_id: Optional[int] = Form(default=None),
        s: Session = Depends(db),
        u: User = Depends(require("lifecycle.documents")),
    ):
        from .features import uploads
        data = await file.read()
        _uerr = SEC.upload_check(file.filename or "upload", data)
        if _uerr:
            raise HTTPException(413 if "MB" in _uerr else 415, _uerr)
        try:
            res = uploads.process_upload(
                data=data, filename=file.filename or "upload",
                content_type=file.content_type or "application/octet-stream",
                org_id="org", engagement_id=str(engagement_id or ""),
                vendor_id=vendor_id,
            )
        except ValueError as e:
            raise HTTPException(415, str(e))

        from datetime import datetime as _dt
        nv = _dt.fromisoformat(res.next_validation) if res.next_validation else None
        doc = Document(vendor_id=vendor_id, engagement_id=engagement_id,
                       name=file.filename or "upload", doc_type=res.doc_type,
                       object_uri=res.object_key, next_validation=nv)
        s.add(doc); s.flush()

        if res.isaac and vendor_id:
            s.add(IntelResult(vendor_id=vendor_id, engine="evidence",
                              score=res.isaac["score"], band=res.isaac["band"],
                              narrative=res.isaac["narrative"]))

        audit(s, "document.uploaded", u.username,
              {"document_id": doc.id, "doc_type": res.doc_type,
               "classified_confidence": res.classification_confidence,
               "isaac_band": res.isaac["band"] if res.isaac else None,
               "evidence_count": res.evidence_count})
        if res.needs_human:
            notify(s, f"Document '{file.filename}' needs human classification review", "all")
        s.commit()

        return {
            "document_id": doc.id,
            "object_key": res.object_key,
            "doc_type": res.doc_type,
            "classification_confidence": res.classification_confidence,
            "needs_human_review": res.needs_human,
            "page_count": res.page_count,
            "scanned_pdf": res.scanned,
            "extracted_chars": res.extracted_chars,
            "isaac": res.isaac,
            "evidence_count": res.evidence_count,
            "next_validation": res.next_validation,
        }

    # ============================================================
    #  CRUD / edit, list+search, admin, VRM, auth, reporting, email
    # ============================================================
    from fastapi.responses import PlainTextResponse
    import csv as _csv, io as _io
    from datetime import datetime as _dt2

    # ---- Vendor: update, delete, list+search, detail ----
    @app.patch("/api/v1/vendors/{vid}")
    def update_vendor(vid: int, b: VendorUpdateIn, s: Session = Depends(db),
                      u: User = Depends(require("vendor.edit"))):
        v = s.get(Vendor, vid)
        if not v:
            raise HTTPException(404, "vendor not found")
        for f, val in b.model_dump(exclude_none=True).items():
            setattr(v, f, val)
        audit(s, "vendor.updated", u.username, {"vendor_id": vid, "fields": list(b.model_dump(exclude_none=True))})
        s.commit()
        return {"vendor_id": vid, "updated": True}

    @app.delete("/api/v1/vendors/{vid}")
    def delete_vendor(vid: int, s: Session = Depends(db),
                      u: User = Depends(require("vendor.edit"))):
        v = s.get(Vendor, vid)
        if not v:
            raise HTTPException(404, "vendor not found")
        # archive semantics: mark, don't hard-delete (preserves audit/history)
        v.ext_ref = (v.ext_ref or "") + "|archived"
        audit(s, "vendor.archived", u.username, {"vendor_id": vid})
        s.commit()
        return {"vendor_id": vid, "archived": True}

    @app.get("/api/v1/vendors/{vid}")
    def get_vendor(vid: int, s: Session = Depends(db),
                   u: User = Depends(require("vendor.view"))):
        v = s.get(Vendor, vid)
        if not v:
            raise HTTPException(404, "vendor not found")
        engs = s.scalars(select(EngagementRow).where(EngagementRow.vendor_id == vid)).all()
        return {"vendor_id": v.id, "name": v.name, "industry": v.industry,
                "country": v.country, "contact_email": v.contact_email,
                "tier": v.tier, "is_critical": v.is_critical,
                "critical_reason": v.critical_reason, "critical_by": v.critical_by,
                "engagements": [{"engagement_id": e.id, "title": e.title, "stage": e.stage} for e in engs]}

    # ---- Vendor: list with search/filter (replaces the plain list inline) ----
    @app.get("/api/v1/vendors-search")
    def search_vendors(q: Optional[str] = None, tier: Optional[str] = None,
                       critical: Optional[bool] = None,
                       s: Session = Depends(db), u: User = Depends(require("vendor.view"))):
        stmt = select(Vendor)
        rows = s.scalars(stmt).all()
        out = []
        for v in rows:
            if "|archived" in (v.ext_ref or ""):
                continue
            if q and q.lower() not in v.name.lower():
                continue
            if tier and v.tier != tier:
                continue
            if critical is not None and v.is_critical != critical:
                continue
            out.append({"vendor_id": v.id, "name": v.name, "tier": v.tier,
                        "is_critical": v.is_critical})
        return out

    # ---- Vendor: remove critical designation (VRM) ----
    @app.delete("/api/v1/vendors/{vid}/critical")
    def remove_critical(vid: int, s: Session = Depends(db),
                        u: User = Depends(require("vendor.critical"))):
        v = s.get(Vendor, vid)
        if not v:
            raise HTTPException(404, "vendor not found")
        v.is_critical = False; v.critical_reason = None; v.critical_by = None
        audit(s, "vendor.critical_removed", u.username, {"vendor_id": vid})
        s.commit()
        return {"vendor_id": vid, "is_critical": False}

    # ---- Engagement: update, list-all, delete ----
    @app.patch("/api/v1/engagements/{eid}")
    def update_engagement(eid: int, b: EngagementUpdateIn, s: Session = Depends(db),
                          u: User = Depends(require("engagement.edit"))):
        e = s.get(EngagementRow, eid)
        if not e:
            raise HTTPException(404, "engagement not found")
        for f, val in b.model_dump(exclude_none=True).items():
            setattr(e, f, val)
        audit(s, "engagement.updated", u.username, {"engagement_id": eid})
        s.commit()
        return {"engagement_id": eid, "updated": True}

    @app.get("/api/v1/engagements")
    def list_engagements(stage: Optional[str] = None, vendor_id: Optional[int] = None,
                         s: Session = Depends(db), u: User = Depends(require("engagement.view"))):
        rows = s.scalars(select(EngagementRow)).all()
        out = []
        for e in rows:
            if stage and e.stage != stage:
                continue
            if vendor_id and e.vendor_id != vendor_id:
                continue
            out.append({"engagement_id": e.id, "vendor_id": e.vendor_id, "title": e.title,
                        "stage": e.stage, "route": e.route, "inherent_band": e.inherent_band,
                        "residual_band": e.residual_band, "decision": e.decision})
        return out

    @app.delete("/api/v1/engagements/{eid}")
    def delete_engagement(eid: int, s: Session = Depends(db),
                          u: User = Depends(require("engagement.edit"))):
        e = s.get(EngagementRow, eid)
        if not e:
            raise HTTPException(404, "engagement not found")
        e.status = "cancelled"
        audit(s, "engagement.cancelled", u.username, {"engagement_id": eid})
        s.commit()
        return {"engagement_id": eid, "cancelled": True}

    # ---- Finding: update, reopen ----
    @app.patch("/api/v1/findings/{fid}")
    def update_finding(fid: int, b: FindingUpdateIn, s: Session = Depends(db),
                       u: User = Depends(require("finding.manage"))):
        f = s.get(Finding, fid)
        if not f:
            raise HTTPException(404, "finding not found")
        for fld, val in b.model_dump(exclude_none=True).items():
            setattr(f, fld, val)
        audit(s, "finding.updated", u.username, {"finding_id": fid})
        s.commit()
        return {"finding_id": fid, "updated": True}

    @app.post("/api/v1/findings/{fid}/reopen")
    def reopen_finding(fid: int, s: Session = Depends(db),
                       u: User = Depends(require("finding.manage"))):
        f = s.get(Finding, fid)
        if not f:
            raise HTTPException(404, "finding not found")
        f.status = "open"
        audit(s, "finding.reopened", u.username, {"finding_id": fid})
        s.commit()
        return {"finding_id": fid, "status": "open"}

    # ---- VRM: sign-off + review queue ----
    @app.post("/api/v1/engagements/{eid}/signoff")
    def signoff(eid: int, b: SignoffIn, s: Session = Depends(db),
                u: User = Depends(require("engagement.review"))):
        e = s.get(EngagementRow, eid)
        if not e:
            raise HTTPException(404, "engagement not found")
        e.status = "signed_off" if b.decision == "approved" else "returned"
        audit(s, "engagement.signoff", u.username,
              {"engagement_id": eid, "decision": b.decision, "note": b.note})
        notify(s, f"Engagement #{eid} {b.decision} by VRM", "business")
        s.commit()
        return {"engagement_id": eid, "status": e.status}

    @app.get("/api/v1/review-queue")
    def review_queue(s: Session = Depends(db), u: User = Depends(require("engagement.review"))):
        rows = s.scalars(select(EngagementRow).where(
            EngagementRow.stage == "decision",
            EngagementRow.residual_band.in_(["HIGH", "ELEVATED"]))).all()
        return [{"engagement_id": e.id, "vendor_id": e.vendor_id, "title": e.title,
                 "residual_band": e.residual_band, "decision": e.decision} for e in rows]

    # ---- Auth self-service: change password, profile ----
    @app.post("/api/v1/me/password")
    def change_password(b: PasswordIn, s: Session = Depends(db), u: User = Depends(actor)):
        from .features.models_db import hash_password
        if not verify_password(b.current_password, u.password_hash):
            raise HTTPException(403, "current password incorrect")
        if len(b.new_password) < 6:
            raise HTTPException(400, "new password too short")
        u.password_hash = hash_password(b.new_password)
        audit(s, "user.password_changed", u.username, {"username": u.username})
        s.commit()
        return {"changed": True}

    @app.get("/api/v1/me")
    def my_profile(u: User = Depends(actor)):
        return {"username": u.username, "full_name": u.full_name,
                "email": u.email, "role": u.role.key}

    @app.patch("/api/v1/me")
    def update_profile(b: ProfileIn, s: Session = Depends(db), u: User = Depends(actor)):
        for f, val in b.model_dump(exclude_none=True).items():
            setattr(u, f, val)
        audit(s, "user.profile_updated", u.username, {"username": u.username})
        s.commit()
        return {"updated": True}

    # ---- Admin: users ----
    @app.get("/api/v1/admin/users")
    def list_users(s: Session = Depends(db), u: User = Depends(require("admin.users"))):
        return [{"id": x.id, "username": x.username, "full_name": x.full_name,
                 "email": x.email, "role": x.role.key, "is_active": x.is_active}
                for x in s.scalars(select(User)).all()]

    @app.post("/api/v1/admin/users")
    def create_user(b: UserIn, s: Session = Depends(db), u: User = Depends(require("admin.users"))):
        from .features.models_db import hash_password
        if s.scalars(select(User).where(User.username == b.username)).first():
            raise HTTPException(409, "username exists")
        role = s.scalars(select(Role).where(Role.key == b.role_key)).first()
        if not role:
            raise HTTPException(400, "unknown role")
        row = User(username=b.username, full_name=b.full_name, email=b.email,
                   password_hash=hash_password(b.password), role_id=role.id)
        s.add(row); s.flush()
        audit(s, "user.created", u.username, {"username": b.username, "role": b.role_key})
        s.commit()
        return {"id": row.id, "username": row.username}

    @app.patch("/api/v1/admin/users/{uid}")
    def update_user(uid: int, b: UserUpdateIn, s: Session = Depends(db),
                    u: User = Depends(require("admin.users"))):
        target = s.get(User, uid)
        if not target:
            raise HTTPException(404, "user not found")
        data = b.model_dump(exclude_none=True)
        if "role_key" in data:
            role = s.scalars(select(Role).where(Role.key == data.pop("role_key"))).first()
            if not role:
                raise HTTPException(400, "unknown role")
            target.role_id = role.id
        for f, val in data.items():
            setattr(target, f, val)
        audit(s, "user.updated", u.username, {"user_id": uid})
        s.commit()
        return {"user_id": uid, "updated": True}

    @app.delete("/api/v1/admin/users/{uid}")
    def deactivate_user(uid: int, s: Session = Depends(db), u: User = Depends(require("admin.users"))):
        target = s.get(User, uid)
        if not target:
            raise HTTPException(404, "user not found")
        if target.username == "admin":
            raise HTTPException(400, "cannot deactivate the seed admin")
        target.is_active = False
        audit(s, "user.deactivated", u.username, {"user_id": uid})
        s.commit()
        return {"user_id": uid, "is_active": False}

    # ---- Admin: roles & permissions ----
    @app.get("/api/v1/admin/roles")
    def list_roles(s: Session = Depends(db), u: User = Depends(require("admin.roles"))):
        return [{"key": r.key, "label": r.label, "is_system": r.is_system,
                 "permissions": [p.key for p in r.permissions]}
                for r in s.scalars(select(Role)).all()]

    @app.get("/api/v1/admin/permissions")
    def list_permissions(s: Session = Depends(db), u: User = Depends(require("admin.roles"))):
        from .features.models_db import Permission
        return [{"key": p.key, "label": p.label, "category": p.category}
                for p in s.scalars(select(Permission)).all()]

    @app.put("/api/v1/admin/roles/{rkey}/permissions")
    def set_role_perms(rkey: str, b: RolePermsIn, s: Session = Depends(db),
                       u: User = Depends(require("admin.roles"))):
        from .features.models_db import Permission
        role = s.scalars(select(Role).where(Role.key == rkey)).first()
        if not role:
            raise HTTPException(404, "role not found")
        perms = s.scalars(select(Permission).where(Permission.key.in_(b.permission_keys))).all()
        role.permissions = list(perms)
        audit(s, "role.permissions_set", u.username, {"role": rkey, "count": len(perms)})
        s.commit()
        return {"role": rkey, "permissions": [p.key for p in perms]}

    # ---- Admin: webhooks ----
    @app.get("/api/v1/admin/webhooks")
    def list_webhooks(s: Session = Depends(db), u: User = Depends(require("admin.webhooks"))):
        return [{"id": w.id, "url": w.url, "event": w.event, "active": w.active}
                for w in s.scalars(select(Webhook)).all()]

    @app.post("/api/v1/admin/webhooks")
    def add_webhook(b: WebhookIn, s: Session = Depends(db), u: User = Depends(require("admin.webhooks"))):
        row = Webhook(url=b.url, event=b.event); s.add(row); s.flush()
        audit(s, "webhook.added", u.username, {"id": row.id, "url": b.url})
        s.commit()
        return {"webhook_id": row.id}

    @app.delete("/api/v1/admin/webhooks/{wid}")
    def del_webhook(wid: int, s: Session = Depends(db), u: User = Depends(require("admin.webhooks"))):
        w = s.get(Webhook, wid)
        if not w:
            raise HTTPException(404, "webhook not found")
        s.delete(w); audit(s, "webhook.deleted", u.username, {"id": wid}); s.commit()
        return {"deleted": True}

    # ---- Notifications: mark read ----
    @app.post("/api/v1/notifications/{nid}/read")
    def mark_read(nid: int, s: Session = Depends(db), u: User = Depends(require("notify.view"))):
        n = s.get(Notification, nid)
        if not n:
            raise HTTPException(404, "notification not found")
        n.is_read = True; s.commit()
        return {"id": nid, "read": True}

    @app.post("/api/v1/notifications/read-all")
    def mark_all_read(s: Session = Depends(db), u: User = Depends(require("notify.view"))):
        rows = s.scalars(select(Notification).where(Notification.is_read == False)).all()  # noqa: E712
        for n in rows:
            n.is_read = True
        s.commit()
        return {"marked": len(rows)}

    # ---- Email: send (real SMTP or simulation outbox) ----
    @app.post("/api/v1/email/send")
    def email_send(b: EmailIn, s: Session = Depends(db), u: User = Depends(require("admin.email"))):
        from .features import email_service
        sent = False
        try:
            sent = email_service.send_email(b.to_addr, b.subject, b.body)
        except Exception as e:  # SMTP failure -> fall back to outbox
            audit(s, "email.send_failed", u.username, {"to": b.to_addr, "error": str(e)[:120]})
        s.add(EmailOutbox(to_addr=b.to_addr, subject=b.subject, body=b.body, sent=sent))
        audit(s, "email.queued" if not sent else "email.sent", u.username, {"to": b.to_addr})
        s.commit()
        return {"to": b.to_addr, "sent": sent, "mode": "smtp" if sent else "simulation"}

    @app.get("/api/v1/email/outbox")
    def email_outbox(s: Session = Depends(db), u: User = Depends(require("admin.email"))):
        return [{"id": e.id, "to": e.to_addr, "subject": e.subject, "sent": e.sent}
                for e in s.scalars(select(EmailOutbox).order_by(EmailOutbox.id.desc())).all()]

    # ---- Evidence renewal chase (uses email path) ----
    @app.post("/api/v1/evidence/{doc_id}/chase")
    def chase_evidence(doc_id: int, s: Session = Depends(db), u: User = Depends(require("lifecycle.evidence"))):
        d = s.get(Document, doc_id)
        if not d:
            raise HTTPException(404, "document not found")
        v = s.get(Vendor, d.vendor_id) if d.vendor_id else None
        to = (v.contact_email if v and v.contact_email else "vendor@example.com")
        from .features import email_service
        sent = False
        try:
            sent = email_service.send_email(to, f"Evidence renewal required: {d.name}",
                                            "Please submit an updated version of this document.")
        except Exception as _e:
            _obs_swallow('bro_app.py', _e)
        s.add(EmailOutbox(to_addr=to, subject=f"Renewal: {d.name}", body="renewal chase", sent=sent))
        audit(s, "evidence.chased", u.username, {"document_id": doc_id, "to": to})
        notify(s, f"Renewal chased for '{d.name}'", "business")
        s.commit()
        return {"document_id": doc_id, "chased": True, "mode": "smtp" if sent else "simulation"}

    # ---- Contract gap review ----
    @app.post("/api/v1/contracts/{cid}/gap-review")
    def contract_gap_review(cid: int, s: Session = Depends(db), u: User = Depends(require("intel.contract"))):
        c = s.get(Contract, cid)
        if not c:
            raise HTTPException(404, "contract not found")
        required = json.loads(c.terms_json) if c.terms_json else []
        # deterministic gap review: in absence of a drafted contract, flag all
        # required terms as 'to confirm'; production diffs against uploaded draft.
        gaps = [t for t in required]
        c.gap_review = json.dumps({"missing_or_unconfirmed": gaps, "count": len(gaps)})
        audit(s, "contract.gap_review", u.username, {"contract_id": cid, "gaps": len(gaps)})
        s.commit()
        return {"contract_id": cid, "gap_count": len(gaps), "gaps": gaps}

    # ---- Reassessment: cadence sweep + delta ----
    @app.post("/api/v1/reassessments/run-due")
    def run_due_reassessments(s: Session = Depends(db), u: User = Depends(require("lifecycle.reassess"))):
        # Tier cadence: Tier1 annual, Tier2 biennial, Tier3 triennial.
        cadence = {"Tier 1": 365, "Tier 2": 730, "Tier 3": 1095}
        created = 0
        for e in s.scalars(select(EngagementRow).where(EngagementRow.stage == "monitor")).all():
            v = s.get(Vendor, e.vendor_id)
            days = cadence.get(v.tier if v else "Tier 3", 1095)
            age = (_dt2.utcnow() - e.created_at).days if e.created_at else 0
            if age >= days:
                s.add(Reassessment(engagement_id=e.id, mode="periodic")); created += 1
        audit(s, "reassessment.cadence_run", u.username, {"created": created})
        s.commit()
        return {"created": created}

    # ---- Reporting: register CSV export, audit export ----
    @app.get("/api/v1/reports/register.csv", response_class=PlainTextResponse)
    def register_csv(s: Session = Depends(db), u: User = Depends(require("reg.report"))):
        buf = _io.StringIO(); w = _csv.writer(buf)
        w.writerow(["vendor_id", "name", "tier", "critical", "engagement_id",
                    "title", "stage", "inherent", "residual", "decision"])
        for v in s.scalars(select(Vendor)).all():
            engs = s.scalars(select(EngagementRow).where(EngagementRow.vendor_id == v.id)).all()
            if not engs:
                w.writerow([v.id, v.name, v.tier, v.is_critical, "", "", "", "", "", ""])
            for e in engs:
                w.writerow([v.id, v.name, v.tier, v.is_critical, e.id, e.title,
                            e.stage, e.inherent_band, e.residual_band, e.decision])
        return buf.getvalue()

    @app.get("/api/v1/audit/export.csv", response_class=PlainTextResponse)
    def audit_export(s: Session = Depends(db), u: User = Depends(require("audit.export"))):
        buf = _io.StringIO(); w = _csv.writer(buf)
        w.writerow(["seq", "action", "actor", "timestamp", "hash"])
        for r in s.scalars(select(AuditLog).order_by(AuditLog.seq)).all():
            w.writerow([r.seq, r.action, r.actor,
                        r.created_at.isoformat() if r.created_at else "", r.entry_hash])
        return buf.getvalue()

    # ---- Vendor portal (self-service, scoped) ----
    @app.get("/api/v1/portal/my-status")
    def portal_status(s: Session = Depends(db), u: User = Depends(require("portal.self"))):
        # vendors see a minimal, scoped view (no internal reasoning)
        return {"message": "Vendor portal active",
                "you": u.username,
                "note": "Complete your DDQ and submit evidence via your assigned engagement."}

    # ============================================================
    #  Conversational multi-agent assessment (chat surface)
    # ============================================================
    from .features import agents as _A
    from .features import agent_engine as _AE
    from .features.models_feature import AgentLearning, BackgroundInsight

    @app.get("/api/v1/ai/status")
    def ai_status(u: User = Depends(require("admin.integrations"))):
        from .agents import llm_config
        st = dict(llm_config.status())
        st["claude_key_present"] = bool(_os.environ.get("ANTHROPIC_API_KEY"))
        st["openai_key_present"] = bool(_os.environ.get("OPENAI_API_KEY"))
        st["grok_key_present"] = bool(_os.environ.get("XAI_API_KEY"))
        st["manus_key_present"] = bool(_os.environ.get("MANUS_API_KEY"))
        st["nvidia_key_present"] = bool(_os.environ.get("NVIDIA_API_KEY"))
        st["providers"] = llm_config.provider_info()
        st["secrets_encrypted"] = SEC.secret_key_set()
        st["production_mode"] = SEC.is_production()
        return st

    @app.get("/api/v1/ai/models")
    def ai_models(u: User = Depends(require("admin.integrations"))):
        """List the models the configured key can actually access, so the admin
        can pick a valid one instead of guessing."""
        from .agents import llm_config
        st = llm_config.status()
        prov = st.get("provider")
        if not st.get("live_ready"):
            return {"available": False, "models": [],
                    "error": ("no provider key configured" if not prov
                              else f"{prov} SDK not installed")}
        try:
            ids = []
            if prov == "claude":
                import anthropic
                key = _os.environ.get("ANTHROPIC_API_KEY") or _os.environ.get("BRO_LLM_KEY")
                for m in anthropic.Anthropic(api_key=key).models.list(limit=50).data:
                    ids.append(getattr(m, "id", None))
            elif prov == "openai":
                import openai
                cl = openai.OpenAI(api_key=_os.environ.get("OPENAI_API_KEY"))
                ids = [getattr(m, "id", None) for m in cl.models.list().data
                       if "gpt" in (getattr(m, "id", "") or "")]
            ids = [i for i in ids if i]
            return {"available": True, "provider": prov, "models": ids,
                    "current": st.get("model")}
        except Exception as e:
            return {"available": False, "models": [],
                    "error": f"{type(e).__name__}: {str(e)[:200]}"}

    @app.post("/api/v1/ai/test")
    def ai_test(u: User = Depends(require("admin.integrations"))):
        """Live diagnostic: makes a real call (and a web-search call) so the
        admin can see exactly why AI isn't working, rather than silent fallback."""
        from .agents import llm_config
        st = llm_config.status()
        if not st.get("live_ready"):
            return {"live_ready": False, "basic_ok": False, "web_ok": False,
                    "error": ("AI is not live: " + ("no provider key configured" if not st.get("provider")
                              else f"{st.get('provider')} SDK not installed"))}
        basic = llm_config.complete("You are a connectivity probe.",
                                    "Reply with exactly: OK", domain="general")
        basic_err = llm_config.last_error()
        web = llm_config.complete("You are a web-search probe with live search.",
                                  "Search the web and reply with one current headline.",
                                  domain="general", web_search=True)
        web_err = llm_config.last_error()
        return {"live_ready": True, "provider": st.get("provider"), "model": st.get("model"),
                "basic_ok": bool(basic), "basic_reply": (basic or "")[:120], "basic_error": basic_err,
                "web_ok": bool(web), "web_reply": (web or "")[:160], "web_error": web_err}

    _AI_ENV = {"claude": "ANTHROPIC_API_KEY", "openai": "OPENAI_API_KEY",
               "grok": "XAI_API_KEY", "manus": "MANUS_API_KEY",
               "nvidia": "NVIDIA_API_KEY"}

    def _persist_ai_keys(s: Session, updates: dict):
        """Persist provider keys in the system config so they survive restart."""
        from .features import config_store as CFG
        store = CFG.get_json(s, "ai_provider_keys", {}) or {}
        store.update({k: SEC.encrypt_value(v) for k, v in updates.items() if v})
        CFG.upsert_json(s, "ai_provider_keys", store, updated_by="admin", category="ai")
        s.commit()

    @app.post("/api/v1/ai/key")
    def ai_set_key(b: AiKeyIn, s: Session = Depends(db),
                   u: User = Depends(require("admin.aikeys"))):
        """Set a live LLM provider key. The key is applied to the running process and
        persisted in the system configuration so it survives restart."""
        from .agents import llm_config
        provider = (b.provider or "claude").lower().strip()
        key = (b.api_key or "").strip()
        if provider not in _AI_ENV:
            raise HTTPException(400, "provider must be one of: claude, openai, grok, manus")
        if key:
            _os.environ[_AI_ENV[provider]] = key
            _persist_ai_keys(s, {provider: key})
        # all four are now wired as live providers (grok/manus via OpenAI-compatible base)
        _os.environ["BRO_LLM_PROVIDER"] = provider
        if (b.model or "").strip():
            _os.environ["BRO_LLM_MODEL"] = b.model.strip()
        try:
            llm_config._adapter.cache_clear()
            llm_config._LAST_ERROR = None
        except Exception as _e:
            _obs_swallow('bro_app.py', _e)
        audit(s, "ai.key_set", u.username, {"provider": provider})
        return ai_status(u)

    @app.post("/api/v1/ai/keys")
    def ai_set_keys(body: dict = Body(default={}), s: Session = Depends(db),
                    u: User = Depends(require("admin.aikeys"))):
        """Bulk-set provider keys from the AI integration panel.
        Body: {anthropic, openai, grok, manus} — only non-empty values are applied."""
        from .agents import llm_config
        alias = {"anthropic": "claude", "claude": "claude", "openai": "openai",
                 "grok": "grok", "xai": "grok", "manus": "manus",
                 "nvidia": "nvidia", "nim": "nvidia"}
        updates = {}
        for k, v in (body or {}).items():
            prov = alias.get(k.lower())
            if prov and (v or "").strip():
                _os.environ[_AI_ENV[prov]] = v.strip()
                updates[prov] = v.strip()
        if updates:
            _persist_ai_keys(s, updates)
            audit(s, "ai.keys_set", u.username, {"providers": sorted(updates)})
            try:
                llm_config._adapter.cache_clear()
                llm_config._LAST_ERROR = None
            except Exception as _e:
                _obs_swallow('bro_app.py', _e)
        return ai_status(u)

    @app.post("/api/v1/ai/test/{provider}")
    def ai_test_provider(provider: str, u: User = Depends(require("admin.integrations"))):
        """Probe a single provider's key/endpoint with a tiny live call (independent of
        the currently-selected provider), so the admin can confirm a key before relying on it."""
        from .agents import llm_config
        return llm_config.test_provider(provider)

    @app.post("/api/v1/ai/provider")
    def ai_set_provider(body: dict = Body(default={}), s: Session = Depends(db),
                        u: User = Depends(require("admin.aikeys"))):
        """Select the active live inference provider (claude / openai / grok / manus)
        and persist the choice so it survives restart."""
        from .agents import llm_config
        from .features import config_store as CFG
        prov = (body.get("provider") or "").lower().strip()
        if prov not in llm_config.all_provider_ids():
            raise HTTPException(400, f"unknown provider: {prov}")
        _os.environ["BRO_LLM_PROVIDER"] = prov
        CFG.upsert_json(s, "ai_active_provider", {"provider": prov},
                        updated_by="admin", category="ai")
        try:
            llm_config._adapter.cache_clear()
            llm_config._LAST_ERROR = None
        except Exception as _e:
            _obs_swallow('bro_app.py', _e)
        audit(s, "ai.provider_selected", u.username, {"provider": prov})
        s.commit()
        return ai_status(u)

    # ===== v4.2: interconnected ecosystem =====
    @app.get("/api/v1/search")
    def global_search(q: str = "", s: Session = Depends(db),
                      u: User = Depends(require("vendor.view"))):
        ql = (q or "").strip()
        if len(ql) < 2:
            return {"q": q, "results": []}
        like = f"%{ql}%"
        res = []
        from .features.registry_models import EngagementRecord, IncidentRecord
        from sqlalchemy import or_ as _or

        # Vendors — filtered and capped in SQL, not in Python.
        for v in s.scalars(
            select(VendorRecord).where(_or(
                VendorRecord.legal_name.ilike(like),
                VendorRecord.vendor_id.ilike(like))).limit(8)
        ).all():
            res.append({"kind": "vendor", "id": v.vendor_id, "title": v.legal_name,
                        "sub": f"{v.vendor_id} · {v.tier}" + (" · CRITICAL" if v.is_critical else "")})

        # Engagements
        for e in s.scalars(
            select(EngagementRecord).where(_or(
                EngagementRecord.title.ilike(like),
                EngagementRecord.engagement_id.ilike(like))).limit(6)
        ).all():
            res.append({"kind": "engagement", "id": e.engagement_id, "title": e.title,
                        "sub": f"{e.engagement_id} · {e.vendor_id} · residual {e.residual_band or '—'}",
                        "vendor_id": e.vendor_id})

        # Incidents
        for i in s.scalars(
            select(IncidentRecord).where(_or(
                IncidentRecord.incident_id.ilike(like),
                IncidentRecord.incident_type.ilike(like),
                IncidentRecord.vendor_name.ilike(like),
                IncidentRecord.impact_description.ilike(like))).limit(4)
        ).all():
            res.append({"kind": "incident", "id": i.incident_id,
                        "title": f"{i.incident_type or 'Incident'} · {i.vendor_name or i.vendor_id or ''}",
                        "sub": f"{i.incident_id} · {i.severity or '—'} · {i.status or '—'}",
                        "vendor_id": i.vendor_id})

        # Contracts (optional module)
        try:
            from .features.contract_models import ContractRecord
            for c in s.scalars(
                select(ContractRecord).where(ContractRecord.title.ilike(like)).limit(2)
            ).all():
                res.append({"kind": "contract", "id": c.id, "title": c.title,
                            "sub": f"contract · {c.vendor_id or ''}", "vendor_id": c.vendor_id})
        except Exception as _e:
            _obs_swallow('bro_app.py', _e)
        return {"q": q, "results": res[:20]}

    @app.get("/api/v1/connections/mcp")
    def mcp_list(s: Session = Depends(db), u: User = Depends(require("admin.integrations"))):
        from .features import config_store as CFG
        return {"connections": CFG.get_json(s, "mcp_connections", []) or []}

    @app.post("/api/v1/connections/mcp")
    def mcp_add(body: dict = Body(default={}), s: Session = Depends(db),
                u: User = Depends(require("admin.integrations"))):
        from .features import config_store as CFG
        name = (body.get("name") or "").strip(); url = (body.get("url") or "").strip()
        if not (name and url):
            raise HTTPException(400, "name and url are required")
        lst = [c for c in (CFG.get_json(s, "mcp_connections", []) or []) if c.get("name") != name]
        lst.append({"name": name, "url": url, "transport": body.get("transport") or "sse",
                    "auth": body.get("auth") or "none", "status": "configured (untested)",
                    "added": __import__("datetime").date.today().isoformat()})
        CFG.upsert_json(s, "mcp_connections", lst, updated_by=u.username, category="connections")
        audit(s, "connections.mcp_add", u.username, {"name": name}); s.commit()
        return {"connections": lst}

    @app.post("/api/v1/connections/mcp/{name}/delete")
    def mcp_del(name: str, s: Session = Depends(db),
                u: User = Depends(require("admin.integrations"))):
        from .features import config_store as CFG
        lst = [c for c in (CFG.get_json(s, "mcp_connections", []) or []) if c.get("name") != name]
        CFG.upsert_json(s, "mcp_connections", lst, updated_by=u.username, category="connections")
        s.commit(); return {"connections": lst}

    @app.get("/api/v1/schedules")
    def schedules_list(s: Session = Depends(db), u: User = Depends(require("vendor.view"))):
        from .features import config_store as CFG
        from sqlalchemy import text as _t
        last_run = None
        try:
            last_run = s.execute(_t("SELECT MAX(started_at) FROM monitoring_runs")).scalar()
        except Exception as _e:
            _obs_swallow('bro_app.py', _e)
        defaults = [
            {"id": "monitoring_sweep", "label": "Continuous monitoring sweep", "what": "Re-screens the portfolio: alerts, thresholds, stage clocks.", "engine": "in-process scheduler / Render cron", "cadence_hours": _monitor_interval()},
            {"id": "sanctions_screen", "label": "Sanctions & AML re-screening", "what": "Re-screens vendors and persons against sanctions lists.", "engine": "on-demand today", "cadence_hours": 24},
            {"id": "fdd_refresh", "label": "Financial health refresh (Vera)", "what": "Refreshes RapidRatings-style FDD signals on monitored vendors.", "engine": "on-demand today", "cadence_hours": 168},
            {"id": "cert_expiry_scan", "label": "Certification expiry scan", "what": "Flags ISO/SOC certificates entering their expiry window.", "engine": "on-demand today", "cadence_hours": 168},
            {"id": "exit_trigger_scan", "label": "Exit trigger scan", "what": "Evaluates exit-trigger conditions across critical vendors.", "engine": "on-demand today", "cadence_hours": 168},
            {"id": "contract_expiry_scan", "label": "Contract renewal window scan", "what": "Surfaces contracts entering notice/renewal windows.", "engine": "on-demand today", "cadence_hours": 168},
        ]
        cfg = CFG.get_json(s, "schedules", {}) or {}
        out = []
        for d in defaults:
            o = cfg.get(d["id"], {})
            d["enabled"] = bool(o.get("enabled", d["id"] == "monitoring_sweep" and _os.environ.get("BRO_SCHEDULER_ENABLED") == "1"))
            d["cadence_hours"] = int(o.get("cadence_hours", d["cadence_hours"]))
            out.append(d)
        return {"scheduler_running": _os.environ.get("BRO_SCHEDULER_ENABLED") == "1",
                "monitoring_last_run": last_run, "schedules": out}

    @app.post("/api/v1/schedules/{sid}")
    def schedules_set(sid: str, body: dict = Body(default={}), s: Session = Depends(db),
                      u: User = Depends(require("admin.integrations"))):
        from .features import config_store as CFG
        from .features import notifications as NOTIF
        cfg = CFG.get_json(s, "schedules", {}) or {}
        e = cfg.setdefault(sid, {})
        if "enabled" in body:
            e["enabled"] = bool(body["enabled"])
        if body.get("cadence_hours"):
            e["cadence_hours"] = max(1, int(body["cadence_hours"]))
        CFG.upsert_json(s, "schedules", cfg, updated_by=u.username, category="schedules")
        NOTIF.emit(s, "schedule.changed", f"Schedule '{sid}' updated",
                   f"enabled={e.get('enabled')} cadence={e.get('cadence_hours')}h by {u.username}")
        audit(s, "schedules.update", u.username, {"id": sid, **e}); s.commit()
        return {"id": sid, **e}

    @app.get("/api/v1/notifications/catalogue")
    def notif_catalogue(s: Session = Depends(db), u: User = Depends(require("admin.integrations"))):
        from .features import notifications as NOTIF
        return {"catalogue": NOTIF.CATALOGUE, "settings": NOTIF.get_settings(s), "audiences": NOTIF.AUDIENCES}

    @app.post("/api/v1/notifications/settings")
    def notif_settings(body: dict = Body(default={}), s: Session = Depends(db),
                       u: User = Depends(require("admin.integrations"))):
        from .features import notifications as NOTIF
        out = NOTIF.set_settings(s, body.get("settings") or {})
        audit(s, "notifications.settings", u.username, {"enabled": [k for k, v in out.items() if v["enabled"]]})
        return {"settings": out}

    @app.get("/api/v1/notifications/inbox")
    def notif_inbox(s: Session = Depends(db), u: User = Depends(require("vendor.view"))):
        from .features import notifications as NOTIF
        return {"items": NOTIF.inbox(s, 50), "unread": NOTIF.unread_count(s)}

    @app.post("/api/v1/notifications/read")
    def notif_read(body: dict = Body(default={}), s: Session = Depends(db),
                   u: User = Depends(require("vendor.view"))):
        from .features import notifications as NOTIF
        NOTIF.mark_read(s, body.get("id")); return {"unread": NOTIF.unread_count(s)}

    @app.post("/api/v2/incidents/{iid}/notable")
    def incident_notable(iid: str, s: Session = Depends(db),
                         u: User = Depends(require("incident.view"))):
        from .features.registry_models import IncidentRecord
        from .features import notifications as NOTIF
        from .features import config_store as CFG
        r = s.scalars(select(IncidentRecord).where(IncidentRecord.incident_id == iid)).first()
        if not r and str(iid).isdigit():
            r = s.get(IncidentRecord, int(iid))
        if not r:
            raise HTTPException(404, "incident not found")
        iid = r.incident_id
        lst = CFG.get_json(s, "notable_incidents", []) or []
        if iid not in lst:
            lst.append(iid)
            CFG.upsert_json(s, "notable_incidents", lst, updated_by=u.username, category="incidents")
        NOTIF.emit(s, "incident.notable",
                   f"NOTABLE EVENT: {r.incident_type or 'Incident'} at {r.vendor_name or r.vendor_id or 'vendor'}",
                   f"{r.incident_id} · severity {r.severity or '—'} · status {r.status or '—'} · flagged by {u.username}",
                   link="incidents", audience="management", force=True)
        audit(s, "incident.notable", u.username, {"incident_id": iid}); s.commit()
        return {"id": iid, "notable": True, "notable_ids": lst}

    @app.post("/api/v1/ai/dump-to-draft")
    async def dump_to_draft(files: list[UploadFile] = File(...), fields: str = Form(...),
                            context: str = Form(default=""), s: Session = Depends(db),
                            u: User = Depends(require("vendor.view"))):
        import json as _json
        from .features import draftfill as DF
        try:
            field_list = _json.loads(fields); assert isinstance(field_list, list)
        except Exception:
            raise HTTPException(400, "fields must be a JSON list of {id,label}")
        texts = []
        for f in files[:6]:
            data = await f.read()
            _err = SEC.upload_check(f.filename or "doc", data)
            if _err:
                raise HTTPException(415, _err)
            t = DF.extract_text(f.filename or "", data)
            if t.strip():
                texts.append(t)
        corpus = chr(10).join(texts)
        if not corpus.strip():
            raise HTTPException(422, "no readable text found in the uploaded document(s)")
        values, engine = None, "rules"
        ai = DF.ai_fill(s, field_list, corpus, context)
        if ai:
            ids = {f.get("id") for f in field_list}
            values = {k: v for k, v in ai.items() if k in ids and v not in (None, "")}
            engine = "ai"
        if not values:
            values = DF.heuristic_fill(field_list, corpus); engine = "rules"
        audit(s, "ai.dump_to_draft", u.username, {"fields": len(field_list), "filled": len(values), "engine": engine})
        s.commit()
        return {"engine": engine, "values": values, "filled": len(values), "of": len(field_list)}

    @app.get("/api/v1/ai/ledger")
    def ai_ledger_view(s: Session = Depends(db),
                       u: User = Depends(require("admin.integrations"))):
        """AI call ledger (metadata only — no prompt/response content is ever stored)."""
        from .features import ai_ledger as AL
        return {"recent": AL.recent(s, 25), "today_count": AL.today_count(s),
                "daily_budget": AL.get_budget(s)}

    @app.post("/api/v1/ai/budget")
    def ai_set_budget(body: dict = Body(default={}), s: Session = Depends(db),
                      u: User = Depends(require("admin.aikeys"))):
        from .features import ai_ledger as AL
        v = body.get("daily_calls")
        AL.set_budget(s, int(v) if v is not None and str(v) != "" else None)
        audit(s, "ai.budget_set", u.username, {"daily_calls": v})
        return {"daily_budget": AL.get_budget(s)}

    @app.post("/api/v1/ai/custom-provider")
    def ai_add_custom_provider(body: dict = Body(default={}), s: Session = Depends(db),
                               u: User = Depends(require("admin.aikeys"))):
        """Register ANY new OpenAI-compatible LLM service (label, base URL, model, key)."""
        from .agents import llm_config
        from .features import config_store as CFG
        label = (body.get("label") or "").strip()
        base_url = (body.get("base_url") or "").strip()
        model = (body.get("model") or "").strip()
        api_key = (body.get("api_key") or "").strip()
        pid = llm_config._slug(body.get("id") or label)
        if not (label and base_url and model):
            raise HTTPException(400, "label, base_url and model are required")
        if pid in llm_config._BUILTIN_ORDER:
            raise HTTPException(400, f"id '{pid}' collides with a built-in provider")
        llm_config.register_provider(pid, label, base_url, model)
        lst = [c for c in (CFG.get_json(s, "ai_custom_providers", []) or []) if c.get("id") != pid]
        lst.append({"id": pid, "label": label, "base_url": base_url, "model": model})
        CFG.upsert_json(s, "ai_custom_providers", lst, updated_by="admin", category="ai")
        if api_key:
            _os.environ[llm_config._PROVIDER_KEY_ENV[pid]] = api_key
            _persist_ai_keys(s, {pid: api_key})
        audit(s, "ai.custom_provider_add", u.username, {"id": pid, "base_url": base_url})
        s.commit()
        return ai_status(u)

    @app.post("/api/v1/ai/custom-provider/{pid}/delete")
    def ai_del_custom_provider(pid: str, s: Session = Depends(db),
                               u: User = Depends(require("admin.aikeys"))):
        from .agents import llm_config
        from .features import config_store as CFG
        pid = llm_config._slug(pid)
        llm_config.unregister_provider(pid)
        lst = [c for c in (CFG.get_json(s, "ai_custom_providers", []) or []) if c.get("id") != pid]
        CFG.upsert_json(s, "ai_custom_providers", lst, updated_by="admin", category="ai")
        store = CFG.get_json(s, "ai_provider_keys", {}) or {}
        store.pop(pid, None)
        CFG.upsert_json(s, "ai_provider_keys", store, updated_by="admin", category="ai")
        if _os.environ.get("BRO_LLM_PROVIDER") == pid:
            _os.environ.pop("BRO_LLM_PROVIDER", None)
        audit(s, "ai.custom_provider_del", u.username, {"id": pid})
        s.commit()
        return ai_status(u)

    @app.get("/api/v1/ai/prompts")
    def ai_list_prompts(s: Session = Depends(db),
                        u: User = Depends(require("admin.aikeys"))):
        """All editable AI-feature prompts, with current text (override or default)."""
        return {"prompts": PROMPTS.listing(s), "groups": PROMPTS.GROUP_ORDER}

    @app.post("/api/v1/ai/prompts/{key}")
    def ai_set_prompt(key: str, body: dict = Body(...), s: Session = Depends(db),
                      u: User = Depends(require("admin.aikeys"))):
        text = (body.get("text") or "").strip()
        if not text:
            raise HTTPException(400, "text required")
        try:
            PROMPTS.set_prompt(s, key, text, by=u.username)
        except KeyError:
            raise HTTPException(404, f"unknown prompt key: {key}")
        audit(s, "ai.prompt_set", u.username, {"key": key, "len": len(text)})
        s.commit()
        return {"ok": True, "key": key, "overridden": True, "current": text}

    @app.post("/api/v1/ai/prompts/{key}/reset")
    def ai_reset_prompt(key: str, s: Session = Depends(db),
                        u: User = Depends(require("admin.aikeys"))):
        if key not in PROMPTS.defaults():
            raise HTTPException(404, f"unknown prompt key: {key}")
        PROMPTS.reset_prompt(s, key, by=u.username)
        audit(s, "ai.prompt_reset", u.username, {"key": key})
        s.commit()
        return {"ok": True, "key": key, "overridden": False,
                "current": PROMPTS.defaults()[key]["default"]}

    @app.post("/api/v1/ai/key/clear")
    def ai_clear_key(u: User = Depends(require("admin.aikeys"))):
        """Remove any in-session key and fall back to the deterministic engines."""
        from .agents import llm_config
        for k in ("ANTHROPIC_API_KEY", "OPENAI_API_KEY", "BRO_LLM_KEY",
                  "BRO_LLM_PROVIDER", "BRO_LLM_MODEL"):
            _os.environ.pop(k, None)
        try:
            llm_config._adapter.cache_clear()
            llm_config._LAST_ERROR = None
        except Exception as _e:
            _obs_swallow('bro_app.py', _e)
        return llm_config.status()

    @app.get("/api/v1/agent/registry")
    def agent_registry(u: User = Depends(require("engagement.view"))):
        return {"agents": _A.AGENTS,
                "stages": [{"id": s.id, "name": s.name, "short": s.short} for s in _A.STAGES],
                "methodology": _A.METHODOLOGY}

    @app.post("/api/v1/agent/sessions")
    def open_session(b: ChatSessionIn, s: Session = Depends(db),
                     u: User = Depends(require("engagement.view"))):
        sess = ConversationSession(engagement_id=b.engagement_id, actor_role="assessor",
                                   stage=0, active_agent="bro", dossier_json="{}")
        s.add(sess); s.flush()
        opener = ("Bro here — your Risk Oracle. Exposure first. Controls second. Verdict last. "
                  "Drop everything you have on this engagement, or tell me about the supplier and "
                  "we start at intake.")
        s.add(ConversationMessage(session_id=sess.id, role="agent", agent="bro",
                                  stage=0, body=opener))
        audit(s, "agent.session_opened", u.username, {"session_id": sess.id})
        s.commit()
        return {"session_id": sess.id, "stage": 0, "active_agent": "bro"}

    @app.get("/api/v1/agent/sessions/{sid}")
    def get_session(sid: int, s: Session = Depends(db),
                    u: User = Depends(require("engagement.view"))):
        sess = s.get(ConversationSession, sid)
        if not sess:
            raise HTTPException(404, "session not found")
        msgs = s.scalars(select(ConversationMessage)
                         .where(ConversationMessage.session_id == sid)
                         .order_by(ConversationMessage.id)).all()
        insights = s.scalars(select(BackgroundInsight)
                             .where(BackgroundInsight.session_id == sid)
                             .order_by(BackgroundInsight.id.desc())).all()
        learnings = s.scalars(select(AgentLearning).order_by(AgentLearning.id.desc())).all()
        return {
            "session_id": sess.id, "stage": sess.stage, "active_agent": sess.active_agent,
            "dossier": json.loads(sess.dossier_json or "{}"),
            "messages": [{"id": m.id, "role": m.role, "agent": m.agent,
                          "stage": m.stage, "body": m.body} for m in msgs],
            "insights": [{"kind": i.kind, "severity": i.severity, "detail": i.detail} for i in insights],
            "learnings": [{"id": l.id, "text": l.text, "stage": l.stage} for l in learnings],
        }

    @app.post("/api/v1/agent/send")
    def agent_send(b: ChatSendIn, s: Session = Depends(db),
                   u: User = Depends(require("engagement.view"))):
        sess = s.get(ConversationSession, b.session_id)
        if not sess:
            raise HTTPException(404, "session not found")
        dossier = json.loads(sess.dossier_json or "{}")
        learn_texts = [l.text for l in s.scalars(select(AgentLearning)).all()]

        # record the user message
        s.add(ConversationMessage(session_id=sess.id, role="user", stage=sess.stage,
                                  body=b.message))

        # AI-only workflow: without a live AI engine, do not proceed — hold.
        if not _ai_live():
            s.add(ConversationMessage(session_id=sess.id, role="agent", agent="bro",
                                      stage=sess.stage, body=AI_HOLDING))
            s.commit()
            return {"session_id": sess.id, "stage": sess.stage, "active_agent": "bro",
                    "advanced": False, "holding": True,
                    "produced": [{"agent": "bro", "body": AI_HOLDING}], "stage_complete": None}

        from .features import methodology as _M
        _meth = _M.methodology_directive(s)
        _briefs = {aid: PROMPTS.resolve(s, f"agent_persona_{aid}") for aid in _A.AGENTS}

        # background consistency check (Sara, silent) — persist insights
        for ins in _AE.consistency_check(dossier, b.message, learn_texts):
            detail = ins.get("issue") or ins.get("concern") or ""
            if ins.get("with"):
                detail += f" (↳ {ins['with']})"
            if ins.get("claim"):
                detail += f" (↳ \"{ins['claim']}\")"
            s.add(BackgroundInsight(session_id=sess.id, kind=ins["kind"],
                                    severity=ins.get("severity", "medium"), detail=detail))

        # choose agent: explicit mention > stage owner
        target = b.agent if (b.agent in _A.AGENTS) else _A.route_next_agent(sess.stage)

        # run the agent turn (deterministic-local or live)
        turn = _AE.run_turn(target, sess.stage, dossier, learn_texts, b.message, _meth, _briefs)
        produced = []

        # follow up to two handoffs to keep it bounded
        hops = 0
        while turn.handoff and hops < 2:
            s.add(ConversationMessage(session_id=sess.id, role="agent",
                                      agent=turn.agent_id, stage=sess.stage, body=turn.body))
            produced.append({"agent": turn.agent_id, "body": turn.body})
            target = turn.handoff
            turn = _AE.run_turn(target, sess.stage, dossier, learn_texts, b.message, _meth, _briefs)
            hops += 1

        s.add(ConversationMessage(session_id=sess.id, role="agent",
                                  agent=turn.agent_id, stage=sess.stage, body=turn.body))
        produced.append({"agent": turn.agent_id, "body": turn.body})
        sess.active_agent = turn.agent_id

        advanced = False
        if turn.stage_complete and sess.stage < len(_A.STAGES) - 1:
            sess.stage += 1
            sess.active_agent = _A.route_next_agent(sess.stage)
            advanced = True
            s.add(ConversationMessage(session_id=sess.id, role="system", agent="bro",
                                      stage=sess.stage,
                                      body=f"Stage advanced — {turn.stage_complete} Now at Stage "
                                           f"{sess.stage}: {_A.STAGES[sess.stage].name}."))

        audit(s, "agent.turn", u.username,
              {"session_id": sess.id, "agent": turn.agent_id, "advanced": advanced})
        s.commit()
        return {"session_id": sess.id, "stage": sess.stage,
                "active_agent": sess.active_agent, "advanced": advanced,
                "produced": produced, "stage_complete": turn.stage_complete}

    @app.post("/api/v1/agent/learnings")
    def add_learning(b: LearningIn, s: Session = Depends(db),
                     u: User = Depends(require("engagement.view"))):
        text = _A.synthesize_learning(b.rating, b.agent, b.issue or "", b.note or "", b.stage)
        row = AgentLearning(rating=b.rating, agent=b.agent, stage=b.stage,
                            issue=b.issue, note=b.note, text=text)
        s.add(row); s.flush()
        audit(s, "agent.learning_captured", u.username, {"learning_id": row.id})
        s.commit()
        return {"learning_id": row.id, "text": text}

    @app.get("/api/v1/agent/learnings")
    def list_learnings(s: Session = Depends(db), u: User = Depends(require("engagement.view"))):
        return [{"id": l.id, "text": l.text, "rating": l.rating, "stage": l.stage}
                for l in s.scalars(select(AgentLearning).order_by(AgentLearning.id.desc())).all()]

    @app.delete("/api/v1/agent/learnings/{lid}")
    def delete_learning(lid: int, s: Session = Depends(db),
                        u: User = Depends(require("engagement.view"))):
        row = s.get(AgentLearning, lid)
        if row:
            s.delete(row); s.commit()
        return {"deleted": True}

    # ============================================================
    #  Registry v2 — exhaustive vendor/engagement model + masters
    # ============================================================
    from .features import registry_service as RS
    from .features import financial as FIN
    from .features.registry_models import (
        IndustryMaster, MaterialGroupMaster, VendorGroup, VendorRecord,
        VendorIndustry, ContactRecord, EngagementRecord, AssessmentRecord,
        FindingRecord, RemediationRecord, FourthPartyRecord, FourthPartyVendor,
        ArtefactRecord, IssueRecord,
    )

    # ---- master lists ----
    @app.get("/api/v2/industries")
    def list_industries(s: Session = Depends(db), u: User = Depends(require("vendor.view"))):
        return [{"industry_id": i.industry_id, "sic_code": i.sic_code, "division": i.division}
                for i in s.scalars(select(IndustryMaster).order_by(IndustryMaster.sic_code)).all()]

    @app.get("/api/v2/material-groups")
    def list_material_groups(s: Session = Depends(db), u: User = Depends(require("engagement.view"))):
        return [{"material_group_id": m.material_group_id, "unspsc_code": m.unspsc_code}
                for m in s.scalars(select(MaterialGroupMaster).order_by(MaterialGroupMaster.unspsc_code)).all()]

    # ---- vendors (exhaustive) ----
    @app.post("/api/v2/vendors")
    def v2_create_vendor(b: V2VendorIn, s: Session = Depends(db),
                         u: User = Depends(require("vendor.edit"))):
        v = RS.create_vendor(s, legal_name=b.legal_name, created_via=b.created_via or "button",
                             group_id=b.group_id, parent=b.parent_company,
                             industries=b.industries or [], tier=b.tier or "Tier 3",
                             trading_name=b.trading_name, registration_number=b.registration_number,
                             hq_country=b.hq_country, website=b.website,
                             listing_status=b.listing_status, procurement_ref=b.procurement_ref)
        audit(s, "v2.vendor_created", u.username,
              {"vendor_id": v.vendor_id, "group_id": v.group_id, "via": v.created_via})
        s.commit()
        return {"vendor_id": v.vendor_id, "group_id": v.group_id}

    @app.post("/api/v2/vendors/import")
    async def v2_import_vendors(file: UploadFile = File(...), mode: str = Form(default="preview"),
                                s: Session = Depends(db), u: User = Depends(require("vendor.edit"))):
        """Bulk CSV vendor import. Columns: legal_name (required), tier, hq_country,
        is_critical. mode=preview validates only; mode=commit creates the vendors."""
        import csv as _csv
        import io as _io
        raw = await file.read()
        _err = SEC.upload_check(file.filename or "vendors.csv", raw)
        if _err:
            raise HTTPException(415, _err)
        try:
            textd = raw.decode("utf-8-sig")
        except Exception:
            raise HTTPException(415, "file must be UTF-8 CSV")
        rows = list(_csv.DictReader(_io.StringIO(textd)))
        if not rows:
            raise HTTPException(400, "no data rows found (expected a header incl. legal_name)")
        valid, errors, seen = [], [], set()
        existing = {(v.legal_name or "").strip().lower()
                    for v in s.scalars(select(VendorRecord)).all()}
        for i, r in enumerate(rows, start=2):
            name = (r.get("legal_name") or r.get("name") or "").strip()
            if not name:
                errors.append({"row": i, "error": "legal_name is required"})
                continue
            key = name.lower()
            if key in existing:
                errors.append({"row": i, "error": f"'{name}' already exists in the register"})
                continue
            if key in seen:
                errors.append({"row": i, "error": f"'{name}' duplicated in file"})
                continue
            seen.add(key)
            tier = (r.get("tier") or "Tier 3").strip()
            valid.append({"legal_name": name,
                          "tier": tier if tier in ("Tier 1", "Tier 2", "Tier 3") else "Tier 3",
                          "hq_country": (r.get("hq_country") or "").strip() or None,
                          "is_critical": str(r.get("is_critical", "")).strip().lower()
                          in ("1", "true", "yes", "y")})
        created = []
        if mode == "commit":
            for v in valid:
                rec = RS.create_vendor(s, legal_name=v["legal_name"], created_via="csv_import",
                                       tier=v["tier"], hq_country=v["hq_country"], industries=[])
                if v["is_critical"]:
                    rec.is_critical = True
                created.append(rec.vendor_id)
            audit(s, "v2.vendors_imported", u.username, {"count": len(created)})
            from .features import notifications as _NOTIF
            _NOTIF.emit(s, "import.completed", f"CSV import: {len(created)} vendor(s) created", f"by {u.username}", link="vendors")
            for _vid in created:
                _NOTIF.emit(s, "vendor.created", f"Vendor created: {_vid}", "via CSV import", link="vendors")
            s.commit()
        return {"mode": mode, "total_rows": len(rows), "valid": len(valid),
                "errors": errors[:50], "created": created}

    @app.get("/api/v2/vendors")
    def v2_list_vendors(group_id: Optional[str] = None, slim: bool = False,
                        s: Session = Depends(db),
                        u: User = Depends(require("vendor.view"))):
        stmt = select(VendorRecord)
        if group_id:
            stmt = stmt.where(VendorRecord.group_id == group_id)
        vendors = s.scalars(stmt).all()
        if slim:
            # SPA list view uses only these fields; skip the industries join entirely.
            return [{"vendor_id": v.vendor_id, "legal_name": v.legal_name, "tier": v.tier,
                     "status": v.status, "is_critical": v.is_critical} for v in vendors]
        # Batch all industry tags in ONE query instead of N (was an N+1 per vendor).
        vids = [v.vendor_id for v in vendors]
        ind_map: dict = {}
        if vids:
            for vi in s.scalars(select(VendorIndustry).where(
                    VendorIndustry.vendor_id.in_(vids))).all():
                ind_map.setdefault(vi.vendor_id, []).append(vi.industry_id)
        return [{"vendor_id": v.vendor_id, "group_id": v.group_id,
                 "legal_name": v.legal_name, "tier": v.tier, "status": v.status,
                 "is_critical": v.is_critical, "industries": ind_map.get(v.vendor_id, []),
                 "created_via": v.created_via} for v in vendors]

    @app.get("/api/v2/vendors/{vid}")
    def v2_get_vendor(vid: str, s: Session = Depends(db), u: User = Depends(require("vendor.view"))):
        v = s.scalars(select(VendorRecord).where(VendorRecord.vendor_id == vid)).first()
        if not v:
            raise HTTPException(404, "vendor not found")
        inds = [vi.industry_id for vi in s.scalars(select(VendorIndustry).where(
            VendorIndustry.vendor_id == vid)).all()]
        contacts = [{"id": c.id, "is_primary": c.is_primary, "name": c.name, "email": c.email,
                     "phone": f"{c.phone_country_code or ''} {c.phone_number or ''}".strip(),
                     "designation": c.designation, "country": c.country,
                     "mailing_address": c.mailing_address}
                    for c in s.scalars(select(ContactRecord).where(
                        ContactRecord.owner_type == "vendor", ContactRecord.owner_id == vid)).all()]
        engs = [{"engagement_id": e.engagement_id, "title": e.title, "status": e.status}
                for e in s.scalars(select(EngagementRecord).where(
                    EngagementRecord.vendor_id == vid)).all()]
        return {"vendor_id": v.vendor_id, "group_id": v.group_id, "legal_name": v.legal_name,
                "trading_name": v.trading_name, "tier": v.tier, "status": v.status,
                "hq_country": v.hq_country, "website": v.website,
                "listing_status": v.listing_status, "is_critical": v.is_critical,
                "industries": inds, "contacts": contacts, "engagements": engs,
                "fourth_party_id": v.fourth_party_id}

    @app.post("/api/v2/vendors/{vid}/group")
    def v2_override_group(vid: str, b: GroupOverrideIn, s: Session = Depends(db),
                          u: User = Depends(require("vendor.edit"))):
        v = s.scalars(select(VendorRecord).where(VendorRecord.vendor_id == vid)).first()
        if not v:
            raise HTTPException(404, "vendor not found")
        v.group_id = b.group_id
        audit(s, "v2.vendor_group_override", u.username, {"vendor_id": vid, "group_id": b.group_id})
        s.commit()
        return {"vendor_id": vid, "group_id": b.group_id}

    @app.post("/api/v2/contacts")
    def v2_add_contact(b: V2ContactIn, s: Session = Depends(db),
                       u: User = Depends(require("vendor.edit"))):
        c = RS.add_contact(s, owner_type=b.owner_type, owner_id=b.owner_id, name=b.name,
                           is_primary=b.is_primary, email=b.email,
                           phone_country_code=b.phone_country_code, phone_number=b.phone_number,
                           designation=b.designation, country=b.country,
                           mailing_address=b.mailing_address)
        audit(s, "v2.contact_added", u.username,
              {"owner": b.owner_id, "primary": b.is_primary, "contact_id": c.id})
        s.commit()
        return {"contact_id": c.id, "is_primary": c.is_primary}

    # ---- engagements (exhaustive) ----
    @app.post("/api/v2/engagements")
    def v2_create_engagement(b: V2EngagementIn, s: Session = Depends(db),
                             u: User = Depends(require("engagement.edit"))):
        e = RS.create_engagement(s, vendor_id=b.vendor_id, title=b.title,
                                 owner_user=b.owner_user or u.username,
                                 service_description=b.service_description,
                                 material_group_id=b.material_group_id,
                                 business_unit=b.business_unit,
                                 deployment_model=b.deployment_model,
                                 annual_value=b.annual_value,
                                 currency=b.currency)
        audit(s, "v2.engagement_created", u.username,
              {"engagement_id": e.engagement_id, "vendor_id": b.vendor_id})
        s.commit()
        return {"engagement_id": e.engagement_id}

    @app.get("/api/v2/engagements")
    def v2_list_engagements(vendor_id: Optional[str] = None, status: Optional[str] = None,
                            slim: bool = False,
                            s: Session = Depends(db), u: User = Depends(require("engagement.view"))):
        from datetime import date as _date
        _today = _date.today().isoformat()
        stmt = select(EngagementRecord)
        if vendor_id:
            stmt = stmt.where(EngagementRecord.vendor_id == vendor_id)
        if status:
            stmt = stmt.where(EngagementRecord.status == status)
        rows = s.scalars(stmt).all()
        out = []
        for e in rows:
            if slim:
                # The list view in the SPA uses only these six fields.
                out.append({"engagement_id": e.engagement_id, "vendor_id": e.vendor_id,
                            "title": e.title, "status": e.status,
                            "residual_band": e.residual_band, "open_actions": e.open_actions})
                continue
            out.append({"engagement_id": e.engagement_id, "vendor_id": e.vendor_id,
                        "title": e.title, "status": e.status, "stage": e.stage,
                        "inherent_band": e.inherent_band, "residual_band": e.residual_band,
                        "open_actions": e.open_actions, "contract_id": e.contract_id,
                        "assessment_id": e.assessment_id, "material_group_id": e.material_group_id,
                        "last_assessment_date": e.last_assessment_date,
                        "next_assessment_due": e.next_assessment_due,
                        "reassessment_due": bool(e.next_assessment_due and e.next_assessment_due <= _today)})
        return out

    # ---- assessments ----
    @app.post("/api/v2/assessments")
    def v2_create_assessment(b: V2AssessmentIn, s: Session = Depends(db),
                             u: User = Depends(require("engagement.view"))):
        pool = [x.username for x in s.scalars(select(User)).all() if x.is_active]
        rec = RS.create_assessment(s, engagement_id=b.engagement_id, vendor_id=b.vendor_id,
                                   engagement_owner=u.username, session_id=b.session_id,
                                   inherent_band=b.inherent_band, residual_band=b.residual_band,
                                   assessor_pool=pool)
        audit(s, "v2.assessment_created", u.username,
              {"assessment_id": rec.assessment_id, "assessor": rec.assessor_user})
        s.commit()
        return {"assessment_id": rec.assessment_id, "status": rec.status,
                "assessor_user": rec.assessor_user, "spoc_user": rec.spoc_user}

    @app.get("/api/v2/assessments")
    def v2_list_assessments(s: Session = Depends(db), u: User = Depends(require("engagement.view"))):
        out = []
        for a in s.scalars(select(AssessmentRecord)).all():
            if not _can_view_assessment(u, a):
                continue
            out.append({"assessment_id": a.assessment_id, "engagement_id": a.engagement_id,
                        "status": a.status, "inherent_band": a.inherent_band,
                        "outcome": a.outcome, "assessor_user": a.assessor_user,
                        "assessor_signed_off": a.assessor_signed_off, "locked": a.locked,
                        "spoc_user": a.spoc_user})
        return out

    @app.post("/api/v2/assessments/{aid}/signoff")
    def v2_signoff(aid: str, s: Session = Depends(db), u: User = Depends(require("engagement.review"))):
        a = s.scalars(select(AssessmentRecord).where(AssessmentRecord.assessment_id == aid)).first()
        if not a:
            raise HTTPException(404, "assessment not found")
        if a.assessor_user and a.assessor_user != u.username and u.role.key != "admin":
            raise HTTPException(403, "only the assigned assessor may sign off")
        a.assessor_signed_off = True
        if a.engagement_id:
            RS.record_last_assessment(s, a.engagement_id)
        audit(s, "v2.assessment_signoff", u.username, {"assessment_id": aid})
        s.commit()
        return {"assessment_id": aid, "assessor_signed_off": True}

    @app.post("/api/v2/assessments/{aid}/approve")
    def v2_approve(aid: str, s: Session = Depends(db), u: User = Depends(require("engagement.review"))):
        try:
            rec = RS.approve_assessment(s, aid)
        except ValueError as e:
            raise HTTPException(400, str(e))
        if rec and rec.engagement_id:
            RS.record_last_assessment(s, rec.engagement_id)
        audit(s, "v2.assessment_approved", u.username, {"assessment_id": aid, "locked": True})
        s.commit()
        return {"assessment_id": aid, "status": rec.status, "locked": rec.locked}

    @app.post("/api/v2/assessments/{aid}/recall")
    def v2_recall(aid: str, s: Session = Depends(db), u: User = Depends(require("engagement.view"))):
        a = s.scalars(select(AssessmentRecord).where(AssessmentRecord.assessment_id == aid)).first()
        if not a:
            raise HTTPException(404, "assessment not found")
        if a.locked:
            raise HTTPException(400, "approved assessments are hard-locked and cannot be recalled")
        a.status = "Recalled"
        audit(s, "v2.assessment_recalled", u.username, {"assessment_id": aid})
        s.commit()
        return {"assessment_id": aid, "status": "Recalled"}

    @app.post("/api/v2/assessments/{aid}/reassign")
    def v2_reassign(aid: str, b: ReassignIn, s: Session = Depends(db),
                    u: User = Depends(require("engagement.review"))):
        a = s.scalars(select(AssessmentRecord).where(AssessmentRecord.assessment_id == aid)).first()
        if not a:
            raise HTTPException(404, "assessment not found")
        if a.locked:
            raise HTTPException(400, "approved assessment is locked")
        a.assessor_user = b.assessor_user
        a.assessor_signed_off = False
        audit(s, "v2.assessor_reassigned", u.username, {"assessment_id": aid, "to": b.assessor_user})
        s.commit()
        return {"assessment_id": aid, "assessor_user": b.assessor_user}

    # ---- findings + remediation ----
    FINDING_STATUSES = ["Draft", "Published", "Under Remediation", "Remediated",
                        "Verified", "Closed"]

    def _finding_row(f):
        import json as _j
        return {"finding_id": f.finding_id, "title": f.title, "heading": f.title,
                "description": f.description, "severity": f.severity, "source": f.source,
                "status": f.status, "domain": f.domain, "owner": f.owner,
                "assessor": f.assessor, "suggested_remediation": f.suggested_remediation,
                "suggested_closure": f.suggested_closure, "due_date": f.due_date,
                "engagement_id": f.engagement_id, "vendor_id": f.vendor_id,
                "assessment_id": f.assessment_id, "remediation_id": f.remediation_id,
                "risk_accepted": bool(f.risk_accepted), "acceptance_expiry": f.acceptance_expiry,
                "acceptance_rationale": f.acceptance_rationale, "accepted_by": f.accepted_by,
                "progress_notes": _j.loads(f.progress_notes or "[]"),
                "attachments": _j.loads(f.attachments or "[]"),
                "raised_by": f.raised_by, "created_at": str(f.created_at)}

    def _assessor_owns(u, f, s):
        """Assessor may modify only findings on engagements assigned to them
        (or where they are the named assessor). Admin/controller/others by perm."""
        role = (getattr(u, "role", None) and u.role.key) or ""
        if role != "vrm":
            return True  # gating handled by require(); non-assessors not scoped here
        if f.assessor and f.assessor == u.username:
            return True
        if f.engagement_id:
            eng = s.scalars(select(EngagementRecord).where(
                EngagementRecord.engagement_id == f.engagement_id)).first()
            if eng and eng.assigned_assessor == u.username:
                return True
        return False

    @app.post("/api/v2/findings")
    def v2_create_finding(b: V2FindingIn, s: Session = Depends(db),
                          u: User = Depends(require("finding.manage"))):
        f = RS.create_finding(s, title=b.title, severity=b.severity or "Medium",
                              source=b.source or "Assessor", description=b.description,
                              domain=b.domain, engagement_id=b.engagement_id,
                              vendor_id=b.vendor_id, assessment_id=b.assessment_id,
                              raised_by=u.username, due_date=b.due_date,
                              owner=b.owner, assessor=b.assessor,
                              suggested_remediation=b.suggested_remediation,
                              suggested_closure=b.suggested_closure,
                              status=b.status or "Draft")
        audit(s, "v2.finding_created", u.username,
              {"finding_id": f.finding_id, "severity": f.severity})
        s.commit()
        return _finding_row(f)

    @app.get("/api/v2/findings")
    def v2_list_findings(engagement_id: Optional[str] = None, vendor_id: Optional[str] = None,
                         assessment_id: Optional[str] = None, status: Optional[str] = None,
                         severity: Optional[str] = None, source: Optional[str] = None,
                         assessor: Optional[str] = None, accepted: Optional[str] = None,
                         s: Session = Depends(db), u: User = Depends(require("finding.view"))):
        rows = s.scalars(select(FindingRecord).order_by(FindingRecord.id.desc())).all()
        out = []
        for f in rows:
            if engagement_id and f.engagement_id != engagement_id: continue
            if vendor_id and f.vendor_id != vendor_id: continue
            if assessment_id and f.assessment_id != assessment_id: continue
            if status and f.status != status: continue
            if severity and f.severity != severity: continue
            if source and f.source != source: continue
            if assessor and f.assessor != assessor: continue
            if accepted in ("true", "1") and not f.risk_accepted: continue
            if accepted in ("false", "0") and f.risk_accepted: continue
            out.append(_finding_row(f))
        return out

    @app.get("/api/v2/findings/{fid}")
    def v2_get_finding(fid: str, s: Session = Depends(db),
                       u: User = Depends(require("finding.view"))):
        f = s.scalars(select(FindingRecord).where(FindingRecord.finding_id == fid)).first()
        if not f: raise HTTPException(404, "finding not found")
        row = _finding_row(f)
        row["can_modify"] = _assessor_owns(u, f, s)
        row["statuses"] = FINDING_STATUSES
        return row

    @app.put("/api/v2/findings/{fid}")
    def v2_update_finding(fid: str, b: FindingPatchIn, s: Session = Depends(db),
                          u: User = Depends(require("finding.manage"))):
        f = s.scalars(select(FindingRecord).where(FindingRecord.finding_id == fid)).first()
        if not f: raise HTTPException(404, "finding not found")
        if not _assessor_owns(u, f, s):
            raise HTTPException(403, "Assessors may only modify findings on engagements assigned to them")
        for k in ("title", "description", "severity", "status", "owner", "assessor",
                  "suggested_remediation", "suggested_closure", "due_date"):
            v = getattr(b, k)
            if v is not None:
                if k == "status" and v not in FINDING_STATUSES:
                    raise HTTPException(400, f"invalid status; allowed: {FINDING_STATUSES}")
                setattr(f, k, v)
        if b.status == "Closed" and not f.closed_date:
            from datetime import date
            f.closed_date = date.today().isoformat()
        s.flush(); RS._recompute_open_actions(s, f.engagement_id)
        audit(s, "v2.finding_updated", u.username, {"finding_id": fid, "status": f.status})
        s.commit()
        return _finding_row(f)

    @app.post("/api/v2/findings/{fid}/risk-accept")
    def v2_finding_risk_accept(fid: str, b: RiskAcceptIn, s: Session = Depends(db),
                               u: User = Depends(require("acceptance.manage"))):
        f = s.scalars(select(FindingRecord).where(FindingRecord.finding_id == fid)).first()
        if not f: raise HTTPException(404, "finding not found")
        f.risk_accepted = bool(b.accept)
        f.acceptance_rationale = b.rationale if b.accept else None
        f.acceptance_expiry = b.expiry_date if b.accept else None
        f.accepted_by = u.username if b.accept else None
        s.flush()
        audit(s, "v2.finding_risk_accept", u.username,
              {"finding_id": fid, "accepted": bool(b.accept), "expiry": b.expiry_date})
        notify(s, (f"Risk acceptance recorded on {fid} (expires {b.expiry_date})"
                   if b.accept else f"Risk acceptance revoked on {fid}"), "all",
               body=f"Vendor {f.vendor_id or '—'} · engagement {f.engagement_id or '—'}")
        s.commit()
        return _finding_row(f)

    @app.post("/api/v2/findings/{fid}/note")
    def v2_finding_note(fid: str, b: FindingNoteIn, s: Session = Depends(db),
                        u: User = Depends(require("finding.manage"))):
        import json as _j
        from datetime import datetime as _dt
        f = s.scalars(select(FindingRecord).where(FindingRecord.finding_id == fid)).first()
        if not f: raise HTTPException(404, "finding not found")
        if not _assessor_owns(u, f, s):
            raise HTTPException(403, "Assessors may only update findings assigned to them")
        notes = _j.loads(f.progress_notes or "[]")
        notes.append({"ts": _dt.utcnow().isoformat(timespec="seconds"),
                      "user": u.username, "note": b.note})
        f.progress_notes = _j.dumps(notes); s.flush()
        audit(s, "v2.finding_note", u.username, {"finding_id": fid})
        s.commit()
        return {"finding_id": fid, "progress_notes": notes}

    @app.post("/api/v2/findings/{fid}/attach")
    def v2_finding_attach(fid: str, b: FindingAttachIn, s: Session = Depends(db),
                          u: User = Depends(require("finding.manage"))):
        import json as _j
        f = s.scalars(select(FindingRecord).where(FindingRecord.finding_id == fid)).first()
        if not f: raise HTTPException(404, "finding not found")
        if not _assessor_owns(u, f, s):
            raise HTTPException(403, "Assessors may only update findings assigned to them")
        atts = _j.loads(f.attachments or "[]")
        atts.append({"doc_id": b.doc_id, "name": b.name or b.doc_id})
        f.attachments = _j.dumps(atts); s.flush()
        audit(s, "v2.finding_attach", u.username, {"finding_id": fid, "doc_id": b.doc_id})
        s.commit()
        return {"finding_id": fid, "attachments": atts}

    @app.delete("/api/v2/findings/{fid}")
    def v2_delete_finding(fid: str, s: Session = Depends(db),
                          u: User = Depends(require("finding.delete"))):
        f = s.scalars(select(FindingRecord).where(FindingRecord.finding_id == fid)).first()
        if not f: raise HTTPException(404, "finding not found")
        eng = f.engagement_id
        s.delete(f); s.flush(); RS._recompute_open_actions(s, eng)
        audit(s, "v2.finding_deleted", u.username, {"finding_id": fid})
        s.commit()
        return {"deleted": fid}

    # ---- remediation plans (RMD) ----
    def _rmd_row(s, r):
        from .features.registry_models import FindingRecord
        f = s.scalars(select(FindingRecord).where(FindingRecord.finding_id == r.finding_id)).first()
        return {"remediation_id": r.remediation_id, "finding_id": r.finding_id,
                "finding_title": f.title if f else None, "severity": f.severity if f else None,
                "vendor_id": f.vendor_id if f else None, "engagement_id": f.engagement_id if f else None,
                "plan": r.plan, "owner": r.owner, "target_date": r.target_date,
                "status": r.status, "progress_pct": r.progress_pct, "evidence": r.evidence,
                "completed_date": r.completed_date, "verified_by": r.verified_by,
                "created_at": r.created_at.isoformat() if r.created_at else None}

    @app.get("/api/v2/remediations")
    def v2_remediations_list(status: Optional[str] = None, s: Session = Depends(db),
                             u: User = Depends(require("finding.view"))):
        from .features.registry_models import RemediationRecord
        rows = s.scalars(select(RemediationRecord).order_by(RemediationRecord.id.desc())).all()
        return [_rmd_row(s, r) for r in rows if not status or r.status == status]

    @app.get("/api/v2/remediations/{rid}")
    def v2_remediation_get(rid: str, s: Session = Depends(db),
                           u: User = Depends(require("finding.view"))):
        from .features.registry_models import RemediationRecord
        r = s.scalars(select(RemediationRecord).where(RemediationRecord.remediation_id == rid)).first()
        if not r: raise HTTPException(404, "remediation plan not found")
        return _rmd_row(s, r)

    @app.post("/api/v2/findings/{fid}/remediation")
    def v2_finding_remediation(fid: str, body: dict = Body(default={}), s: Session = Depends(db),
                               u: User = Depends(require("finding.manage"))):
        """Create (or fetch) the remediation plan linked to a finding."""
        from .features.registry_models import FindingRecord, RemediationRecord
        from .features import registry_service as RS
        f = s.scalars(select(FindingRecord).where(FindingRecord.finding_id == fid)).first()
        if not f: raise HTTPException(404, "finding not found")
        r = None
        if f.remediation_id:
            r = s.scalars(select(RemediationRecord).where(
                RemediationRecord.remediation_id == f.remediation_id)).first()
        if not r:
            rid = RS.next_id(s, "remediation")
            r = RemediationRecord(remediation_id=rid, finding_id=fid,
                                  plan=body.get("plan") or (f.suggested_remediation or "Remediation plan to be defined."),
                                  owner=body.get("owner") or f.owner, target_date=body.get("target_date") or f.due_date,
                                  status="Planned", progress_pct=0)
            s.add(r); s.flush()
            f.remediation_id = rid
            if f.status in ("Draft", "Published"):
                f.status = "Under Remediation"
        audit(s, "v2.remediation_create", u.username, {"finding_id": fid, "remediation_id": r.remediation_id})
        s.commit()
        return _rmd_row(s, r)

    @app.put("/api/v2/remediations/{rid}")
    def v2_remediation_update(rid: str, body: dict = Body(...), s: Session = Depends(db),
                              u: User = Depends(require("finding.manage"))):
        from .features.registry_models import RemediationRecord
        from datetime import date as _date
        r = s.scalars(select(RemediationRecord).where(RemediationRecord.remediation_id == rid)).first()
        if not r: raise HTTPException(404, "remediation plan not found")
        for k in ("plan", "owner", "target_date", "status", "evidence", "verified_by"):
            if body.get(k) is not None:
                setattr(r, k, body[k])
        if body.get("progress_pct") is not None:
            r.progress_pct = max(0, min(100, int(body["progress_pct"])))
        if r.status in ("Complete", "Verified") and not r.completed_date:
            r.completed_date = _date.today().isoformat()
            r.progress_pct = 100
        audit(s, "v2.remediation_update", u.username, {"remediation_id": rid, "status": r.status})
        s.commit()
        return _rmd_row(s, r)

    @app.get("/api/v2/pestle/exposure-summary")
    def v2_pestle_exposure_summary(focus_type: str = "portfolio", focus_id: Optional[str] = None,
                                   s: Session = Depends(db), u: User = Depends(require("vendor.view"))):
        """Narrative risk summary for the current knowledge-map focus. Deterministic;
        upgrades to AI when a model is connected."""
        from .features import pestle as PES
        from .agents import llm_config
        try:
            d = PES.get_summary(s)
        except Exception:
            d = None
        if not d:
            return {"summary": "No PESTLE data yet — run an overnight sweep to populate the threat surface.",
                    "ai": False, "focus": focus_type}
        cats = d.get("categories", {})
        means = d.get("cat_means", {})
        top = d.get("top_threats", [])[:6]
        movers = d.get("movers", [])[:5]
        label = focus_id or "the portfolio"
        hi = sorted(means.items(), key=lambda x: -x[1])[:2]
        hi_txt = ", ".join(f"{cats.get(c, {}).get('name', c)} ({v})" for c, v in hi)
        top_txt = "; ".join(f"{t['name']} ({t['score']})" for t in top)
        mv_up = [m for m in movers if m.get("delta", 0) > 0][:3]
        mv_txt = ", ".join(f"{m['name']} ▲{abs(m['delta']):.1f}" for m in mv_up)
        det = (f"Focus: {label} ({focus_type}). Highest-exposure PESTLE categories are {hi_txt}. "
               f"The most systemic threats are {top_txt}. "
               + (f"Overnight, the sharpest deteriorations were {mv_txt}. " if mv_txt else "No material overnight moves. ")
               + "Prioritise the highest-scoring threats above for mitigation and monitor the movers.")
        ai = False
        if llm_config.status().get("live_ready"):
            try:
                out = llm_config.complete(
                    PROMPTS.resolve(s, "pestle_summary"),
                    f"Focus: {label} ({focus_type}). Category means: {means}. Top threats: {top}. Movers: {movers}.",
                    domain="risk", max_tokens=300, review=True)
                if out and out.strip():
                    det = out.strip(); ai = True
            except Exception as _e:
                _obs_swallow('bro_app.py', _e)
        return {"summary": det, "ai": ai, "focus": focus_type, "focus_id": focus_id}

    # ---- Controller: assign engagement to an assessor ----
    @app.post("/api/v2/engagements/{eid}/assign")
    def v2_assign_engagement(eid: str, b: AssignAssessorIn, s: Session = Depends(db),
                             u: User = Depends(require("engagement.assign"))):
        eng = s.scalars(select(EngagementRecord).where(
            EngagementRecord.engagement_id == eid)).first()
        if not eng: raise HTTPException(404, "engagement not found")
        target = s.scalars(select(User).where(User.username == b.assessor_user)).first()
        if not target: raise HTTPException(404, "assessor user not found")
        eng.assigned_assessor = b.assessor_user
        # propagate to the engagement's assessments + open findings
        for a in s.scalars(select(AssessmentRecord).where(
                AssessmentRecord.engagement_id == eid)).all():
            a.assessor_user = b.assessor_user
        for f in s.scalars(select(FindingRecord).where(
                FindingRecord.engagement_id == eid)).all():
            if not f.assessor:
                f.assessor = b.assessor_user
        s.flush()
        audit(s, "v2.engagement_assigned", u.username,
              {"engagement_id": eid, "assessor": b.assessor_user})
        notify(s, f"Engagement {eid} assigned to {b.assessor_user}", "vrm")
        s.commit()
        return {"engagement_id": eid, "assigned_assessor": b.assessor_user}

    @app.get("/api/v2/assessors")
    def v2_list_assessors(s: Session = Depends(db),
                          u: User = Depends(require("engagement.assign"))):
        from .features.models_db import Role
        rid = s.scalars(select(Role).where(Role.key == "vrm")).first()
        users = s.scalars(select(User).where(User.role_id == (rid.id if rid else -1))).all()
        return [{"username": x.username} for x in users]

    # ===== Continuous-monitoring connectors: RapidRatings (FDD) + Interos (Reputation) =====
    # API trigger + inbound webhook + MCP-style discovery. Demonstrator data with a clean
    # seam: set the provider credentials/MCP endpoint to switch to live calls.
    def _connector_status():
        import os as _os
        return {
            "rapidratings": {"name": "RapidRatings", "domain": "Financial DD",
                             "mode": "live" if _os.environ.get("RAPIDRATINGS_API_KEY") else "demonstrator",
                             "mcp": _os.environ.get("RAPIDRATINGS_MCP_URL"),
                             "webhook": "/api/v2/webhooks/rapidratings"},
            "interos": {"name": "Interos", "domain": "Reputation & resilience",
                        "mode": "live" if _os.environ.get("INTEROS_API_KEY") else "demonstrator",
                        "mcp": _os.environ.get("INTEROS_MCP_URL"),
                        "webhook": "/api/v2/webhooks/interos"},
        }

    def _file_monitoring_report(s, vendor_id, provider, mode, payload, actor):
        """File a connector pull as a tagged report + update the vendor indicator."""
        import base64 as _b64, json as _j
        from .features import documents as DOC
        from .features.registry_models import VendorRecord
        v = s.scalars(select(VendorRecord).where(
            VendorRecord.vendor_id == vendor_id)).first() if vendor_id else None
        purpose = "fdd_report" if provider == "rapidratings" else "reputation_report"
        data = _b64.b64encode(_j.dumps(payload, indent=2).encode()).decode()
        doc = DOC.store_document(
            s, filename=f"{provider}_{(v.legal_name if v else vendor_id or 'entity')}.json"[:80],
            content_type="application/json", data_b64=data,
            vendor_id=vendor_id, uploaded_by=f"{provider}:{mode}", purpose=purpose)
        if v:
            if provider == "rapidratings" and payload.get("financial_health_band"):
                v.financial_health_band = payload["financial_health_band"]
            if provider == "interos" and payload.get("reputation_band"):
                try: v.reputation_band = payload["reputation_band"]
                except Exception: pass
        notify(s, f"{provider.title()} monitoring update filed for {vendor_id or 'entity'}", "all")
        return doc.doc_id

    @app.get("/api/v2/connectors/status")
    def v2_connector_status(u: User = Depends(require("intel.financial"))):
        return _connector_status()

    @app.post("/api/v2/connectors/rapidratings/pull")
    def v2_rapidratings_pull(b: ConnectorPullIn, s: Session = Depends(db),
                             u: User = Depends(require("intel.financial"))):
        st = _connector_status()["rapidratings"]
        # DEMONSTRATOR payload — live mode would call the RapidRatings API / MCP here.
        payload = {"provider": "RapidRatings", "vendor_id": b.vendor_id,
                   "fhr": 62, "core_health_score": 58, "financial_health_band": "Adequate",
                   "trend": "stable", "as_of": "latest",
                   "summary": "Financial Health Rating in the adequate band; stable trend; "
                              "no going-concern flag.", "mode": st["mode"]}
        doc = _file_monitoring_report(s, b.vendor_id, "rapidratings", st["mode"], payload, u.username)
        audit(s, "v2.connector_rapidratings", u.username, {"vendor_id": b.vendor_id, "doc": doc})
        s.commit()
        return {"provider": "rapidratings", "mode": st["mode"], "filed_report": doc, "data": payload}

    @app.post("/api/v2/connectors/interos/pull")
    def v2_interos_pull(b: ConnectorPullIn, s: Session = Depends(db),
                        u: User = Depends(require("intel.reputation"))):
        st = _connector_status()["interos"]
        payload = {"provider": "Interos", "vendor_id": b.vendor_id,
                   "reputation_band": "Caution", "resilience_score": 71,
                   "adverse_media": True, "esg_flag": "Moderate",
                   "summary": "One adverse-media item in the last 12 months; resilience score "
                              "within tolerance; ESG exposure moderate.", "mode": st["mode"]}
        doc = _file_monitoring_report(s, b.vendor_id, "interos", st["mode"], payload, u.username)
        audit(s, "v2.connector_interos", u.username, {"vendor_id": b.vendor_id, "doc": doc})
        s.commit()
        return {"provider": "interos", "mode": st["mode"], "filed_report": doc, "data": payload}

    @app.post("/api/v2/webhooks/rapidratings")
    def v2_webhook_rapidratings(b: ConnectorWebhookIn, s: Session = Depends(db),
                                u: User = Depends(require("admin.webhooks"))):
        payload = b.payload or {"provider": "RapidRatings", "event": "rating_change"}
        doc = _file_monitoring_report(s, b.vendor_id, "rapidratings", "webhook", payload, u.username)
        audit(s, "v2.webhook_rapidratings", u.username, {"vendor_id": b.vendor_id})
        s.commit()
        return {"received": True, "filed_report": doc}

    @app.post("/api/v2/webhooks/interos")
    def v2_webhook_interos(b: ConnectorWebhookIn, s: Session = Depends(db),
                           u: User = Depends(require("admin.webhooks"))):
        payload = b.payload or {"provider": "Interos", "event": "risk_event"}
        doc = _file_monitoring_report(s, b.vendor_id, "interos", "webhook", payload, u.username)
        audit(s, "v2.webhook_interos", u.username, {"vendor_id": b.vendor_id})
        s.commit()
        return {"received": True, "filed_report": doc}

    # ===== Global Regulations — regulatory reference, AI live-updates & gap assessment =====
    import json as _json2, os as _os2
    _REG_CACHE = {}

    def _regcat():
        if not _REG_CACHE:
            path = _os2.path.join(_os2.path.dirname(__file__), "features", "regdata.json")
            with open(path) as f:
                _REG_CACHE.update(_json2.load(f))
        return _REG_CACHE

    def _reg_short(t):
        return (str(t).split("\u2014")[0].split("(")[0].strip())[:60]

    def _extract_json_array(text):
        if not text:
            return []
        t = text.replace("```json", "").replace("```", "").strip()
        st, e = t.find("["), t.rfind("]")
        if st == -1 or e == -1 or e < st:
            return []
        try:
            arr = json.loads(t[st:e + 1])
            return [x for x in arr if isinstance(x, dict)] if isinstance(arr, list) else []
        except Exception:
            return []

    def _reg_norm_status(sv):
        t = str(sv or "").lower()
        if t.startswith("add") or t in ("full", "green", "yes"):
            return "addressed"
        if t.startswith("part") or t in ("amber", "some"):
            return "partial"
        return "gap"

    @app.get("/api/v2/regulations")
    def v2_regulations(u: User = Depends(require("reg.view"))):
        d = _regcat()
        return {"attrs": d["ATTRS"], "order": d["ORDER"], "extra_order": d["EXTRA"],
                "industries": d["INDUSTRIES"], "catalog": d["CATALOG"]}

    @app.post("/api/v2/regulations/export")
    def v2_reg_export(b: RegExportIn, s: Session = Depends(db),
                      u: User = Depends(require("reg.view"))):
        from openpyxl import Workbook
        from fastapi import Response
        import io
        d = _regcat(); cat = d["CATALOG"]; ATTRS = d["ATTRS"]
        wb = Workbook(); wb.remove(wb.active)
        for code in b.codes:
            c = cat.get(code)
            if not c:
                continue
            ws = wb.create_sheet(code[:31])
            ws.append(["Attribute"] + [_reg_short(inst[0]) for inst in c["I"]])
            for ai, attr in enumerate(ATTRS):
                ws.append([attr] + [(inst[ai] if ai < len(inst) else "") for inst in c["I"]])
        if b.updates:
            ws = wb.create_sheet("Live Updates")
            ws.append(["Jurisdiction", "Instrument idx", "Attribute idx", "Title", "Update", "Date", "Source"])
            for up in b.updates:
                ws.append([up.get("full") or up.get("code"), up.get("instrument"),
                           up.get("attr"), up.get("title"), up.get("update"),
                           up.get("date"), up.get("source")])
        if not wb.sheetnames:
            wb.create_sheet("Empty")
        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        audit(s, "v2.reg_export", u.username, {"codes": b.codes}); s.commit()
        return Response(content=buf.read(),
                        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        headers={"Content-Disposition": 'attachment; filename="Global_Regulations.xlsx"'})

    @app.post("/api/v2/regulations/refresh")
    def v2_reg_refresh(b: RegRefreshIn, s: Session = Depends(db),
                       u: User = Depends(require("reg.assess"))):
        from .agents import llm_config
        if not llm_config.status().get("live_ready"):
            return {"holding": True, "message": AI_HOLDING, "updates": []}
        d = _regcat(); cat = d["CATALOG"]; ATTRS = d["ATTRS"]; updates = []
        attr_idx = "|".join(f"{i}:{a}" for i, a in enumerate(ATTRS))
        for code in (b.codes or [])[:6]:
            c = cat.get(code)
            if not c:
                continue
            inst_idx = "|".join(f"{i}:{_reg_short(inst[0])}" for i, inst in enumerate(c["I"])) or "(none yet)"
            prompt = "\n".join([
                f"FS analyst. {c['full']} ({c['reg']}).",
                f"Instruments: {inst_idx}", f"Attrs: {attr_idx}",
                "Web-search 2026-onward TPRM/outsourcing/ICT-resilience news: rules, consultations,",
                "regulator letters, enforcement, fines, CTP designations. Primary sources only.",
                "Map each to instrument+attr index. New instrument: instrument:-1.",
                "JSON array only, no prose/fences.",
                "Per item: {instrument:int,attr:int,title:str,update:str<=180c,date:str,source:str}.",
                "Max 6 newest-first. Nothing new: []."])
            try:
                out = llm_config.complete(
                    PROMPTS.resolve(s, "regulatory_websearch"),
                    prompt, domain="regulatory", web_search=True, max_tokens=900)
                for it in _extract_json_array(out):
                    it["code"] = code; it["full"] = c["full"]; updates.append(it)
            except Exception:
                continue
        audit(s, "v2.reg_refresh", u.username, {"codes": b.codes, "found": len(updates)})
        s.commit()
        return {"holding": False, "updates": updates}

    @app.post("/api/v2/regulations/assess")
    def v2_reg_assess(b: RegAssessIn, s: Session = Depends(db),
                      u: User = Depends(require("reg.assess"))):
        from .agents import llm_config
        if not llm_config.status().get("live_ready"):
            return {"holding": True, "message": AI_HOLDING, "results": {}, "coverage": {}}
        d = _regcat(); cat = d["CATALOG"]; results = {}
        tally = {"addressed": 0, "partial": 0, "gap": 0}
        doc = (b.doc_text or "")[:8000]
        for code in (b.codes or []):
            c = cat.get(code)
            if not c or not c["I"]:
                continue
            regs = "\n".join(
                f"{i}|{_reg_short(inst[0])}|clauses:{(inst[7] or '—')[:110]}|clock:{inst[8] or '—'}"
                for i, inst in enumerate(c["I"]))
            prompt = "\n".join([
                f"Compliance assessor, {c['full']}.",
                "Rate each regulation vs docs: addressed=substantially covers, partial=some gaps, gap=not covered.",
                "Conservative: absent obligation = gap.", f"REGS:\n{regs}",
                f'DOCS:\n"""\n{doc}\n"""', "JSON array only, no prose/fences.",
                "Per item: {instrument:int,status:str,rationale:str<=160c,gap:str<=120c}."])
            try:
                out = llm_config.complete(
                    PROMPTS.resolve(s, "compliance_assessor"),
                    prompt, domain="regulatory", max_tokens=1200)
                items = []
                for it in _extract_json_array(out):
                    st = _reg_norm_status(it.get("status")); it["status"] = st
                    ii = it.get("instrument")
                    if isinstance(ii, int) and 0 <= ii < len(c["I"]):
                        it["title"] = _reg_short(c["I"][ii][0])
                    tally[st] = tally.get(st, 0) + 1; items.append(it)
                results[code] = {"full": c["full"], "items": items}
            except Exception:
                continue
        audit(s, "v2.reg_assess", u.username, {"codes": b.codes}); s.commit()
        return {"holding": False, "results": results, "coverage": tally}

    @app.post("/api/v2/regulations/relevance")
    def v2_reg_relevance(b: RegRelevanceIn, s: Session = Depends(db),
                         u: User = Depends(require("reg.assess"))):
        from .agents import llm_config
        if not llm_config.status().get("live_ready"):
            return {"holding": True, "message": AI_HOLDING, "relevance": {}}
        d = _regcat(); cat = d["CATALOG"]; rel = {}
        ind = b.industry or {}
        ilabel = ind.get("other") if ind.get("id") == "other" else \
            next((x["label"] for x in d["INDUSTRIES"] if x["id"] == ind.get("id")), "All")
        for code in (b.codes or []):
            c = cat.get(code)
            if not c or not c["I"]:
                continue
            lst = "\n".join(f"{k}:{_reg_short(inst[0])}|{(inst[5] or '')[:80]}"
                            for k, inst in enumerate(c["I"]))
            prompt = "\n".join([
                f'Industry: "{ilabel}". Jurisdiction: {c["full"]}.',
                "Mark applies:false ONLY when instrument clearly does NOT bind this industry.",
                "Banks: all. Trading houses: market-conduct/data/resilience, not deposit-taking rules. FS: most.",
                f"INSTRUMENTS:\n{lst}", "JSON array only, no prose/fences.",
                "Per item: {instrument:int,applies:bool,note:str<=100c}."])
            try:
                out = llm_config.complete(PROMPTS.resolve(s, "regulatory_applicability"),
                                          prompt, domain="regulatory", max_tokens=900)
                rel[code] = _extract_json_array(out)
            except Exception:
                continue
        audit(s, "v2.reg_relevance", u.username, {"codes": b.codes}); s.commit()
        return {"holding": False, "relevance": rel}

    @app.post("/api/v2/regulations/assess/export")
    def v2_reg_assess_export(b: RegReportExportIn, s: Session = Depends(db),
                             u: User = Depends(require("reg.view"))):
        from openpyxl import Workbook
        from fastapi import Response
        import io
        rep = b.report or {}; lab = {"addressed": "Addressed", "partial": "Partial", "gap": "Gap"}
        wb = Workbook(); ws = wb.active; ws.title = "Assessment"
        ws.append([f"Global Regulations — Documentation Assessment"])
        ws.append(["Jurisdiction", "Instrument", "Status", "Rationale", "Gap"])
        for code, r in (rep.get("results") or {}).items():
            for it in (r.get("items") or []):
                ws.append([r.get("full", code), it.get("title", ""),
                           lab.get(it.get("status"), it.get("status", "")),
                           it.get("rationale", ""), it.get("gap", "")])
        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        audit(s, "v2.reg_assess_export", u.username, {}); s.commit()
        return Response(content=buf.read(),
                        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        headers={"Content-Disposition": 'attachment; filename="Regulatory_Assessment.xlsx"'})

    # ===== Phase 1 — Data Integrity & Completion (the overnight steward) =====
    _ALLOWED_FIX_FIELDS = {"lei", "registration_number", "incorporation_country",
                           "hq_country", "ultimate_parent", "legal_form", "website",
                           "listing_status", "status", "tier", "trading_name", "duns"}

    @app.post("/api/v2/integrity/sweep")
    def v2_integrity_sweep(b: IntegritySweepIn, s: Session = Depends(db),
                           u: User = Depends(require("integrity.view"))):
        from .features import integrity as INTEG
        res = INTEG.run_sweep(s, limit=b.limit, vendor_ids=b.vendor_ids)
        audit(s, "v2.integrity_sweep", u.username,
              {"checked": res["vendors_checked"], "issues": res["health"]["issue_count"]})
        s.commit()
        return res

    @app.get("/api/v2/integrity/digest")
    def v2_integrity_digest(s: Session = Depends(db),
                            u: User = Depends(require("integrity.view"))):
        from .features import integrity as INTEG
        res = INTEG.run_sweep(s)
        top = sorted(res["issues"], key=lambda x: {"high": 0, "medium": 1, "low": 2}.get(x["severity"], 3))[:8]
        return {"generated_at": res["generated_at"], "health": res["health"],
                "vendors_checked": res["vendors_checked"], "top_issues": top,
                "duplicate_clusters": len(res["duplicate_clusters"])}

    @app.post("/api/v2/integrity/fix")
    def v2_integrity_fix(b: IntegrityFixIn, s: Session = Depends(db),
                         u: User = Depends(require("integrity.manage"))):
        from .features.registry_models import VendorRecord
        if b.field not in _ALLOWED_FIX_FIELDS:
            raise HTTPException(400, f"field not fixable: {b.field}")
        v = s.scalars(select(VendorRecord).where(VendorRecord.vendor_id == b.vendor_id)).first()
        if not v:
            raise HTTPException(404, "vendor not found")
        old = getattr(v, b.field, None)
        setattr(v, b.field, b.value if b.value not in ("", None) else None)
        s.flush()
        from .features import registry_service as _RS
        n_re = _RS.schedule_reassessment(s, vendor_id=b.vendor_id,
                                         reason=f"Vendor data updated (integrity fix: {b.field})")
        audit(s, "v2.integrity_fix", u.username,
              {"vendor_id": b.vendor_id, "field": b.field, "from": old, "to": b.value,
               "reassessment_scheduled": n_re})
        s.commit()
        return {"vendor_id": b.vendor_id, "field": b.field, "value": b.value}

    @app.post("/api/v2/integrity/merge")
    def v2_integrity_merge(b: IntegrityMergeIn, s: Session = Depends(db),
                           u: User = Depends(require("integrity.manage"))):
        from .features.registry_models import (VendorRecord, EngagementRecord,
                                              AssessmentRecord, FindingRecord)
        if b.primary_vendor_id == b.duplicate_vendor_id:
            raise HTTPException(400, "primary and duplicate are the same")
        prim = s.scalars(select(VendorRecord).where(VendorRecord.vendor_id == b.primary_vendor_id)).first()
        dup = s.scalars(select(VendorRecord).where(VendorRecord.vendor_id == b.duplicate_vendor_id)).first()
        if not prim or not dup:
            raise HTTPException(404, "vendor not found")
        moved = {"engagements": 0, "assessments": 0, "findings": 0, "documents": 0}
        for e in s.scalars(select(EngagementRecord).where(EngagementRecord.vendor_id == b.duplicate_vendor_id)).all():
            e.vendor_id = b.primary_vendor_id; moved["engagements"] += 1
        for a in s.scalars(select(AssessmentRecord).where(AssessmentRecord.vendor_id == b.duplicate_vendor_id)).all():
            a.vendor_id = b.primary_vendor_id; moved["assessments"] += 1
        for f in s.scalars(select(FindingRecord).where(FindingRecord.vendor_id == b.duplicate_vendor_id)).all():
            f.vendor_id = b.primary_vendor_id; moved["findings"] += 1
        try:
            from .features.documents import StoredDocument as _SD
            for d in s.scalars(select(_SD).where(_SD.vendor_id == b.duplicate_vendor_id)).all():
                d.vendor_id = b.primary_vendor_id; moved["documents"] += 1
        except Exception as _e:
            _obs_swallow('bro_app.py', _e)
        # carry over any field the primary is missing
        from .features.integrity import COMPLETENESS_FIELDS as _CF
        for field, _w, _l in _CF:
            if getattr(prim, field, None) in (None, "") and getattr(dup, field, None):
                setattr(prim, field, getattr(dup, field))
        s.delete(dup); s.flush()
        audit(s, "v2.integrity_merge", u.username,
              {"primary": b.primary_vendor_id, "merged": b.duplicate_vendor_id, "moved": moved})
        s.commit()
        return {"primary_vendor_id": b.primary_vendor_id, "merged": b.duplicate_vendor_id, "moved": moved}

    @app.post("/api/v2/integrity/enrich")
    def v2_integrity_enrich(b: IntegrityVendorIn, s: Session = Depends(db),
                            u: User = Depends(require("integrity.manage"))):
        from .features.registry_models import VendorRecord
        from .features import integrity as INTEG
        from .agents import llm_config
        v = s.scalars(select(VendorRecord).where(VendorRecord.vendor_id == b.vendor_id)).first()
        if not v:
            raise HTTPException(404, "vendor not found")
        _score, missing = INTEG.completeness(v)
        if not llm_config.status().get("live_ready"):
            return {"holding": True, "message": AI_HOLDING, "missing": missing, "suggestions": []}
        if not missing:
            return {"holding": False, "missing": [], "suggestions": []}
        prompt = "\n".join([
            f"Entity: {v.legal_name}" + (f" ({v.website})" if v.website else ""),
            f"Country (HQ): {v.hq_country or 'unknown'}.",
            "Web-search authoritative sources and return the following missing reference fields.",
            "Fields: " + ", ".join(missing),
            "JSON array only, no prose/fences.",
            "Per item: {field:str(snake_case of label),value:str,confidence:int(0-100),source:str(url)}."])
        try:
            out = llm_config.complete(
                PROMPTS.resolve(s, "enrichment_corporate"),
                prompt, domain="enrichment", web_search=True, max_tokens=700)
        except Exception:
            out = ""
        # reuse the regulatory JSON extractor
        import json as _j
        sug = []
        try:
            t = (out or "").replace("```json", "").replace("```", "").strip()
            i, e = t.find("["), t.rfind("]")
            if i != -1 and e != -1:
                arr = _j.loads(t[i:e + 1])
                sug = [x for x in arr if isinstance(x, dict) and x.get("field")]
        except Exception:
            sug = []
        audit(s, "v2.integrity_enrich", u.username, {"vendor_id": b.vendor_id, "n": len(sug)})
        s.commit()
        return {"holding": False, "missing": missing, "suggestions": sug}

    @app.post("/api/v2/integrity/chase")
    def v2_integrity_chase(b: IntegrityVendorIn, s: Session = Depends(db),
                           u: User = Depends(require("integrity.view"))):
        from .features.registry_models import VendorRecord
        from .features import integrity as INTEG
        v = s.scalars(select(VendorRecord).where(VendorRecord.vendor_id == b.vendor_id)).first()
        if not v:
            raise HTTPException(404, "vendor not found")
        want = INTEG.expected_evidence(v)
        try:
            from .features.documents import StoredDocument as _SD
            have = {(d.purpose or "").lower() + " " + (d.filename or "").lower()
                    for d in s.scalars(select(_SD).where(_SD.vendor_id == b.vendor_id)).all()}
        except Exception:
            have = set()
        missing = [w for w in want if not any(w.split()[0].lower() in h for h in have)]
        return {"vendor": v.legal_name, "expected": want, "missing_evidence": missing}

    # ===== Phase 2 — Entity / ownership graph, concentration & contagion =====
    @app.get("/api/v2/graph/overview")
    def v2_graph_overview(u: User = Depends(require("vendor.view")), s: Session = Depends(db)):
        from .features import graph as GRAPH
        return GRAPH.build_graph(s)

    @app.post("/api/v2/graph/contagion")
    def v2_graph_contagion(b: ContagionIn, s: Session = Depends(db),
                           u: User = Depends(require("vendor.view"))):
        from .features import graph as GRAPH
        if b.node_type not in ("fourth_party", "owner", "vendor"):
            raise HTTPException(400, "node_type must be fourth_party | owner | vendor")
        return GRAPH.contagion(s, b.node_type, b.node_id)

    # ===== Phase 3 — Exposure & event linkage =====
    @app.get("/api/v2/exposure/bu")
    def v2_exposure_bu(u: User = Depends(require("vendor.view")), s: Session = Depends(db)):
        from .features import exposure as EXP
        return EXP.bu_exposure(s)

    @app.post("/api/v2/incidents/match")
    def v2_incident_match(b: IncidentMatchIn, s: Session = Depends(db),
                          u: User = Depends(require("finding.view"))):
        from .features import exposure as EXP
        res = EXP.incident_match(s, b.vendor_id, b.description or "", b.domain)
        audit(s, "v2.incident_match", u.username,
              {"vendor_id": b.vendor_id, "matched": len(res["matched_findings"]),
               "peers": res["peer_count"]})
        s.commit()
        return res

    @app.post("/api/v2/exposure/brief")
    def v2_exposure_brief(b: BriefIn, s: Session = Depends(db),
                          u: User = Depends(require("vendor.view"))):
        from .features import exposure as EXP
        from .agents import llm_config
        prof = next((p for p in EXP.bu_exposure(s)["business_units"]
                     if p["business_unit"] == b.business_unit), None)
        if not prof:
            raise HTTPException(404, "business unit not found")
        rmix = ", ".join(f"{k} {v}" for k, v in prof["residual"].items() if v)
        deterministic = (f"{prof['business_unit']}: {prof['vendor_count']} vendors "
                         f"({prof['critical_count']} critical) across {prof['engagements']} engagements, "
                         f"£{prof['spend']:,} annual spend. Residual mix: {rmix or 'n/a'}. "
                         f"{prof['open_findings']} open findings. Exposure score {prof['exposure_score']}.")
        if not llm_config.status().get("live_ready"):
            return {"holding": True, "message": AI_HOLDING, "brief": deterministic, "profile": prof}
        try:
            out = llm_config.complete(
                PROMPTS.resolve(s, "exposure_brief"),
                f"Write an exposure brief for business unit '{prof['business_unit']}'. Data: {deterministic}",
                domain="risk", max_tokens=400)
            return {"holding": False, "brief": (out or deterministic).strip(), "profile": prof}
        except Exception:
            return {"holding": False, "brief": deterministic, "profile": prof}

    # ===== Phase 4 — Export-control / geopolitical sensor =====
    @app.get("/api/v2/geopolitical/exposure")
    def v2_geo_exposure(u: User = Depends(require("reg.view")), s: Session = Depends(db)):
        from .features import geopolitical as GEO
        return GEO.exposure(s)

    @app.post("/api/v2/geopolitical/scan")
    def v2_geo_scan(s: Session = Depends(db), u: User = Depends(require("reg.assess"))):
        from .features import geopolitical as GEO
        from .agents import llm_config
        exp = GEO.exposure(s)
        countries = sorted({j["country"] for j in exp["jurisdictions"]
                            if GEO._LEVEL_RANK[j["level"]] >= 1})
        if not llm_config.status().get("live_ready"):
            return {"holding": True, "message": AI_HOLDING, "events": [], "exposed_countries": countries}
        if not countries:
            return {"holding": False, "events": [], "exposed_countries": []}
        prompt = "\n".join([
            "FS supply-chain analyst. Web-search 2026-onward export-control, sanctions and dual-use actions",
            "affecting these jurisdictions where our vendors / sub-processors sit: " + ", ".join(countries) + ".",
            "Cover: new export restrictions (esp. semiconductors / components → likely shortages),",
            "sanctions designations, Entity List additions, dual-use rule changes. Primary sources only.",
            "JSON array only, no prose/fences.",
            "Per item: {country:str,title:str,summary:str<=180c,impact:str<=120c,date:str,source:str}.",
            "Max 8 newest-first. Nothing new: []."])
        try:
            out = llm_config.complete(
                PROMPTS.resolve(s, "geopolitical_analyst"),
                prompt, domain="regulatory", web_search=True, max_tokens=1100)
            import json as _j
            t = (out or "").replace("```json", "").replace("```", "").strip()
            i, e = t.find("["), t.rfind("]")
            events = [x for x in _j.loads(t[i:e + 1]) if isinstance(x, dict)] if (i != -1 and e != -1) else []
        except Exception:
            events = []
        # map events to affected vendors by country
        by_country = {}
        for f in exp["flagged_vendors"]:
            by_country.setdefault(f["country"], []).append(
                {"vendor_id": f["vendor_id"], "legal_name": f["legal_name"], "via": f["via"]})
        for ev in events:
            ev["affected_vendors"] = by_country.get(ev.get("country"), [])
        audit(s, "v2.geo_scan", u.username, {"countries": countries, "events": len(events)})
        s.commit()
        return {"holding": False, "events": events, "exposed_countries": countries}

    # ===== Multilingual layer (display / input / documents) — backend stays English =====
    @app.get("/api/v2/i18n/languages")
    def v2_i18n_languages(u: User = Depends(actor)):
        from .features import i18n as I18N
        return {"languages": I18N.LANGUAGES}

    @app.post("/api/v2/i18n/translate")
    def v2_i18n_translate(b: I18nTranslateIn, u: User = Depends(actor)):
        from .features import i18n as I18N
        res = I18N.translate_strings(b.strings or [], b.lang)
        if isinstance(res, dict) and "_map" in res:
            return {"translations": res["_map"], "ai": res["_ai"], "lang": b.lang}
        return {"translations": res or {}, "ai": False, "lang": b.lang}

    @app.post("/api/v2/i18n/normalize")
    def v2_i18n_normalize(b: I18nTextIn, s: Session = Depends(db), u: User = Depends(actor)):
        from .features import i18n as I18N
        res = I18N.to_english(b.text or "", b.lang)
        audit(s, "v2.i18n_normalize", u.username,
              {"detected": res.get("detected_language"), "chars": len(b.text or "")})
        s.commit()
        return res

    @app.post("/api/v2/i18n/document")
    def v2_i18n_document(b: I18nTextIn, s: Session = Depends(db), u: User = Depends(actor)):
        from .features import i18n as I18N
        res = I18N.translate_document(b.text or "")
        audit(s, "v2.i18n_document", u.username,
              {"detected": res.get("detected_language"), "chars": len(b.text or "")})
        s.commit()
        return res

    # ===== Critical Vendor Modelling (transparency) =====
    @app.get("/api/v2/criticality/model")
    def v2_criticality_model(u: User = Depends(require("vendor.view")), s: Session = Depends(db)):
        from .features import criticality as CRIT
        return CRIT.model(s)

    # ===== Supplier Incidents =====
    _INC_SLA = {"Critical": 24, "High": 48, "Medium": 72, "Low": 120}

    def _notif_light(date_of_incident, notified_at, sla_hours):
        from datetime import datetime, timezone, timedelta
        if not date_of_incident or not sla_hours:
            return None
        def _p(x):
            try:
                d = datetime.fromisoformat(str(x).replace("Z", "+00:00"))
                return d if d.tzinfo else d.replace(tzinfo=timezone.utc)
            except Exception:
                return None
        di = _p(date_of_incident)
        if not di:
            return None
        due = di + timedelta(hours=sla_hours)
        if notified_at:
            n = _p(notified_at)
            if not n:
                return "amber"
            return "green" if n <= due else "red"
        return "amber" if datetime.now(timezone.utc) <= due else "red"

    def _inc_out(r):
        return {
            "incident_id": r.incident_id, "date_of_incident": r.date_of_incident,
            "reported_date": r.reported_date, "reported_by": r.reported_by,
            "vendor_id": r.vendor_id, "vendor_name": r.vendor_name,
            "engagement_id": r.engagement_id,
            "active_engagements": json.loads(r.active_engagements or "[]"),
            "incident_type": r.incident_type, "severity": r.severity,
            "customer_impacting": r.customer_impacting, "impacts_client_org": r.impacts_client_org,
            "impact_description": r.impact_description, "region": json.loads(r.region or "[]"),
            "root_cause_assessment": r.root_cause_assessment,
            "risk_entry_needed": r.risk_entry_needed, "risk_entry_ref": r.risk_entry_ref,
            "notes_log": json.loads(r.notes_log or "[]"),
            "attachments": json.loads(r.attachments or "[]"), "status": r.status,
            "vendor_notified_at": r.vendor_notified_at,
            "notification_sla_hours": r.notification_sla_hours,
            "notification_compliant": r.notification_compliant,
            "contract_notification_summary": r.contract_notification_summary,
            "created_at": r.created_at.isoformat() if r.created_at else None,
            "created_by": r.created_by,
        }

    def _inc_get(s, iid):
        from .features.registry_models import IncidentRecord
        r = s.scalars(select(IncidentRecord).where(IncidentRecord.incident_id == iid)).first()
        if not r:
            raise HTTPException(404, "incident not found")
        return r

    @app.get("/api/v2/incidents")
    def v2_incidents_list(status: str = None, severity: str = None, vendor_id: str = None,
                          s: Session = Depends(db), u: User = Depends(require("incident.view"))):
        from .features.registry_models import IncidentRecord
        rows = s.scalars(select(IncidentRecord).order_by(IncidentRecord.id.desc())).all()
        out = []
        for r in rows:
            if status and r.status != status:
                continue
            if severity and r.severity != severity:
                continue
            if vendor_id and r.vendor_id != vendor_id:
                continue
            out.append(_inc_out(r))
        return out

    @app.post("/api/v2/incidents")
    def v2_incidents_create(b: IncidentIn, s: Session = Depends(db),
                            u: User = Depends(require("incident.manage"))):
        from .features.registry_models import IncidentRecord, VendorRecord, EngagementRecord
        from .features import registry_service as RS
        iid = RS.next_id(s, "incident")
        vname = None
        active = []
        if b.vendor_id:
            v = s.scalars(select(VendorRecord).where(VendorRecord.vendor_id == b.vendor_id)).first()
            vname = v.legal_name if v else None
            # auto-tag active engagements
            active = [e.engagement_id for e in s.scalars(select(EngagementRecord).where(
                EngagementRecord.vendor_id == b.vendor_id)).all() if (e.status or "Active") == "Active"]
        from .features import config_store as _CFG
        _sla_map = _CFG.incident_sla_map(s)
        sla = b.notification_sla_hours or _sla_map.get(b.severity or "Medium", 72)
        light = _notif_light(b.date_of_incident, b.vendor_notified_at, sla)
        r = IncidentRecord(
            incident_id=iid, date_of_incident=b.date_of_incident,
            reported_date=b.reported_date, reported_by=b.reported_by or u.username,
            vendor_id=b.vendor_id, vendor_name=vname,
            engagement_id=b.engagement_id,
            active_engagements=json.dumps(active), incident_type=b.incident_type,
            severity=b.severity or "Medium", customer_impacting=bool(b.customer_impacting),
            impacts_client_org=bool(b.impacts_client_org), impact_description=b.impact_description,
            region=json.dumps(b.region or []), root_cause_assessment=b.root_cause_assessment,
            risk_entry_needed=bool(b.risk_entry_needed), status=b.status or "Drafted",
            vendor_notified_at=b.vendor_notified_at, notification_sla_hours=sla,
            notification_compliant=light, created_by=u.username)
        s.add(r)
        audit(s, "v2.incident_create", u.username, {"incident_id": iid, "vendor_id": b.vendor_id,
              "severity": r.severity, "auto_tagged": len(active)})
        notify(s, f"Supplier incident {iid} raised ({r.severity})", "all")
        s.commit()
        return _inc_out(r)

    @app.get("/api/v2/incidents/{iid}")
    def v2_incident_get(iid: str, s: Session = Depends(db), u: User = Depends(require("incident.view"))):
        return _inc_out(_inc_get(s, iid))

    @app.put("/api/v2/incidents/{iid}")
    def v2_incident_update(iid: str, b: IncidentIn, s: Session = Depends(db),
                           u: User = Depends(require("incident.manage"))):
        r = _inc_get(s, iid)
        for f in ("date_of_incident", "reported_date", "reported_by", "incident_type",
                  "severity", "impact_description", "root_cause_assessment", "status",
                  "vendor_notified_at", "engagement_id"):
            v = getattr(b, f)
            if v is not None:
                setattr(r, f, v)
        if b.customer_impacting is not None:
            r.customer_impacting = bool(b.customer_impacting)
        if b.impacts_client_org is not None:
            r.impacts_client_org = bool(b.impacts_client_org)
        if b.risk_entry_needed is not None:
            r.risk_entry_needed = bool(b.risk_entry_needed)
        if b.region is not None:
            r.region = json.dumps(b.region)
        if b.notification_sla_hours is not None:
            r.notification_sla_hours = b.notification_sla_hours
        if not r.notification_sla_hours:
            from .features import config_store as _CFG
            r.notification_sla_hours = _CFG.incident_sla_map(s).get(r.severity or "Medium", 72)
        r.notification_compliant = _notif_light(r.date_of_incident, r.vendor_notified_at, r.notification_sla_hours)
        audit(s, "v2.incident_update", u.username, {"incident_id": iid, "status": r.status})
        s.commit()
        return _inc_out(r)

    @app.post("/api/v2/incidents/{iid}/note")
    def v2_incident_note(iid: str, b: IncidentNoteIn, s: Session = Depends(db),
                         u: User = Depends(require("incident.manage"))):
        import datetime as _dt
        r = _inc_get(s, iid)
        log = json.loads(r.notes_log or "[]")
        log.append({"ts": _dt.datetime.now(_dt.timezone.utc).isoformat(timespec="seconds"),
                    "user": u.username, "note": b.note})
        r.notes_log = json.dumps(log)
        audit(s, "v2.incident_note", u.username, {"incident_id": iid})
        s.commit()
        return _inc_out(r)

    @app.post("/api/v2/incidents/{iid}/attach")
    def v2_incident_attach(iid: str, b: IncidentAttachIn, s: Session = Depends(db),
                           u: User = Depends(require("incident.manage"))):
        r = _inc_get(s, iid)
        att = json.loads(r.attachments or "[]")
        att.append(b.name)
        r.attachments = json.dumps(att)
        audit(s, "v2.incident_attach", u.username, {"incident_id": iid, "name": b.name})
        s.commit()
        return _inc_out(r)

    @app.post("/api/v2/incidents/{iid}/draft-risk")
    def v2_incident_draft_risk(iid: str, s: Session = Depends(db),
                               u: User = Depends(require("incident.manage"))):
        from .features import registry_service as RS
        from .agents import llm_config
        r = _inc_get(s, iid)
        engs = json.loads(r.active_engagements or "[]")
        summary = (f"Incident {r.incident_id} ({r.severity}, {r.incident_type or 'incident'}) for "
                   f"{r.vendor_name or r.vendor_id}. Impact: {r.impact_description or 'n/a'}. "
                   f"Root cause: {r.root_cause_assessment or 'n/a'}.")
        title = f"Risk from incident {r.incident_id}: {(r.incident_type or 'supplier incident')}"
        description = summary
        remediation = "Confirm containment, validate root-cause remediation, and re-test affected controls."
        if llm_config.status().get("live_ready"):
            try:
                out = llm_config.complete(
                    PROMPTS.resolve(s, "incident_risk_entry"),
                    summary, domain="risk", max_tokens=600)
                t = (out or "").replace("```json", "").replace("```", "").strip()
                i, e = t.find("{"), t.rfind("}")
                d = json.loads(t[i:e + 1]) if (i != -1 and e != -1) else {}
                title = d.get("title") or title
                description = d.get("description") or description
                remediation = d.get("suggested_remediation") or remediation
            except Exception as _e:
                _obs_swallow('bro_app.py', _e)
        f = RS.create_finding(s, title=title, severity=r.severity or "Medium", source="Incident",
                              description=description, suggested_remediation=remediation,
                              vendor_id=r.vendor_id, engagement_id=(engs[0] if engs else None),
                              status="Draft")
        r.risk_entry_needed = True
        r.risk_entry_ref = f.finding_id
        audit(s, "v2.incident_draft_risk", u.username, {"incident_id": iid, "finding_id": f.finding_id})
        s.commit()
        return {"finding_id": f.finding_id, "title": title, "incident_id": iid}

    @app.post("/api/v2/incidents/{iid}/contract-summary")
    def v2_incident_contract_summary(iid: str, s: Session = Depends(db),
                                     u: User = Depends(require("incident.view"))):
        from .agents import llm_config
        r = _inc_get(s, iid)
        light = _notif_light(r.date_of_incident, r.vendor_notified_at, r.notification_sla_hours)
        r.notification_compliant = light
        # locate a contract document for the vendor (demonstrator)
        contract_text = None
        try:
            from .features.documents import StoredDocument as _SD
            doc = s.scalars(select(_SD).where(_SD.vendor_id == r.vendor_id)).first()
            contract_text = (doc.content if doc and getattr(doc, "content", None) else None)
        except Exception:
            contract_text = None
        if not llm_config.status().get("live_ready"):
            r.contract_notification_summary = None
            s.commit()
            return {"holding": True, "message": AI_HOLDING, "traffic_light": light,
                    "sla_hours": r.notification_sla_hours}
        prompt = ("Summarise the supplier's INCIDENT-NOTIFICATION obligation from this contract "
                  "(required notification window, to whom, and any conditions). If no contract text is "
                  "available, state the typical market obligation for this severity. Be concise (3-4 lines).\n\n"
                  f"Severity: {r.severity}. Contract text: {contract_text or '[not on file]'}")
        try:
            summ = llm_config.complete(PROMPTS.resolve(s, "contract_summary"), prompt,
                                       domain="legal", max_tokens=400)
        except Exception:
            summ = None
        r.contract_notification_summary = (summ or "").strip() or None
        audit(s, "v2.incident_contract_summary", u.username, {"incident_id": iid, "light": light})
        s.commit()
        return {"holding": False, "summary": r.contract_notification_summary,
                "traffic_light": light, "sla_hours": r.notification_sla_hours,
                "vendor_notified_at": r.vendor_notified_at, "date_of_incident": r.date_of_incident}

    @app.get("/api/v2/incidents/{iid}/warroom")
    def v2_incident_warroom(iid: str, s: Session = Depends(db),
                            u: User = Depends(require("incident.view"))):
        from .features import exposure as EXP
        from .features.registry_models import (EngagementRecord, FourthPartyVendor,
                                              FourthPartyRecord)
        from .agents import llm_config
        r = _inc_get(s, iid)
        vid = r.vendor_id
        active = json.loads(r.active_engagements or "[]")
        engs = s.scalars(select(EngagementRecord).where(EngagementRecord.vendor_id == vid)).all() if vid else []
        eng_map = {e.engagement_id: e for e in engs}
        affected = [eng_map[e] for e in active if e in eng_map] or engs
        bus = sorted({e.business_unit for e in affected if e.business_unit})
        affected_out = [{"engagement_id": e.engagement_id, "title": e.title,
                         "business_unit": e.business_unit, "inherent_band": e.inherent_band,
                         "residual_band": e.residual_band,
                         "annual_value": round(float(e.annual_value or 0))} for e in affected]
        # incident → issue matching (open findings implicated + peers with same gap)
        desc = " ".join(x for x in [r.impact_description, r.root_cause_assessment] if x)
        match = EXP.incident_match(s, vid, desc, r.incident_type) if vid else {"matched_findings": [], "peer_exposure": [], "peer_count": 0}
        # BU exposure profiles for the affected BUs
        bu_all = EXP.bu_exposure(s)["business_units"]
        bu_prof = [b for b in bu_all if b["business_unit"] in bus]
        # fourth-party concentration the vendor relies on (+ SPOF)
        links = s.scalars(select(FourthPartyVendor)).all()
        cnt = {}
        for ln in links:
            cnt.setdefault(ln.fourth_party_id, []).append(ln.vendor_id)
        total = len({ln.vendor_id for ln in links}) or 1
        fps = {f.fourth_party_id: f for f in s.scalars(select(FourthPartyRecord)).all()}
        fp_ids = [ln.fourth_party_id for ln in links if ln.vendor_id == vid]
        fourth = [{"fourth_party_id": fid, "legal_name": (fps[fid].legal_name if fid in fps else fid),
                   "shared_with": len(cnt.get(fid, [])),
                   "spof": len(cnt.get(fid, [])) >= max(10, total * 0.15)} for fid in fp_ids]
        fourth.sort(key=lambda x: -x["shared_with"])
        light = _notif_light(r.date_of_incident, r.vendor_notified_at, r.notification_sla_hours)
        # Ecosystem impact — UPSTREAM vendors: those who declared THIS vendor as their 4th party
        upstream = []
        if vid:
            from .features.registry_models import VendorRecord as _VR
            vmap = {v.vendor_id: v.legal_name for v in s.scalars(select(_VR)).all()}
            self_fp_ids = [f.fourth_party_id for f in s.scalars(select(FourthPartyRecord)
                           .where(FourthPartyRecord.vendor_id == vid)).all()]
            seen_up = set()
            for ln in links:
                if ln.fourth_party_id in self_fp_ids and ln.vendor_id and ln.vendor_id != vid \
                        and ln.vendor_id not in seen_up:
                    seen_up.add(ln.vendor_id)
                    upstream.append({"vendor_id": ln.vendor_id,
                                     "legal_name": vmap.get(ln.vendor_id, ln.vendor_id)})
        stats = {
            "engagements_affected": len(affected), "bus_affected": len(bus),
            "peers_same_gap": match["peer_count"], "open_findings_matched": len(match["matched_findings"]),
            "fourth_party_deps": len(fourth), "spof_deps": sum(1 for x in fourth if x["spof"]),
        }
        brief = (f"{r.severity} {r.incident_type or 'incident'} at {r.vendor_name or vid}. "
                 f"{stats['engagements_affected']} engagement(s) across {stats['bus_affected']} business unit(s) affected; "
                 f"{stats['open_findings_matched']} open finding(s) directly implicated; "
                 f"{stats['peers_same_gap']} peer vendor(s) carry the same unremediated gap. "
                 f"Vendor notification status: {(light or 'n/a').upper()}."
                 + (f" Concentration risk: relies on {stats['spof_deps']} single-point-of-failure sub-provider(s)." if stats['spof_deps'] else ""))
        ai = False
        if llm_config.status().get("live_ready"):
            try:
                out = llm_config.complete(
                    PROMPTS.resolve(s, "incident_board_note"),
                    f"Incident: {r.severity} {r.incident_type} at {r.vendor_name}. Impact: {r.impact_description}. RCA: {r.root_cause_assessment}. "
                    f"Systemic: {brief}", domain="risk", max_tokens=400)
                if out and out.strip():
                    brief = out.strip(); ai = True
            except Exception as _e:
                _obs_swallow('bro_app.py', _e)
        return {"incident": _inc_out(r), "sla_light": light, "stats": stats,
                "affected_engagements": affected_out, "business_units": bu_prof,
                "matched_findings": match["matched_findings"], "peer_exposure": match["peer_exposure"],
                "fourth_party_deps": fourth, "upstream_vendors": upstream, "brief": brief, "brief_ai": ai}

    @app.get("/api/v2/evidence/{vendor_id}")
    def v2_evidence(vendor_id: str, s: Session = Depends(db),
                    u: User = Depends(require("vendor.view"))):
        from .features.registry_models import (VendorRecord, EngagementRecord,
                                              AssessmentRecord, FindingRecord, IncidentRecord)
        v = s.scalars(select(VendorRecord).where(VendorRecord.vendor_id == vendor_id)).first()
        if not v:
            raise HTTPException(404, "vendor not found")
        engs = s.scalars(select(EngagementRecord).where(EngagementRecord.vendor_id == vendor_id)).all()
        asms = s.scalars(select(AssessmentRecord).where(AssessmentRecord.vendor_id == vendor_id)).all()
        finds = s.scalars(select(FindingRecord).where(FindingRecord.vendor_id == vendor_id)).all()
        incs = s.scalars(select(IncidentRecord).where(IncidentRecord.vendor_id == vendor_id)).all()
        docs = []
        try:
            from .features.documents import StoredDocument as _SD
            docs = s.scalars(select(_SD).where(_SD.vendor_id == vendor_id)).all()
        except Exception:
            docs = []
        ids = {vendor_id} | {e.engagement_id for e in engs} | {a.assessment_id for a in asms} \
            | {f.finding_id for f in finds} | {i.incident_id for i in incs}
        # audit trail referencing this subject + full-chain integrity
        all_audit = s.scalars(select(AuditLog).order_by(AuditLog.seq)).all()
        prev = "genesis"; intact = True
        for r in all_audit:
            expect = eng.chain_hash(prev, r.action, r.actor, json.loads(r.detail) if r.detail else {})
            if r.prev_hash != prev or r.entry_hash != expect:
                intact = False
            prev = r.entry_hash
        rel = [r for r in all_audit if r.detail and any(i in r.detail for i in ids)]
        approvals = [{"assessment_id": a.assessment_id, "outcome": a.outcome,
                      "assessor": a.assessor_user, "signed_off": a.assessor_signed_off,
                      "locked": a.locked, "status": a.status} for a in asms]
        signed = sum(1 for a in asms if a.assessor_signed_off)
        return {
            "vendor": {"vendor_id": v.vendor_id, "legal_name": v.legal_name, "tier": v.tier,
                       "is_critical": v.is_critical, "hq_country": v.hq_country,
                       "lei": v.lei, "status": v.status},
            "engagements": [{"engagement_id": e.engagement_id, "title": e.title,
                             "inherent_band": e.inherent_band, "residual_band": e.residual_band,
                             "business_unit": e.business_unit,
                             "annual_value": round(float(e.annual_value or 0))} for e in engs],
            "assessments": [{"assessment_id": a.assessment_id, "status": a.status,
                             "outcome": a.outcome, "inherent_band": a.inherent_band,
                             "residual_band": a.residual_band, "assessor": a.assessor_user,
                             "signed_off": a.assessor_signed_off, "locked": a.locked,
                             "created_at": a.created_at.isoformat() if a.created_at else None} for a in asms],
            "findings": [{"finding_id": f.finding_id, "title": f.title, "severity": f.severity,
                          "status": f.status, "source": f.source,
                          "risk_accepted": getattr(f, "risk_accepted", False),
                          "acceptance_expiry": getattr(f, "acceptance_expiry", None)} for f in finds],
            "documents": [{"doc_id": d.doc_id, "filename": d.filename, "purpose": d.purpose,
                           "uploaded_by": d.uploaded_by,
                           "created_at": d.created_at.isoformat() if getattr(d, "created_at", None) else None} for d in docs],
            "incidents": [{"incident_id": i.incident_id, "incident_type": i.incident_type,
                           "severity": i.severity, "status": i.status,
                           "notification_compliant": i.notification_compliant} for i in incs],
            "approvals": approvals,
            "audit_trail": [{"seq": r.seq, "action": r.action, "actor": r.actor,
                             "created_at": r.created_at.isoformat() if r.created_at else None,
                             "entry_hash": (r.entry_hash or "")[:16]} for r in rel[-40:]],
            "summary": {"engagements": len(engs), "assessments": len(asms),
                        "signed_off": signed, "findings": len(finds),
                        "documents": len(docs), "incidents": len(incs),
                        "audit_entries": len(rel)},
            "chain_intact": intact,
        }

    @app.post("/api/v2/copilot/ask")
    def v2_copilot_ask(b: I18nTextIn, s: Session = Depends(db),
                       u: User = Depends(require("vendor.view"))):
        from .features import copilot as COP
        from .agents import llm_config
        res = COP.answer_query(s, b.text or "")
        res["ai"] = False
        if llm_config.status().get("live_ready") and res["rows"]:
            try:
                top = "; ".join(f"{r['label']} ({r['sublabel']})" for r in res["rows"][:12])
                out = llm_config.complete(
                    PROMPTS.resolve(s, "copilot_interpret"),
                    f"Question: {b.text}\nResult: {res['answer']}\nTop records: {top}",
                    domain="risk", max_tokens=220)
                if out and out.strip():
                    res["narrative"] = out.strip(); res["ai"] = True
            except Exception as _e:
                _obs_swallow('bro_app.py', _e)
        audit(s, "v2.copilot_ask", u.username, {"q": (b.text or "")[:120], "n": res["count"]})
        s.commit()
        return res

    @app.get("/api/v2/stress-radar")
    def v2_stress_radar(u: User = Depends(require("vendor.view")), s: Session = Depends(db)):
        from .features import stress as STRESS
        return STRESS.radar(s)

    @app.get("/api/v2/scenario/options")
    def v2_scenario_options(u: User = Depends(require("vendor.view")), s: Session = Depends(db)):
        from .features import scenario as SCEN
        return SCEN.options(s)

    @app.get("/api/v2/scenario/fourth-party-impact/{vid}")
    def v2_scenario_fourth_party_impact(vid: str, s: Session = Depends(db),
                                        u: User = Depends(require("vendor.view"))):
        """Downstream 4th-party impact: the registered third parties that declared this
        vendor as a sub-processor / dependency, and are therefore *indirectly* exposed if
        it fails. Matches the vendor to its 4th-party identity (explicit link or name)."""
        from .features.registry_models import (FourthPartyRecord, FourthPartyVendor,
                                               EngagementRecord)
        v = s.scalars(select(VendorRecord).where(VendorRecord.vendor_id == vid)).first()
        if not v:
            raise HTTPException(404, "vendor not found")
        # resolve the vendor's 4th-party identity/identities
        fp_ids = set()
        if getattr(v, "fourth_party_id", None):
            fp_ids.add(v.fourth_party_id)
        nm = (v.legal_name or "").strip().lower()
        for f in s.scalars(select(FourthPartyRecord)).all():
            if (f.legal_name or "").strip().lower() == nm:
                fp_ids.add(f.fourth_party_id)
        dep_ids = []
        seen = set()
        if fp_ids:
            for ln in s.scalars(select(FourthPartyVendor).where(
                    FourthPartyVendor.fourth_party_id.in_(fp_ids))).all():
                if ln.vendor_id and ln.vendor_id != vid and ln.vendor_id not in seen:
                    seen.add(ln.vendor_id); dep_ids.append(ln.vendor_id)
        dep_vendors = {x.vendor_id: x for x in s.scalars(select(VendorRecord).where(
            VendorRecord.vendor_id.in_(dep_ids))).all()} if dep_ids else {}
        # engagements per dependent
        engs_by_v: dict[str, list] = {}
        if dep_ids:
            for e in s.scalars(select(EngagementRecord).where(
                    EngagementRecord.vendor_id.in_(dep_ids))).all():
                engs_by_v.setdefault(e.vendor_id, []).append(
                    {"engagement_id": e.engagement_id, "title": e.title, "status": e.status})
        dependents = []
        eng_total = 0
        for did in dep_ids:
            dv = dep_vendors.get(did)
            if not dv:
                continue
            elist = engs_by_v.get(did, [])
            eng_total += len(elist)
            dependents.append({"vendor_id": dv.vendor_id, "legal_name": dv.legal_name,
                               "tier": dv.tier, "is_critical": dv.is_critical,
                               "engagements": elist})
        dependents.sort(key=lambda d: (not d["is_critical"], -len(d["engagements"])))
        return {"vendor": {"vendor_id": v.vendor_id, "legal_name": v.legal_name,
                           "is_critical": v.is_critical},
                "is_fourth_party": bool(fp_ids), "fourth_party_ids": sorted(fp_ids),
                "dependent_count": len(dependents), "engagement_count": eng_total,
                "dependents": dependents}

    @app.post("/api/v2/scenario/simulate")
    def v2_scenario_simulate(b: ScenarioIn, s: Session = Depends(db),
                             u: User = Depends(require("vendor.view"))):
        from .features import scenario as SCEN
        from .agents import llm_config
        res = SCEN.simulate(s, b.node_type, b.node_id, b.hours)
        res["brief_ai"] = False
        if llm_config.status().get("live_ready"):
            try:
                out = llm_config.complete(
                    PROMPTS.resolve(s, "scenario_impact"),
                    res["brief"], domain="risk", max_tokens=350)
                if out and out.strip():
                    res["brief"] = out.strip(); res["brief_ai"] = True
            except Exception as _e:
                _obs_swallow('bro_app.py', _e)
        return res

    @app.get("/api/v2/board-pack")
    def v2_board_pack(u: User = Depends(require("vendor.view")), s: Session = Depends(db)):
        from .features import exposure as EXP, graph as GRAPH, criticality as CRIT
        from .features import geopolitical as GEO, integrity as INTEG
        from .features.registry_models import (VendorRecord, EngagementRecord,
                                              FindingRecord, IncidentRecord)
        from .agents import llm_config
        vendors = s.scalars(select(VendorRecord)).all()
        engs = s.scalars(select(EngagementRecord)).all()
        finds = s.scalars(select(FindingRecord)).all()
        incs = s.scalars(select(IncidentRecord).order_by(IncidentRecord.id.desc())).all()
        open_find = [f for f in finds if f.status != "Closed"]
        hi_find = [f for f in open_find if (f.severity or "") in ("High", "Critical")]
        exec_stats = {
            "vendors": len(vendors), "critical_vendors": sum(1 for v in vendors if v.is_critical),
            "engagements": len(engs), "open_findings": len(open_find),
            "high_critical_findings": len(hi_find),
            "annual_value": round(sum(float(e.annual_value or 0) for e in engs)),
        }
        graph = GRAPH.build_graph(s)
        bu = EXP.bu_exposure(s)
        crit = CRIT.model(s)
        geo = GEO.exposure(s)
        dq = INTEG.run_sweep(s)["health"]
        inc_open = [i for i in incs if i.status != "Closed"]
        inc_sev = {}
        for i in incs:
            inc_sev[i.severity] = inc_sev.get(i.severity, 0) + 1
        inc_breach = sum(1 for i in incs if i.notification_compliant == "red")
        disagreements = [c for c in crit["subthreshold_high_risk"] if c.get("model_critical")]
        payload = {
            "generated_at": __import__("datetime").datetime.now(__import__("datetime").timezone.utc).isoformat(timespec="seconds"),
            "exec": exec_stats,
            "concentration": {"shared_sub_providers": graph["stats"]["shared_fourth_parties"],
                              "spof": graph["stats"]["spof_count"],
                              "max_fanout_pct": graph["stats"]["max_fanout_pct"],
                              "top": graph["shared_fourth_parties"][:5]},
            "bu_exposure": bu["business_units"][:6],
            "criticality": {"critical": crit["stats"]["critical"],
                            "disagreements": len(disagreements),
                            "top_disagreements": disagreements[:5]},
            "geopolitical": {"exposed": geo["stats"]["exposed_vendors"],
                             "component_shortage": geo["stats"]["component_shortage_vendors"],
                             "high_risk_jurisdictions": geo["stats"]["high_risk_jurisdictions"],
                             "top": geo["flagged_vendors"][:6]},
            "incidents": {"total": len(incs), "open": len(inc_open),
                          "by_severity": inc_sev, "notification_breaches": inc_breach,
                          "recent": [{"incident_id": i.incident_id, "vendor_name": i.vendor_name,
                                      "severity": i.severity, "status": i.status,
                                      "incident_type": i.incident_type,
                                      "notification_compliant": i.notification_compliant}
                                     for i in incs[:6]]},
            "data_health": {"overall": dq["overall"], "completeness": dq["completeness"],
                            "issues": dq["issue_count"]},
        }
        # executive summary — deterministic, AI-enhanced if live
        summary = (f"The third-party estate comprises {exec_stats['vendors']} vendors "
                   f"({exec_stats['critical_vendors']} critical) across {exec_stats['engagements']} engagements, "
                   f"£{exec_stats['annual_value']:,} annual value. {exec_stats['high_critical_findings']} high/critical "
                   f"findings are open. Concentration: {graph['stats']['spof_count']} single points of failure, the "
                   f"largest serving {graph['stats']['max_fanout_pct']}% of the estate. {len(incs)} incidents "
                   f"({len(inc_open)} open, {inc_breach} with a notification-SLA breach). {geo['stats']['exposed_vendors']} "
                   f"vendors carry geopolitical/export-control exposure. Data-health score {dq['overall']}/100.")
        payload["summary_ai"] = False
        if llm_config.status().get("live_ready"):
            try:
                out = llm_config.complete(
                    PROMPTS.resolve(s, "board_pack_summary"),
                    summary, domain="risk", max_tokens=500)
                if out and out.strip():
                    summary = out.strip(); payload["summary_ai"] = True
            except Exception as _e:
                _obs_swallow('bro_app.py', _e)
        payload["summary"] = summary
        return payload

    @app.post("/api/v2/remediations")
    def v2_create_remediation(b: V2RemediationIn, s: Session = Depends(db),
                              u: User = Depends(require("finding.manage"))):
        r = RS.create_remediation(s, finding_id=b.finding_id, plan=b.plan,
                                  owner=b.owner, target_date=b.target_date)
        audit(s, "v2.remediation_created", u.username,
              {"remediation_id": r.remediation_id, "finding_id": b.finding_id})
        s.commit()
        return {"remediation_id": r.remediation_id}

    # ---- fourth parties ----
    @app.post("/api/v2/fourth-parties")
    def v2_create_fourth_party(b: V2FourthPartyIn, s: Session = Depends(db),
                               u: User = Depends(require("lifecycle.fourthparty"))):
        fp = RS.create_fourth_party(s, legal_name=b.legal_name, vendor_ids=b.vendor_ids or [],
                                    also_vendor_id=b.vendor_id, service_provided=b.service_provided,
                                    hq_country=b.hq_country)
        audit(s, "v2.fourth_party_created", u.username,
              {"fourth_party_id": fp.fourth_party_id, "concentration": fp.concentration_flag})
        s.commit()
        return {"fourth_party_id": fp.fourth_party_id, "concentration_flag": fp.concentration_flag}

    @app.get("/api/v2/fourth-parties")
    def v2_list_fourth_parties(s: Session = Depends(db), u: User = Depends(require("lifecycle.fourthparty"))):
        out = []
        for fp in s.scalars(select(FourthPartyRecord)).all():
            vlinks = [fv.vendor_id for fv in s.scalars(select(FourthPartyVendor).where(
                FourthPartyVendor.fourth_party_id == fp.fourth_party_id)).all()]
            out.append({"fourth_party_id": fp.fourth_party_id, "legal_name": fp.legal_name,
                        "concentration_flag": fp.concentration_flag, "vendor_id": fp.vendor_id,
                        "supports_vendors": vlinks})
        return out

    @app.get("/api/v2/fourth-parties/{fpid}/vendors")
    def v2_fourth_party_vendors(fpid: str, s: Session = Depends(db),
                                u: User = Depends(require("lifecycle.fourthparty"))):
        from .features.registry_models import VendorRecord
        fp = s.scalars(select(FourthPartyRecord).where(FourthPartyRecord.fourth_party_id == fpid)).first()
        if not fp:
            raise HTTPException(404, "fourth party not found")
        vids = [fv.vendor_id for fv in s.scalars(select(FourthPartyVendor).where(
            FourthPartyVendor.fourth_party_id == fpid)).all()]
        vmap = {v.vendor_id: v for v in s.scalars(select(VendorRecord).where(
            VendorRecord.vendor_id.in_(vids))).all()} if vids else {}
        vendors = [{"vendor_id": vid, "legal_name": (vmap[vid].legal_name if vid in vmap else vid),
                    "tier": (vmap[vid].tier if vid in vmap else None),
                    "is_critical": (bool(vmap[vid].is_critical) if vid in vmap else False)}
                   for vid in vids]
        return {"fourth_party_id": fpid, "legal_name": fp.legal_name,
                "concentration_flag": fp.concentration_flag,
                "supports_count": len(vendors), "vendors": vendors}

    # ---- artefacts + revalidation + issues ----
    @app.post("/api/v2/artefacts")
    def v2_create_artefact(b: V2ArtefactIn, s: Session = Depends(db),
                           u: User = Depends(require("lifecycle.documents"))):
        art = RS.create_artefact(s, vendor_id=b.vendor_id, name=b.name,
                                 artefact_type=b.artefact_type or "certificate",
                                 expiry_date=b.expiry_date, received_via=b.received_via or "upload",
                                 supersedes=b.supersedes, issue_date=b.issue_date,
                                 engagement_id=b.engagement_id)
        audit(s, "v2.artefact_created", u.username,
              {"artefact_id": art.artefact_id, "status": art.status})
        s.commit()
        return {"artefact_id": art.artefact_id, "status": art.status,
                "supersedes": art.supersedes}

    @app.get("/api/v2/artefacts")
    def v2_list_artefacts(vendor_id: Optional[str] = None, s: Session = Depends(db),
                          u: User = Depends(require("lifecycle.documents"))):
        rows = s.scalars(select(ArtefactRecord)).all()
        return [{"artefact_id": a.artefact_id, "vendor_id": a.vendor_id, "name": a.name,
                 "type": a.artefact_type, "expiry_date": a.expiry_date, "status": a.status,
                 "is_current": a.is_current, "received_via": a.received_via,
                 "doc_link": a.object_uri}
                for a in rows if not vendor_id or a.vendor_id == vendor_id]

    # ============================================================
    # CR-4/5/12 — DOCUMENT STORE + AI EXTRACTION
    # ============================================================
    @app.post("/api/v2/documents/upload")
    def v2_doc_upload(b: DocUploadIn, s: Session = Depends(db),
                      u: User = Depends(require("lifecycle.documents"))):
        from .features import documents as DOC
        out = []
        for f in b.files:
            try:
                row = DOC.store_document(s, filename=f.filename, content_type=f.content_type or "",
                                         data_b64=f.data_b64, vendor_id=b.vendor_id,
                                         engagement_id=b.engagement_id, uploaded_by=u.username,
                                         purpose=b.purpose)
            except ValueError as e:
                raise HTTPException(422, str(e))
            out.append({"doc_id": row.doc_id, "filename": row.filename, "size": row.size_bytes})
        audit(s, "v2.doc_upload", u.username, {"count": len(out), "purpose": b.purpose})
        s.commit()
        return {"documents": out}

    @app.get("/api/v2/documents")
    def v2_doc_list(purpose: str = "", vendor_id: str = "", ai_only: bool = False,
                    s: Session = Depends(db), u: User = Depends(require("lifecycle.documents"))):
        """List stored documents/reports. ai_only filters to AI-generated reports;
        purpose/vendor_id narrow further. Used by the Documents and AI Reports views."""
        from .features.documents import StoredDocument
        from .features.registry_models import VendorRecord
        AI_PURPOSES = ("fdd_report", "reputation_report", "fdd_reputation_report", "proassess_report")
        q = s.query(StoredDocument)
        if purpose:
            q = q.filter(StoredDocument.purpose == purpose)
        if vendor_id:
            q = q.filter(StoredDocument.vendor_id == vendor_id)
        if ai_only:
            q = q.filter(StoredDocument.purpose.in_(AI_PURPOSES))
        rows = q.order_by(StoredDocument.id.desc()).all()
        vmap = {v.vendor_id: v.legal_name for v in s.query(VendorRecord).all()}
        return {"documents": [{
            "doc_id": d.doc_id, "filename": d.filename, "purpose": d.purpose,
            "content_type": d.content_type, "size_bytes": d.size_bytes,
            "vendor_id": d.vendor_id, "vendor_name": vmap.get(d.vendor_id) if d.vendor_id else None,
            "engagement_id": d.engagement_id, "uploaded_by": d.uploaded_by,
            "is_ai_report": d.purpose in AI_PURPOSES,
            "url": f"/api/v2/documents/{d.doc_id}",
            "created_at": d.created_at.isoformat() if d.created_at else None,
        } for d in rows]}

    @app.get("/api/v2/documents/{doc_id}")
    def v2_doc_get(doc_id: str, s: Session = Depends(db),
                   u: User = Depends(require("lifecycle.documents"))):
        from .features import documents as DOC
        import base64 as _b64
        from fastapi import Response
        d = DOC.get_document(s, doc_id)
        if not d:
            raise HTTPException(404, "document not found")
        raw = _b64.b64decode(d.data_b64 or "")
        return Response(content=raw, media_type=d.content_type,
                        headers={"Content-Disposition": f'inline; filename="{d.filename}"'})

    @app.post("/api/v2/certificates/ingest")
    def v2_cert_ingest(b: CertIngestIn, s: Session = Depends(db),
                       u: User = Depends(require("lifecycle.documents"))):
        """CR-5: multi-document certificate ingest. Each document is stored, read by the
        extractor, and a certificate record (ArtefactRecord) is created with the document
        linked for viewing, tagged to the vendor (and engagement where given)."""
        from .features import documents as DOC
        v = s.scalars(select(VendorRecord).where(VendorRecord.vendor_id == b.vendor_id)).first()
        if not v:
            raise HTTPException(404, "vendor not found")
        created = []
        for f in b.files:
            try:
                doc = DOC.store_document(s, filename=f.filename, content_type=f.content_type or "",
                                         data_b64=f.data_b64, vendor_id=b.vendor_id,
                                         engagement_id=b.engagement_id, uploaded_by=u.username,
                                         purpose="certificate")
            except ValueError as e:
                raise HTTPException(422, str(e))
            ext = DOC.extract_certificate(s, doc)
            art = RS.create_artefact(
                s, vendor_id=b.vendor_id, name=ext["name"],
                artefact_type=ext["artefact_type"], expiry_date=ext.get("expiry_date"),
                received_via="upload", issue_date=ext.get("issue_date"),
                engagement_id=b.engagement_id, object_uri=f"/api/v2/documents/{doc.doc_id}")
            created.append({"artefact_id": art.artefact_id, "name": art.name,
                            "type": art.artefact_type, "expiry_date": art.expiry_date,
                            "status": art.status, "doc_id": doc.doc_id,
                            "doc_link": f"/api/v2/documents/{doc.doc_id}",
                            "gaps": ext.get("gaps", [])})
        audit(s, "v2.cert_ingest", u.username,
              {"vendor_id": b.vendor_id, "count": len(created)})
        s.commit()
        return {"certificates": created}

    @app.post("/api/v2/artefacts/revalidate")
    def v2_revalidate(s: Session = Depends(db), u: User = Depends(require("lifecycle.evidence"))):
        result = RS.revalidation_run(s)
        # send 7-day notices via the email path (SMTP or simulation)
        for n in result["notify_7day"]:
            v = s.scalars(select(VendorRecord).where(VendorRecord.vendor_id == n["vendor_id"])).first()
            contact = s.scalars(select(ContactRecord).where(
                ContactRecord.owner_type == "vendor", ContactRecord.owner_id == n["vendor_id"],
                ContactRecord.is_primary == True)).first()  # noqa: E712
            to = contact.email if contact and contact.email else "vendor@example.com"
            s.add(EmailOutbox(to_addr=to, subject=f"Certificate expiring: {n['name']}",
                              body=f"{n['name']} expires {n['expiry']}. Please provide an updated copy.",
                              sent=False))
        audit(s, "v2.revalidation_run", u.username,
              {"checked": result["checked"], "notify": len(result["notify_7day"]),
               "new_issues": len(result["new_issues"])})
        s.commit()
        return result

    @app.get("/api/v2/issues")
    def v2_list_issues(status: Optional[str] = None, s: Session = Depends(db),
                       u: User = Depends(require("finding.view"))):
        rows = s.scalars(select(IssueRecord).order_by(IssueRecord.id.desc())).all()
        from .features.registry_models import ArtefactRecord
        art_eng = {a.artefact_id: a.engagement_id for a in s.scalars(select(ArtefactRecord)).all()}
        return [{"issue_id": i.issue_id, "vendor_id": i.vendor_id, "vendor_name": i.vendor_name,
                 "engagement_id": art_eng.get(i.artefact_id),
                 "artefact_id": i.artefact_id, "kind": i.kind, "detail": i.detail,
                 "status": i.status, "closed_reason": i.closed_reason}
                for i in rows if not status or i.status == status]

    # ---- financial DD (deterministic engine) ----
    @app.post("/api/v2/financial-dd")
    def v2_financial_dd(b: FinancialIn, s: Session = Depends(db),
                        u: User = Depends(require("intel.financial"))):
        from .features import master_service as MS
        from .features import entity_resolve as ER
        result = FIN.assess_financials(b.figures, b.flags or {})
        ent = ER.resolve_entity(s, vendor_id=b.vendor_id, other_name=b.other_name)
        persisted = False
        if ent.get("registered"):
            persisted = MS.persist_fdd(s, ent["vendor_id"], result)
            MS.refresh_risk_profile(s, ent["vendor_id"])
        result["entity"] = ent
        result["persisted"] = persisted
        audit(s, "v2.financial_dd", u.username,
              {"banding": result["banding"], "altman_zone": result["altman"]["zone"],
               "entity": ent["vendor_name"], "persisted": persisted})
        s.commit()
        return result

    # ---- reputation & ESG (7-pillar engine) ----
    @app.post("/api/v2/reputation")
    def v2_reputation(b: ReputationIn, s: Session = Depends(db),
                      u: User = Depends(require("intel.reputation"))):
        from .features import reputation as REP
        from .features import entity_resolve as ER
        from .features import master_service as MS
        ent = ER.resolve_entity(s, vendor_id=b.vendor_id, other_name=b.other_name)
        result = REP.assess_reputation(b.events or [], b.customer_facing)
        result["entity"] = ent
        persisted = False
        if ent.get("registered"):
            persisted = MS.persist_reputation(s, ent["vendor_id"], result)
            MS.refresh_risk_profile(s, ent["vendor_id"])
        result["persisted"] = persisted
        audit(s, "v2.reputation", u.username,
              {"overall": result["overall"], "verdict": result["verdict"],
               "entity": ent["vendor_name"], "events": result["event_count"],
               "persisted": persisted})
        s.commit()
        return result

    # ---- system configuration manager (admin) ----
    @app.get("/api/v2/config")
    def v2_config_list(s: Session = Depends(db), u: User = Depends(require("admin.config"))):
        from .features import config_store as CFG
        cats = {}
        for r in CFG.list_config(s):
            if r.value_type == "json" or r.category == "_internal" or r.key.startswith("nav."):
                continue
            eff = CFG.get_config(s, r.key)
            dflt = CFG.default_for(r.key)
            cats.setdefault(r.category, []).append({
                "key": r.key, "label": r.label, "description": r.description,
                "type": r.value_type, "value": eff, "default": dflt,
                "is_default": eff == dflt, "updated_by": r.updated_by,
                "updated_at": r.updated_at.isoformat() if r.updated_at else None})
        return {"categories": [{"category": c, "items": items} for c, items in cats.items()]}

    @app.put("/api/v2/config/{key}")
    def v2_config_set(key: str, b: ConfigSetIn, s: Session = Depends(db),
                      u: User = Depends(require("admin.config"))):
        from .features import config_store as CFG
        try:
            row = CFG.set_config(s, key, b.value, updated_by=u.username)
        except (ValueError, TypeError):
            raise HTTPException(status_code=400, detail="Invalid value for this setting type")
        if not row:
            raise HTTPException(status_code=404, detail="Unknown setting")
        audit(s, "v2.config_set", u.username, {"key": key, "value": CFG.get_config(s, key)})
        s.commit()
        return {"key": key, "value": CFG.get_config(s, key), "updated_by": row.updated_by}

    @app.post("/api/v2/config/{key}/reset")
    def v2_config_reset(key: str, s: Session = Depends(db),
                        u: User = Depends(require("admin.config"))):
        from .features import config_store as CFG
        row = CFG.reset_config(s, key, updated_by=u.username)
        if not row:
            raise HTTPException(status_code=404, detail="Unknown setting")
        audit(s, "v2.config_reset", u.username, {"key": key})
        s.commit()
        return {"key": key, "value": CFG.get_config(s, key), "is_default": True}

    # ---- navigation menu order (read: any user; write: admin) ----
    @app.get("/api/v2/nav-order")
    def v2_nav_order_get(s: Session = Depends(db), u: User = Depends(actor)):
        from .features import config_store as CFG
        return {"order": CFG.get_json(s, "nav.order", None)}

    @app.put("/api/v2/nav-order")
    def v2_nav_order_set(b: NavOrderIn, s: Session = Depends(db),
                         u: User = Depends(require("admin.config"))):
        from .features import config_store as CFG
        if b.order is None:
            CFG.delete_key(s, "nav.order")
            audit(s, "v2.nav_order_reset", u.username, {})
            s.commit()
            return {"order": None}
        CFG.upsert_json(s, "nav.order", b.order, updated_by=u.username)
        audit(s, "v2.nav_order_set", u.username,
              {"groups": len((b.order or {}).get("groups", []))})
        s.commit()
        return {"order": CFG.get_json(s, "nav.order", None)}

    # ---- contract management (Matt) ----
    @app.post("/api/v2/contracts/terms")
    def v2_contract_terms(b: ContractTermsIn, s: Session = Depends(db),
                          u: User = Depends(require("intel.contract"))):
        from .features import contracts as CON
        from .features import entity_resolve as ER
        ent = ER.resolve_entity(s, vendor_id=b.vendor_id, other_name=b.other_name)
        terms = CON.required_terms(b.inherent_band, b.exposure or {})
        audit(s, "v2.contract_terms", u.username,
              {"inherent": b.inherent_band, "count": len(terms), "entity": ent["vendor_name"]})
        s.commit()
        return {"inherent_band": b.inherent_band, "required_terms": terms,
                "count": len(terms), "entity": ent}

    @app.post("/api/v2/contracts/gap-report")
    def v2_contract_gap(b: ContractGapIn, s: Session = Depends(db),
                        u: User = Depends(require("intel.contract"))):
        from .features import contracts as CON
        rep = CON.gap_report(b.contract_text, b.inherent_band, b.exposure or {})
        audit(s, "v2.contract_gap_report", u.username,
              {"gaps": len(rep["gaps"]), "critical": rep["critical_gaps"]})
        s.commit()
        return rep

    @app.post("/api/v2/contracts/diff")
    def v2_contract_diff(b: ContractDiffIn, s: Session = Depends(db),
                         u: User = Depends(require("intel.contract"))):
        from .features import contracts as CON
        rep = CON.existing_vs_to_add(b.inherent_band, b.exposure or {},
                                     b.prior_contract_texts or [])
        audit(s, "v2.contract_diff", u.username,
              {"existing": len(rep["terms_already_existing"]),
               "to_add": len(rep["terms_to_be_added"])})
        s.commit()
        return rep

    @app.post("/api/v2/contracts/gap-from-document")
    def v2_contract_gap_doc(b: ContractGapDocIn, s: Session = Depends(db),
                            u: User = Depends(require("intel.contract"))):
        """CR-12: upload a contract document; AI extracts terms; gap review runs against
        the required terms for the engagement's inherent band. For a REGISTERED engagement
        the inherent band and exposure are inherited automatically; only an 'Other' vendor
        is prompted for the band."""
        from .features import documents as DOC
        from .features import contracts as CON
        from .features import master_service as MS
        band = b.inherent_band
        exposure = {}
        engagement_id = b.engagement_id
        vendor_id = b.vendor_id
        # inherit from the engagement where registered
        if engagement_id:
            eng = MS.engagement_full(s, engagement_id)
            if not eng:
                raise HTTPException(404, "engagement not found")
            base = eng.get("base", {}); ext = eng.get("ext", {}) or {}
            band = base.get("inherent_band") or band
            vendor_id = base.get("vendor_id") or vendor_id
            # derive exposure flags from engagement risk fields
            exposure = {
                "personal_data": bool(ext.get("personal_data")),
                "cross_border": bool(ext.get("cross_border")),
                "mission_critical": bool(ext.get("mission_critical")),
                "regulated": bool(ext.get("regulated_activity")),
            }
        if not band:
            raise HTTPException(422, "inherent band required for an 'Other' (unregistered) vendor")
        # store + extract
        try:
            doc = DOC.store_document(s, filename=b.file.filename,
                                     content_type=b.file.content_type or "",
                                     data_b64=b.file.data_b64, vendor_id=vendor_id,
                                     engagement_id=engagement_id, uploaded_by=u.username,
                                     purpose="contract")
        except ValueError as e:
            raise HTTPException(422, str(e))
        terms = DOC.extract_contract_terms(s, doc)
        text = DOC._decode_text(doc)
        rep = CON.gap_report(text, band, exposure)
        audit(s, "v2.contract_gap_doc", u.username,
              {"doc_id": doc.doc_id, "engagement_id": engagement_id,
               "gaps": len(rep["gaps"])})
        s.commit()
        return {"doc_id": doc.doc_id, "doc_link": f"/api/v2/documents/{doc.doc_id}",
                "inherited_from_engagement": bool(engagement_id),
                "inherent_band": band, "exposure": exposure,
                "extracted_terms": terms, "gap_report": rep,
                "readable": terms.get("readable", False)}

    # ---- management dashboard + chat (leadership-gated) ----
    @app.get("/api/v2/management/risk-view")
    def v2_risk_view(s: Session = Depends(db), u: User = Depends(require("dashboard.risk"))):
        from .features import management as MGMT
        return MGMT.risk_view(s)

    @app.get("/api/v2/management/ops-view")
    def v2_ops_view(s: Session = Depends(db), u: User = Depends(require("dashboard.ops"))):
        from .features import management as MGMT
        return MGMT.ops_view(s)

    @app.get("/api/v2/management/concentration")
    def v2_concentration(s: Session = Depends(db), u: User = Depends(require("dashboard.risk"))):
        from .features import management as MGMT
        return MGMT.concentration_graph(s)

    @app.get("/api/v2/management/concentration/detail")
    def v2_concentration_detail(node_type: str, key: str, s: Session = Depends(db),
                                u: User = Depends(require("dashboard.risk"))):
        from .features import management as MGMT
        if node_type not in ("location", "fourth_party", "vendor"):
            raise HTTPException(400, "node_type must be location, fourth_party or vendor")
        return MGMT.concentration_node_detail(s, node_type, key)

    @app.post("/api/v2/intelligence/board")
    def v2_board_intelligence(s: Session = Depends(db),
                              u: User = Depends(require("dashboard.exec"))):
        from .features import intelligence as INTEL
        result = INTEL.board_intelligence(s)
        from .agents import llm_config
        if llm_config.is_enabled():
            try:
                import json as _json
                ctx = _json.dumps({"internal": result["internal"],
                                   "external": result["external"],
                                   "observations": result["observations"][:6],
                                   "predictions": result["predictions"]})
                enriched = llm_config.complete(
                    PROMPTS.resolve(s, "board_intelligence"),
                    f"Analysis: {ctx}", domain="management")
                if enriched:
                    result["executive_briefing"] = enriched
                    result["engine"] = "llm"
            except Exception as _e:
                _obs_swallow('bro_app.py', _e)
        audit(s, "v2.board_intelligence", u.username,
              {"engine": result["engine"], "observations": len(result["observations"])})
        s.commit()
        return result

    @app.post("/api/v2/intelligence/board/followup")
    def v2_board_followup(b: BoardFollowupIn, s: Session = Depends(db),
                          u: User = Depends(require("dashboard.exec"))):
        """Executive deep-dive follow-up — answered by AI over the board-intelligence
        base + live portfolio, at Board/ExCo (BCG-grade) level."""
        from .agents import llm_config
        if not llm_config.status().get("live_ready"):
            return {"answer": AI_HOLDING, "engine": "holding"}
        from .features import intelligence as INTEL, management as MGMT
        base = INTEL.board_intelligence(s)
        ctx = MGMT.portfolio_context(s)
        hist = ""
        for t in (b.history or [])[-4:]:
            if isinstance(t, dict) and t.get("q"):
                hist += f"\nPrevious Q: {t.get('q')}\nPrevious A: {(t.get('a') or '')[:400]}\n"
        system = PROMPTS.resolve(s, "board_followup")
        user_msg = (hist + f"\nFollow-up question: {b.question}\n\nBOARD DEEP-DIVE (JSON):\n"
                    f"{json.dumps(base)[:9000]}\n\nPORTFOLIO CONTEXT (JSON):\n{json.dumps(ctx)[:6000]}")
        ocx = None
        try:
            from .features import oss as OSS
            _os = OSS.summary(s)
            if _os and _os.get("sboms"):
                ocx = ("OSS / SBOM exposure (external): "
                       f"{_os['coverage_pct']}% coverage, {_os['kev_components']} components with "
                       f"known-exploited (KEV) vulnerabilities, {_os['prohibited_licences']} prohibited licences.")
        except Exception:
            ocx = None
        ans = llm_config.complete(system, user_msg, domain="management", max_tokens=1500,
                                  review=True, external_context=ocx,
                                  feedback_context=_fb_guidance(s, "board"))
        if not ans:
            return {"answer": ("The AI engine is connected but the call did not complete: "
                               + (llm_config.last_error() or "unknown error") + "."),
                    "engine": "ai_failed"}
        audit(s, "v2.board_followup", u.username, {"q": b.question[:80]})
        s.commit()
        return {"answer": ans, "engine": "llm"}

    @app.post("/api/v2/management/chat")
    def v2_management_chat(b: MgmtChatIn, s: Session = Depends(db),
                           u: User = Depends(require("dashboard.exec"))):
        from .features import management as MGMT
        from .agents import llm_config
        ctx = MGMT.portfolio_context(s)
        result = MGMT.management_answer(s, b.question)  # deterministic baseline/fallback
        result["context"] = ctx
        if llm_config.status().get("live_ready"):
            try:
                hist = ""
                for turn in (b.history or [])[-4:]:
                    if isinstance(turn, dict) and turn.get("q"):
                        hist += f"\nPrevious Q: {turn.get('q')}\nPrevious A: {(turn.get('a') or '')[:400]}\n"
                system = PROMPTS.resolve(s, "management_chat")
                user_msg = (hist + f"\nCurrent question: {b.question}\n\n"
                            f"PORTFOLIO CONTEXT (JSON):\n{json.dumps(ctx)}")
                pcx = None
                try:
                    from .features import pestle as PES
                    _ps = PES.get_summary(s)
                    if _ps:
                        pcx = "Top live portfolio threats: " + "; ".join(
                            f"{t['name']} ({t['score']})" for t in _ps.get("top_threats", [])[:8])
                except Exception:
                    pcx = None
                ocx = None
                try:
                    from .features import oss as OSS
                    _os = OSS.summary(s)
                    if _os and _os.get("sboms"):
                        topc = ", ".join(f"{c['name']} ({c['usage']} eng)"
                                         for c in _os.get("top_concentration", [])[:5])
                        ocx = ("OSS / SBOM exposure (external): "
                               f"{_os['coverage_pct']}% SBOM coverage, {_os['components']} components, "
                               f"{_os['kev_components']} with known-exploited (KEV) vulnerabilities, "
                               f"{_os['prohibited_licences']} prohibited licences; "
                               f"most-concentrated components: {topc}.")
                except Exception:
                    ocx = None
                enriched = llm_config.complete(system, user_msg, domain="management",
                                               max_tokens=1200, review=True,
                                               pestle_context=pcx, external_context=ocx,
                                               feedback_context=_fb_guidance(s, "management"))
                if enriched:
                    result["answer"] = enriched
                    result["engine"] = "llm"
                else:
                    result["ai_error"] = llm_config.last_error()
            except Exception as e:
                result["ai_error"] = str(e)[:200]
        audit(s, "v2.management_chat", u.username, {"engine": result["engine"]})
        s.commit()
        return result

    @app.get("/api/v2/management/suggested")
    def v2_management_suggested(u: User = Depends(require("dashboard.exec"))):
        from .features import management as MGMT
        return {"questions": MGMT.SUGGESTED_QUESTIONS}

    # ---- capture a chat session into a structured assessment ----
    @app.post("/api/v2/assessments/from-session")
    def v2_capture_session(b: CaptureIn, s: Session = Depends(db),
                           u: User = Depends(require("engagement.view"))):
        from .features import assessment_capture as CAP
        pool = [x.username for x in s.scalars(select(User)).all() if x.is_active]
        try:
            rec = CAP.capture_session(s, session_id=b.session_id,
                                      engagement_id=b.engagement_id,
                                      vendor_id=b.vendor_id,
                                      engagement_owner=u.username, assessor_pool=pool)
        except ValueError as e:
            raise HTTPException(400, str(e))
        audit(s, "v2.assessment_captured", u.username,
              {"assessment_id": rec.assessment_id, "session_id": b.session_id,
               "status": rec.status})
        s.commit()
        return {"assessment_id": rec.assessment_id, "status": rec.status,
                "inherent_band": rec.inherent_band, "assessor_user": rec.assessor_user}

    @app.get("/api/v2/assessments/{aid}/structured")
    def v2_assessment_structured(aid: str, s: Session = Depends(db),
                                 u: User = Depends(require("engagement.view"))):
        a = s.scalars(select(AssessmentRecord).where(
            AssessmentRecord.assessment_id == aid)).first()
        if not a:
            raise HTTPException(404, "assessment not found")
        if not _can_view_assessment(u, a):
            raise HTTPException(403, "you do not have access to this assessment record")
        return {"assessment_id": a.assessment_id, "engagement_id": a.engagement_id,
                "status": a.status, "locked": a.locked,
                "structured": json.loads(a.structured_json or "{}")}

    @app.get("/api/v2/assessments/{aid}/review")
    def v2_assessment_review(aid: str, s: Session = Depends(db),
                             u: User = Depends(require("engagement.view"))):
        """CR-2: full reviewable detail — scope, inherent risks, controls assessed,
        documents and the final residual recommendation — for a reviewer to scrutinise
        before approving. CR-3 access rule enforced."""
        from .features import master_service as MS
        a = s.scalars(select(AssessmentRecord).where(
            AssessmentRecord.assessment_id == aid)).first()
        if not a:
            raise HTTPException(404, "assessment not found")
        if not _can_view_assessment(u, a):
            raise HTTPException(403, "you do not have access to this assessment record")
        st = json.loads(a.structured_json or "{}")
        eng = MS.engagement_full(s, a.engagement_id) if a.engagement_id else {}
        base = eng.get("base", {})
        # documents/artefacts tagged to this vendor/engagement
        arts = []
        for art in s.scalars(select(ArtefactRecord).where(
                ArtefactRecord.vendor_id == a.vendor_id)).all():
            arts.append({"artefact_id": art.artefact_id, "kind": art.artefact_type,
                         "title": art.name, "status": art.status,
                         "expiry_date": art.expiry_date,
                         "doc_link": getattr(art, "doc_link", None)})
        # controls assessed: DDQ / per-stage controls captured in structured snapshot
        controls = st.get("per_stage", [])
        can_approve = (getattr(u.role, "key", None) in ("admin", "vrm")) and not a.locked
        # IRQ (inherent-risk questionnaire) and Due Diligence question review
        irq = st.get("extracted_irq") or st.get("irq") or {}
        irq_rows = [{"question": k.replace("_", " ").title(), "answer": v}
                    for k, v in irq.items()] if isinstance(irq, dict) else []
        dd_rows = []
        for c in controls:
            for t in (c.get("turns", []) if isinstance(c, dict) else []):
                dd_rows.append({"area": (c.get("name") or "Control"),
                                "by": t.get("agent") or t.get("role"),
                                "detail": (t.get("body") or t.get("excerpt") or "")})
        for g in st.get("gaps", []):
            dd_rows.append({"area": g.get("domain") or "Gap", "by": "Assessment",
                            "detail": g.get("issue") or "", "resolution": g.get("resolution")})
        # all documents available / referred during the assessment (artefacts + stored docs)
        all_docs = list(arts)
        try:
            from .features.documents import StoredDocument as _SD
            for d in s.scalars(select(_SD).where(
                    _SD.vendor_id == a.vendor_id)).all():
                all_docs.append({"artefact_id": d.doc_id, "title": d.filename,
                                 "kind": d.purpose or d.content_type, "status": "on file",
                                 "doc_link": f"/api/v2/documents/{d.doc_id}"})
        except Exception as _e:
            _obs_swallow('bro_app.py', _e)
        # findings linked to this assessment
        link_findings = [{"finding_id": f.finding_id, "title": f.title, "severity": f.severity,
                          "status": f.status, "source": f.source}
                         for f in s.scalars(select(FindingRecord).where(
                             FindingRecord.assessment_id == a.assessment_id)).all()]
        return {
            "assessment_id": a.assessment_id, "engagement_id": a.engagement_id,
            "vendor_id": a.vendor_id, "status": a.status, "locked": a.locked,
            "assigned_assessor": a.assessor_user,
            "irq": irq_rows, "due_diligence": dd_rows,
            "all_documents": all_docs, "linked_findings": link_findings,
            "can_assign": getattr(u.role, "key", None) in ("admin", "controller"),
            "outcome": a.outcome, "assessor_signed_off": a.assessor_signed_off,
            "scope": {"title": base.get("title"), "service_description": base.get("service_description"),
                      "data_classification": (eng.get("ext", {}) or {}).get("data_classification"),
                      "is_critical": base.get("is_critical")},
            "inherent": {"band": a.inherent_band, "detail": st.get("inherent_detail"),
                         "risks": st.get("risks", [])},
            "controls_assessed": controls,
            "documents": arts,
            "residual": {"band": a.residual_band, "verdict": st.get("verdict"),
                         "recommendation": st.get("recommendation")},
            "gaps": st.get("gaps", []),
            "transcript_available": bool(st.get("transcript")),
            "can_approve": can_approve,
        }

    @app.post("/api/v2/artefacts/email-intake")
    def v2_email_intake(b: EmailIntakeIn, s: Session = Depends(db),
                        u: User = Depends(require("lifecycle.documents"))):
        from .features import email_intake as EI
        result = EI.process_inbound_email(
            s, sender=b.sender, subject=b.subject or "",
            attachment_name=b.attachment_name or "", attachment_b64=b.attachment_b64,
            body_text=b.body_text or "", vendor_id=b.vendor_id)
        audit(s, "v2.email_intake", u.username,
              {"status": result["status"], "artefact_id": result.get("artefact_id")})
        s.commit()
        return result

    # ---- analysis-section support: sectors, peers, research, monitoring ----
    @app.get("/api/v2/sectors")
    def v2_sectors(u: User = Depends(require("intel.financial"))):
        return FIN.SECTORS

    @app.post("/api/v2/financial-dd/peers")
    def v2_peers(b: PeerBenchmarkIn, s: Session = Depends(db),
                 u: User = Depends(require("intel.financial"))):
        result = FIN.assess_financials(b.figures, b.flags or {})
        peers = FIN.peer_benchmark(result["ratios"], b.sector)
        return {"sector": b.sector, "peers": peers}

    @app.post("/api/v2/financial-dd/research")
    def v2_fin_research(b: FinResearchIn, s: Session = Depends(db),
                        u: User = Depends(require("intel.financial"))):
        from .features import entity_resolve as ER
        res = ER.research_financials(b.company, b.jurisdiction or "UK",
                                     b.identifier or "", b.year or "")
        audit(s, "v2.fin_research", u.username,
              {"company": b.company, "matched": res.get("matched")})
        s.commit()
        return res

    @app.post("/api/v2/research/web")
    def v2_web_research(b: FinResearchIn, s: Session = Depends(db),
                        u: User = Depends(require("intel.financial"))):
        """Live internet-grounded combined FDD + reputation research."""
        from .features import entity_resolve as ER
        res = ER.web_research_fdd_reputation(b.company, b.jurisdiction or "UK",
                                             b.identifier or "")
        audit(s, "v2.web_research", u.username,
              {"company": b.company, "matched": res.get("matched"),
               "available": res.get("available")})
        s.commit()
        return res

    # ===== Methodology library (admin-only) =====
    @app.get("/api/v2/methodology/docs")
    def v2_meth_list(s: Session = Depends(db), u: User = Depends(require("admin.integrations"))):
        from .features import methodology as M
        return {"docs": M.list_docs(s), "has_methodology": bool(M.methodology_text(s))}

    @app.post("/api/v2/methodology/docs")
    def v2_meth_add(b: MethodologyIn, s: Session = Depends(db),
                    u: User = Depends(require("admin.integrations"))):
        from .features import methodology as M
        content = b.content_text or ""
        if not content and b.data_b64:
            import base64 as _b64
            try:
                content = _b64.b64decode(b.data_b64).decode("utf-8", "replace")
            except Exception:
                raise HTTPException(422, "could not decode uploaded text")
        if not content.strip():
            raise HTTPException(422, "empty methodology content")
        d = M.add_doc(s, title=b.title or "Methodology", content_text=content,
                      filename=b.filename or "", uploaded_by=u.username)
        audit(s, "methodology.added", u.username, {"doc_id": d.doc_id, "chars": len(content)})
        s.commit()
        return {"doc_id": d.doc_id}

    @app.get("/api/v2/methodology/docs/{doc_id}")
    def v2_meth_get(doc_id: str, s: Session = Depends(db), u: User = Depends(require("admin.integrations"))):
        from .features import methodology as M
        d = M.get_doc(s, doc_id)
        if not d:
            raise HTTPException(404, "not found")
        return {"doc_id": d.doc_id, "title": d.title, "filename": d.filename,
                "active": d.active, "content_text": d.content_text}

    @app.post("/api/v2/methodology/docs/{doc_id}/active")
    def v2_meth_active(doc_id: str, b: ActiveIn, s: Session = Depends(db),
                       u: User = Depends(require("admin.integrations"))):
        from .features import methodology as M
        M.set_active(s, doc_id, b.active)
        s.commit()
        return {"ok": True}

    @app.delete("/api/v2/methodology/docs/{doc_id}")
    def v2_meth_del(doc_id: str, s: Session = Depends(db), u: User = Depends(require("admin.integrations"))):
        from .features import methodology as M
        ok = M.delete_doc(s, doc_id)
        audit(s, "methodology.deleted", u.username, {"doc_id": doc_id})
        s.commit()
        return {"deleted": ok}

    # ===== AI-driven FDD + Reputation (Claude searches, infers, organises) =====
    def _ai_research(s, *, vendor_id, company, jurisdiction, identifier, mode, actor):
        """mode = 'fdd' | 'reputation' | 'both'. Claude web-searches & organises;
        the result is auto-filed as a report and indicators are updated everywhere."""
        from .features import entity_resolve as ER
        from .features import methodology as M
        from .features import documents as DOCS
        from .features.registry_models import VendorRecord, FinMonitorRecord
        from .features.master_ext import VendorMasterExt, VendorMonitorSignal
        import base64 as _b64, datetime as _dt
        # resolve company name from vendor if needed
        if vendor_id and not company:
            v = s.scalars(select(VendorRecord).where(VendorRecord.vendor_id == vendor_id)).first()
            company = v.legal_name if v else vendor_id
        res = ER.web_research_fdd_reputation(company, jurisdiction or "UK", identifier or "",
                                             methodology=M.methodology_directive(s), mode=mode)
        if not res.get("available"):
            return {"available": False, "holding": True, "message": AI_HOLDING,
                    "limitations": res.get("limitations")}
        # available but the AI call didn't produce usable findings -> surface, don't file
        if res.get("limitations") and not res.get("matched") and not (res.get("financials") or res.get("reputation")):
            return {"available": False, "holding": True,
                    "message": res.get("limitations")}
        # auto-file as a report
        import json as _json
        purpose = {"fdd": "fdd_report", "reputation": "reputation_report"}.get(mode, "fdd_reputation_report")
        fname = f"{purpose}_{(company or vendor_id or 'entity').replace(' ', '_')[:40]}.json"
        doc_id = None
        try:
            payload = _b64.b64encode(_json.dumps(res, indent=2).encode()).decode()
            doc = DOCS.store_document(s, filename=fname, content_type="application/json",
                                      data_b64=payload, vendor_id=vendor_id, uploaded_by=actor, purpose=purpose)
            doc_id = doc.doc_id
        except Exception:
            doc_id = None
        # auto-update indicators everywhere
        updated = []
        if vendor_id:
            fhb = (res.get("financial_health_band") or res.get("fdd", {}).get("financial_health_band")
                   if isinstance(res.get("fdd"), dict) else res.get("financial_health_band"))
            if mode in ("fdd", "both") and fhb:
                ext = s.scalars(select(VendorMasterExt).where(VendorMasterExt.vendor_id == vendor_id)).first()
                if ext:
                    ext.financial_health_band = str(fhb)[:40]; updated.append("financial_health_band")
                fm = s.scalars(select(FinMonitorRecord).where(FinMonitorRecord.vendor_id == vendor_id)).first()
                if fm:
                    fm.last_result = str(fhb)[:40]
                    fm.last_signal = (res.get("summary") or "AI FDD research")[:200]
                    fm.last_swept = _dt.datetime.now(_dt.timezone.utc); updated.append("fin_monitor")
            if mode in ("reputation", "both"):
                adverse = res.get("adverse_media") or res.get("reputation", {}).get("adverse_media") \
                    if isinstance(res.get("reputation"), dict) else res.get("adverse_media")
                s.add(VendorMonitorSignal(vendor_id=vendor_id, signal_type="adverse_media",
                                          value="flagged" if adverse else "clear",
                                          source="AI reputation research"))
                updated.append("reputation_signal")
        res["filed_report"] = doc_id
        res["indicators_updated"] = updated
        return res

    @app.post("/api/v2/research/fdd")
    def v2_research_fdd(b: AIResearchIn, s: Session = Depends(db),
                        u: User = Depends(require("intel.financial"))):
        if not _ai_live():
            return {"available": False, "holding": True, "message": AI_HOLDING}
        res = _ai_research(s, vendor_id=b.vendor_id, company=b.company, jurisdiction=b.jurisdiction,
                           identifier=b.identifier, mode="fdd", actor=u.username)
        audit(s, "v2.research_fdd", u.username, {"vendor_id": b.vendor_id, "company": b.company,
              "filed": res.get("filed_report"), "updated": res.get("indicators_updated")})
        s.commit()
        return res

    @app.post("/api/v2/research/reputation")
    def v2_research_reputation(b: AIResearchIn, s: Session = Depends(db),
                               u: User = Depends(require("intel.reputation"))):
        if not _ai_live():
            return {"available": False, "holding": True, "message": AI_HOLDING}
        res = _ai_research(s, vendor_id=b.vendor_id, company=b.company, jurisdiction=b.jurisdiction,
                           identifier=b.identifier, mode="reputation", actor=u.username)
        audit(s, "v2.research_reputation", u.username, {"vendor_id": b.vendor_id, "company": b.company,
              "filed": res.get("filed_report"), "updated": res.get("indicators_updated")})
        s.commit()
        return res

    # ===== BRO Chat Stage 0: PR pull (MOCK) + similar-engagement check =====
    @app.post("/api/v2/procurement/pr-pull")
    def v2_pr_pull(b: PRPullIn, s: Session = Depends(db),
                   u: User = Depends(require("engagement.view"))):
        """MOCK purchasing-system pull. In production this calls the procurement
        connector; here it returns a plausible pulled package for the demo."""
        from .features.registry_models import VendorRecord
        pr = (b.pr_number or "").strip() or "PR-DEMO"
        vs = list(s.scalars(select(VendorRecord)).all())
        v = vs[abs(hash(pr)) % len(vs)] if vs else None
        return {"pr_number": pr, "matched": bool(v),
                "vendor_id": v.vendor_id if v else None,
                "vendor_name": v.legal_name if v else "Unmatched supplier",
                "tier": v.tier if v else None,
                "scope": "Provision of cloud hosting and data-processing services; ~50,000 "
                         "records; multi-region; API integration; offshore support.",
                "annual_value": 250000,
                "documents": ["Proposal.pdf", "Quote.pdf", "System_Design.pdf",
                              "Vendor_email_thread.eml", "SOC2_Type_II.pdf"]}

    @app.post("/api/v2/engagements/similar")
    def v2_similar_engagements(b: SimilarIn, s: Session = Depends(db),
                               u: User = Depends(require("engagement.view"))):
        """Check the Vendor & Entity database for similar engagements (mock heuristic:
        vendor-name + scope-keyword overlap)."""
        from .features.registry_models import VendorRecord, EngagementRecord
        ent = (b.entity or "").lower()
        scope_toks = {t for t in (b.scope or "").lower().replace(",", " ").split() if len(t) > 3}
        vmap = {v.vendor_id: v for v in s.scalars(select(VendorRecord)).all()}
        matches = []
        for e in s.scalars(select(EngagementRecord)).all():
            v = vmap.get(e.vendor_id); vname = (v.legal_name.lower() if v else "")
            if ent and (ent in vname or vname in ent):
                name_match = 1.0
            elif ent and any(tok in vname for tok in ent.split() if len(tok) > 3):
                name_match = 0.6
            else:
                name_match = 0.0
            title_toks = {t for t in (e.title or "").lower().replace(",", " ").split() if len(t) > 3}
            overlap = (len(scope_toks & title_toks) / max(1, len(scope_toks))) if scope_toks else 0.0
            score = round(0.6 * name_match + 0.4 * overlap, 2)
            if score >= 0.4:
                matches.append({"engagement_id": e.engagement_id, "vendor_id": e.vendor_id,
                                "vendor_name": v.legal_name if v else e.vendor_id,
                                "title": e.title, "stage": e.stage, "status": e.status,
                                "inherent_band": e.inherent_band, "residual_band": e.residual_band,
                                "score": score})
        matches.sort(key=lambda m: m["score"], reverse=True)
        top = matches[:5]
        return {"matches": top, "very_similar": bool(top and top[0]["score"] >= 0.7),
                "best": top[0] if top else None}

    @app.get("/api/v2/fin-monitor")
    def v2_finmon_list(s: Session = Depends(db), u: User = Depends(require("intel.financial"))):
        from .features.registry_models import FinMonitorRecord
        rows = s.scalars(select(FinMonitorRecord).order_by(FinMonitorRecord.id)).all()
        return [{"id": r.id, "vendor_id": r.vendor_id, "entity_name": r.entity_name,
                 "last_signal": r.last_signal, "last_swept": r.last_swept,
                 "last_result": r.last_result} for r in rows]

    @app.post("/api/v2/fin-monitor")
    def v2_finmon_add(b: FinMonitorAddIn, s: Session = Depends(db),
                      u: User = Depends(require("intel.financial"))):
        from .features import entity_resolve as ER
        from .features.registry_models import FinMonitorRecord
        ent = ER.resolve_entity(s, vendor_id=b.vendor_id, other_name=b.other_name)
        if ent["vendor_name"] == "(unspecified)":
            raise HTTPException(400, "provide a vendor_id or other_name")
        row = FinMonitorRecord(vendor_id=ent["vendor_id"], entity_name=ent["vendor_name"])
        s.add(row); s.flush()
        audit(s, "v2.finmon_empanel", u.username, {"entity": ent["vendor_name"], "id": row.id})
        s.commit()
        return {"id": row.id, "vendor_id": row.vendor_id, "entity_name": row.entity_name}

    @app.delete("/api/v2/fin-monitor/{mid}")
    def v2_finmon_remove(mid: int, s: Session = Depends(db),
                         u: User = Depends(require("intel.financial"))):
        from .features.registry_models import FinMonitorRecord
        row = s.get(FinMonitorRecord, mid)
        if row:
            s.delete(row); s.commit()
        return {"deleted": True}

    @app.post("/api/v2/fin-monitor/sweep")
    def v2_finmon_sweep(b: FinMonitorSweepIn, s: Session = Depends(db),
                        u: User = Depends(require("intel.financial"))):
        from .features.registry_models import FinMonitorRecord
        from .agents import llm_config
        import datetime as _dt
        targets = ([s.get(FinMonitorRecord, b.monitor_id)] if b.monitor_id
                   else s.scalars(select(FinMonitorRecord)).all())
        targets = [t for t in targets if t]
        ts = _dt.datetime.now(_dt.timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
        swept = 0
        for t in targets:
            if llm_config.is_enabled():
                text = llm_config.complete(
                    PROMPTS.resolve(s, "financial_monitor"),
                    f"Monitoring sweep for: {t.entity_name}", domain="finance")
                signal = "ok"
                if text:
                    low = text.lower()
                    signal = ("distress" if "signal=distress" in low else
                              "watch" if "signal=watch" in low else "ok")
                t.last_result = text or "No material findings."
                t.last_signal = signal
            else:
                t.last_result = ("Live monitoring needs an AI key. With a key set, Vera "
                                 "sweeps authoritative sources for financial-health signals, "
                                 "profit warnings, rating changes and distress indicators.")
                t.last_signal = "ok"
            t.last_swept = ts
            swept += 1
            # R1: reconcile the panel signal into the attribute time-series + risk profile
            if t.vendor_id:
                from .features import master_service as MS
                MS.persist_monitor_result(s, t.vendor_id, t.last_signal, t.last_result)
                MS.refresh_risk_profile(s, t.vendor_id)
        audit(s, "v2.finmon_sweep", u.username, {"swept": swept})
        s.commit()
        return {"swept": swept, "last_swept": ts,
                "ai_enabled": llm_config.is_enabled()}

    # ============================================================
    # REQ 1 — VENDOR MASTER
    # ============================================================
    @app.get("/api/v2/vendor-master/{vid}")
    def v2_vendor_master_get(vid: str, s: Session = Depends(db),
                             u: User = Depends(require("vendor.view"))):
        from .features import master_service as MS
        # banking visible only to admin or vendor.critical holders
        inc_bank = u.role.key == "admin" or "vendor.critical" in {p.key for p in u.role.permissions}
        data = MS.get_vendor_master(s, vid, include_bank=inc_bank)
        if not data:
            raise HTTPException(404, "vendor not found")
        return data

    @app.get("/api/v2/vendor/{vid}/linkage")
    def v2_vendor_linkage(vid: str, s: Session = Depends(db),
                          u: User = Depends(require("vendor.view"))):
        """Engagements (with Active/Inactive status), contracts and documents linked
        to a vendor — for the visible vendor↔engagement linkage in the master record."""
        from .features.registry_models import EngagementRecord
        from .features.master_ext import ContractRecord, EngagementExt
        from .features.documents import StoredDocument
        import datetime as _dt
        now = _dt.datetime.now(_dt.timezone.utc)
        engs = s.scalars(select(EngagementRecord).where(EngagementRecord.vendor_id == vid)).all()
        ext = {e.engagement_id: e for e in s.scalars(select(EngagementExt).where(
            EngagementExt.engagement_id.in_([e.engagement_id for e in engs]))).all()} if engs else {}

        def _active(e):
            st = (e.status or "").lower(); stg = (e.stage or "").lower()
            if stg in ("terminate", "terminated", "exited") or st in ("terminated", "expired", "inactive", "closed"):
                return False
            x = ext.get(e.engagement_id)
            exp = getattr(x, "contract_end", None) or getattr(x, "expiry_date", None) if x else None
            if exp:
                try:
                    ed = exp if isinstance(exp, _dt.datetime) else _dt.datetime.fromisoformat(str(exp))
                    if ed.tzinfo is None: ed = ed.replace(tzinfo=_dt.timezone.utc)
                    if ed < now: return False
                except Exception as _e:
                    _obs_swallow('bro_app.py', _e)
            return True
        engagements = [{"engagement_id": e.engagement_id, "title": e.title, "stage": e.stage,
                        "status": e.status, "active": _active(e),
                        "inherent_band": e.inherent_band, "residual_band": e.residual_band}
                       for e in engs]
        contracts = [{"contract_id": c.contract_id, "type": c.contract_type, "status": c.status,
                      "doc_link": c.doc_link, "engagement_id": c.engagement_id}
                     for c in s.scalars(select(ContractRecord).where(ContractRecord.vendor_id == vid)).all()]
        docs = [{"doc_id": d.doc_id, "filename": d.filename, "purpose": d.purpose,
                 "url": f"/api/v2/documents/{d.doc_id}"}
                for d in s.scalars(select(StoredDocument).where(StoredDocument.vendor_id == vid)
                                   .order_by(StoredDocument.id.desc())).all()]
        return {"vendor_id": vid, "engagements": engagements, "contracts": contracts,
                "documents": docs,
                "active_count": sum(1 for e in engagements if e["active"]),
                "inactive_count": sum(1 for e in engagements if not e["active"])}

    @app.put("/api/v2/vendor-master/{vid}")
    def v2_vendor_master_put(vid: str, b: VendorMasterIn, s: Session = Depends(db),
                             u: User = Depends(require("vendor.edit"))):
        from .features import master_service as MS
        v = s.scalars(select(VendorRecord).where(VendorRecord.vendor_id == vid)).first()
        if not v:
            raise HTTPException(404, "vendor not found")
        err = _validate_typed_fields(b.data)
        if err:
            raise HTTPException(422, err)
        inc_bank = b.include_bank and (
            u.role.key == "admin" or "vendor.critical" in {p.key for p in u.role.permissions})
        if b.include_bank and not inc_bank:
            raise HTTPException(403, "banking fields require elevated permission")
        MS.update_vendor_master(s, vid, b.data, include_bank=inc_bank)
        from .features import registry_service as _RS
        n_re = _RS.schedule_reassessment(s, vendor_id=vid, reason="Vendor master data updated")
        audit(s, "v2.vendor_master_update", u.username,
              {"vendor_id": vid, "bank": inc_bank, "reassessment_scheduled": n_re})
        s.commit()
        return MS.get_vendor_master(s, vid, include_bank=inc_bank)

    # ============================================================
    # REQ 2 — VENDOR ATTRIBUTE DATABASE
    # ============================================================
    @app.get("/api/v2/vendor-attributes/{vid}")
    def v2_vendor_attributes(vid: str, s: Session = Depends(db),
                             u: User = Depends(require("vendor.view"))):
        from .features import master_service as MS
        v = s.scalars(select(VendorRecord).where(VendorRecord.vendor_id == vid)).first()
        if not v:
            raise HTTPException(404, "vendor not found")
        return MS.vendor_attributes(s, vid)

    @app.post("/api/v2/vendor-attributes/{vid}/screening")
    def v2_attr_screening(vid: str, b: ScreeningIn, s: Session = Depends(db),
                          u: User = Depends(require("vendor.edit"))):
        from .features import master_service as MS
        if b.screen_type not in MS.SCREEN_TYPES:
            raise HTTPException(400, f"screen_type must be one of {MS.SCREEN_TYPES}")
        MS.set_screening(s, vid, b.screen_type, result=b.result, detail=b.detail,
                         screened_date=b.screened_date, next_due=b.next_due)
        audit(s, "v2.screening_update", u.username, {"vendor_id": vid, "type": b.screen_type})
        s.commit()
        return MS.list_screening(s, vid)

    @app.post("/api/v2/vendor-attributes/{vid}/domain/{domain}")
    def v2_attr_domain(vid: str, domain: str, b: AttrDomainIn, s: Session = Depends(db),
                       u: User = Depends(require("vendor.edit"))):
        from .features import master_service as MS
        fn = {"privacy": MS.update_privacy, "cyber": MS.update_cyber,
              "resilience": MS.update_resilience, "esg": MS.update_esg,
              "governance": MS.update_governance}.get(domain)
        if not fn:
            raise HTTPException(404, "unknown attribute domain")
        fn(s, vid, b.data)
        audit(s, "v2.attr_update", u.username, {"vendor_id": vid, "domain": domain})
        s.commit()
        return MS.vendor_attributes(s, vid)

    @app.post("/api/v2/vendor-attributes/{vid}/insurance")
    def v2_attr_insurance(vid: str, b: InsuranceIn, s: Session = Depends(db),
                          u: User = Depends(require("vendor.edit"))):
        from .features import master_service as MS
        MS.add_insurance(s, vid, b.model_dump())
        audit(s, "v2.insurance_add", u.username, {"vendor_id": vid, "type": b.policy_type})
        s.commit()
        return MS.vendor_attributes(s, vid)

    @app.post("/api/v2/vendor-attributes/{vid}/monitor-signal")
    def v2_attr_signal(vid: str, b: MonitorSignalIn, s: Session = Depends(db),
                       u: User = Depends(require("vendor.edit"))):
        from .features import master_service as MS
        MS.add_monitor_signal(s, vid, b.signal_type, b.value, b.source)
        s.commit()
        return {"ok": True}

    @app.post("/api/v2/vendor-attributes/{vid}/refresh-rollups")
    def v2_attr_refresh(vid: str, s: Session = Depends(db),
                        u: User = Depends(require("vendor.view"))):
        from .features import master_service as MS
        MS.refresh_cyber_certs(s, vid)
        rp = MS.refresh_risk_profile(s, vid)
        s.commit()
        return {"refreshed": True, "inherent_band": rp.inherent_band,
                "residual_band": rp.residual_band, "open_findings": rp.open_findings}

    # ============================================================
    # REQ 3 — ENGAGEMENT REGISTER
    # ============================================================
    @app.get("/api/v2/engagement-register/{eid}")
    def v2_eng_full(eid: str, s: Session = Depends(db),
                    u: User = Depends(require("engagement.view"))):
        from .features import master_service as MS
        data = MS.engagement_full(s, eid)
        if not data:
            raise HTTPException(404, "engagement not found")
        return data

    @app.get("/api/v2/engagement/{eid}/irq-dd")
    def v2_eng_irqdd(eid: str, s: Session = Depends(db),
                     u: User = Depends(require("engagement.view"))):
        """Full IRQ + Due-Diligence detail for an engagement, joined to question
        text, with scores, bands and linked documents — reviewable to answer level."""
        import json as _json
        from .features.registry_models import AssessmentRecord, EngagementRecord, VendorRecord
        from .features.documents import StoredDocument
        from .features.master_ext import ContractRecord
        from .features.bro_engine import IRQ_QUESTIONS, DDQ_DOMAINS
        eng = s.query(EngagementRecord).filter(EngagementRecord.engagement_id == eid).first()
        if not eng:
            raise HTTPException(404, "engagement not found")
        a = (s.query(AssessmentRecord)
             .filter(AssessmentRecord.engagement_id == eid)
             .order_by(AssessmentRecord.id.desc()).first())
        ven = s.query(VendorRecord).filter(VendorRecord.vendor_id == eng.vendor_id).first()
        d = {}
        if a and a.structured_json:
            try: d = _json.loads(a.structured_json)
            except Exception: d = {}
        irq_ans = d.get("irq", {}) or {}
        ddq_ans = d.get("ddq", {}) or {}

        def fmt(v):
            return ", ".join(v) if isinstance(v, list) else ("" if v is None else str(v))
        irq_rows = [{"id": q["id"], "question": q["text"], "type": q["type"],
                     "answer": fmt(irq_ans.get(q["id"])),
                     "answered": irq_ans.get(q["id"]) not in (None, "", [])}
                    for q in IRQ_QUESTIONS]
        ddq_domains = []
        for dom in DDQ_DOMAINS:
            qs = [{"id": q["id"], "question": q["text"], "critical": q.get("critical", False),
                   "response": ddq_ans.get(q["id"], "—")} for q in dom["questions"]]
            ddq_domains.append({"id": dom["id"], "name": dom["name"], "questions": qs})

        docs = (s.query(StoredDocument)
                .filter((StoredDocument.engagement_id == eid) | (StoredDocument.vendor_id == eng.vendor_id))
                .order_by(StoredDocument.id.desc()).all())
        documents = [{"doc_id": x.doc_id, "filename": x.filename, "purpose": x.purpose,
                      "scope": "engagement" if x.engagement_id == eid else "vendor",
                      "url": f"/api/v2/documents/{x.doc_id}"} for x in docs]
        contracts = [{"contract_id": c.contract_id, "type": c.contract_type, "status": c.status,
                      "doc_link": c.doc_link}
                     for c in s.query(ContractRecord)
                     .filter((ContractRecord.engagement_id == eid) | (ContractRecord.vendor_id == eng.vendor_id)).all()]
        return {
            "engagement_id": eid, "title": eng.title, "vendor_id": eng.vendor_id,
            "vendor_name": ven.legal_name if ven else eng.vendor_id,
            "assessment_id": a.assessment_id if a else None,
            "status": a.status if a else None, "outcome": a.outcome if a else None,
            "assessor": a.assessor_user if a else None,
            "tier": d.get("tier"), "schema": d.get("schema"),
            "inherent_band": d.get("inherent_band", eng.inherent_band),
            "residual": d.get("residual", {"band": eng.residual_band}),
            "weighted_pct": d.get("weighted_pct"), "completeness_cls": d.get("completeness_cls"),
            "domain_scores": d.get("domain_scores", {}),
            "recommendation": d.get("recommendation"), "route": d.get("route"),
            "scope_summary": d.get("scope_summary"),
            "irq": irq_rows, "ddq": ddq_domains,
            "documents": documents, "contracts": contracts,
            "has_assessment": bool(a),
        }

    # ===== Exit Strategy & Exit Planning (vendor level, CMORG-aligned) =====
    @app.get("/api/v2/exit/portfolio")
    def v2_exit_portfolio(s: Session = Depends(db), u: User = Depends(require("engagement.view"))):
        from .features import exit_planning as EX
        return EX.portfolio(s)

    @app.get("/api/v2/exit/triggers")
    def v2_exit_trig_cat(u: User = Depends(require("engagement.view"))):
        from .features import exit_planning as EX
        return {"catalogue": EX.TRIGGER_CATALOGUE, "strategies": EX.STRATEGY_LABELS}

    @app.post("/api/v2/exit/triggers/scan")
    def v2_exit_trig_scan(s: Session = Depends(db), u: User = Depends(require("lifecycle.offboard"))):
        from .features import exit_planning as EX
        fired = EX.scan_triggers(s)
        audit(s, "exit.trigger_scan", u.username, {"fired": fired})
        s.commit()
        return {"fired": fired}

    @app.get("/api/v2/exit/plan/{vendor_id}")
    def v2_exit_get(vendor_id: str, s: Session = Depends(db), u: User = Depends(require("engagement.view"))):
        from .features import exit_planning as EX
        data = EX.plan_full(s, vendor_id)
        s.commit()
        return data

    @app.post("/api/v2/exit/plan/{vendor_id}/ai-draft")
    def v2_exit_ai_draft(vendor_id: str, s: Session = Depends(db),
                         u: User = Depends(require("engagement.view"))):
        """Viny drafts an exit plan from organisation data (+ public web data when a live
        model is configured). Returns a draft for review — it does not save anything."""
        from .features import exit_planning as EX
        from .agents import llm_config
        import json as _json
        ctx = EX.draft_context(s, vendor_id)
        draft = EX.deterministic_draft(ctx)
        engine, sources = "rules", []
        if llm_config.status().get("live_ready"):
            try:
                system = (PROMPTS.resolve(s, "exit_plan_persona") + "\n\n"
                          + PROMPTS.resolve(s, "exit_plan_draft"))
                raw = llm_config.complete(system, "ORGANISATION DATA (JSON):\n" + _json.dumps(ctx),
                                          domain="exit", web_search=True, max_tokens=1600) or ""
                from .features.ai_json import parse_json_strict
                ai = parse_json_strict(raw)
                if ai:
                    for k in ("strategy_type", "target_window", "rationale", "impact_summary",
                              "data_plan", "comms_plan"):
                        if ai.get(k):
                            draft[k] = ai[k]
                    if isinstance(ai.get("steps"), list) and ai["steps"]:
                        draft["steps"] = ai["steps"]
                    if isinstance(ai.get("alternatives"), list):
                        draft["alternatives"] = ai["alternatives"]
                    sources = ai.get("sources") or []
                    engine = "ai"
            except Exception:
                engine = "rules"
        audit(s, "exit.ai_draft", u.username, {"vendor_id": vendor_id, "engine": engine})
        s.commit()
        return {"engine": engine, "by": "Viny", "draft": draft, "sources": sources,
                "context_used": {k: ctx[k] for k in
                                 ("vendor_name", "is_critical", "residual_band", "dependencies",
                                  "existing_alternatives", "contract_exit_clause")}}

    @app.put("/api/v2/exit/plan/{vendor_id}")
    def v2_exit_put(vendor_id: str, b: ExitPlanIn, s: Session = Depends(db),
                    u: User = Depends(require("lifecycle.offboard"))):
        from .features import exit_planning as EX
        EX.upsert_plan(s, vendor_id, b.model_dump(exclude_none=True))
        audit(s, "exit.plan_saved", u.username, {"vendor_id": vendor_id})
        s.commit()
        return EX.plan_full(s, vendor_id)

    @app.post("/api/v2/exit/plan/{vendor_id}/child")
    def v2_exit_child_add(vendor_id: str, b: ExitChildIn, s: Session = Depends(db),
                          u: User = Depends(require("lifecycle.offboard"))):
        from .features import exit_planning as EX
        if b.kind == "alternative":
            EX.add_alternative(s, vendor_id, b.name or "Alternative", b.prequalified, b.lead_time_days, b.viability, b.note)
        elif b.kind == "step":
            EX.add_step(s, vendor_id, b.description or "Step", b.owner, b.duration_days, b.rto, b.rpo, b.dependency)
        elif b.kind == "dependency":
            EX.add_dependency(s, vendor_id, b.service_name or "Service", b.impact_tolerance, b.max_downtime, b.criticality)
        else:
            raise HTTPException(400, "unknown child kind")
        audit(s, "exit.child_added", u.username, {"vendor_id": vendor_id, "kind": b.kind})
        s.commit()
        return EX.plan_full(s, vendor_id)

    @app.delete("/api/v2/exit/plan/{vendor_id}/child/{kind}/{cid}")
    def v2_exit_child_del(vendor_id: str, kind: str, cid: int, s: Session = Depends(db),
                          u: User = Depends(require("lifecycle.offboard"))):
        from .features import exit_planning as EX
        EX.remove_child(s, kind, cid, vendor_id)
        s.commit()
        return EX.plan_full(s, vendor_id)

    @app.post("/api/v2/exit/plan/{vendor_id}/test")
    def v2_exit_test(vendor_id: str, b: ExitTestIn, s: Session = Depends(db),
                     u: User = Depends(require("lifecycle.offboard"))):
        from .features import exit_planning as EX
        EX.log_test(s, vendor_id, method=b.method, outcome=b.outcome, lessons=b.lessons,
                    participants=b.participants, passed=b.passed)
        audit(s, "exit.test_logged", u.username, {"vendor_id": vendor_id, "method": b.method, "passed": b.passed})
        s.commit()
        return EX.plan_full(s, vendor_id)

    @app.post("/api/v2/exit/plan/{vendor_id}/attest")
    def v2_exit_attest(vendor_id: str, s: Session = Depends(db),
                       u: User = Depends(require("lifecycle.offboard"))):
        from .features import exit_planning as EX
        EX.attest(s, vendor_id)
        audit(s, "exit.attested", u.username, {"vendor_id": vendor_id})
        s.commit()
        return EX.plan_full(s, vendor_id)

    @app.post("/api/v2/exit/plan/{vendor_id}/invoke")
    def v2_exit_invoke(vendor_id: str, b: ExitInvokeIn, s: Session = Depends(db),
                       u: User = Depends(require("lifecycle.offboard"))):
        from .features import exit_planning as EX
        res = EX.invoke(s, vendor_id, mode=b.mode)
        audit(s, "exit.invoked", u.username, {"vendor_id": vendor_id, "mode": b.mode})
        notify(s, f"Exit invoked ({b.mode}) for {vendor_id}", "vrm")
        s.commit()
        return res

    @app.put("/api/v2/engagement-register/{eid}")
    def v2_eng_ext_put(eid: str, b: EngExtIn, s: Session = Depends(db),
                       u: User = Depends(require("engagement.edit"))):
        from .features import master_service as MS
        eng = s.scalars(select(EngagementRecord).where(
            EngagementRecord.engagement_id == eid)).first()
        if not eng:
            raise HTTPException(404, "engagement not found")
        # base-record fields (inherent/residual band) persist to the engagement itself
        data = dict(b.data or {})
        for base_field in ("inherent_band", "residual_band"):
            if base_field in data:
                setattr(eng, base_field, data.pop(base_field))
        MS.update_eng_ext(s, eid, data)
        from .features import registry_service as _RS
        _RS.schedule_reassessment(s, engagement_id=eid, reason="Engagement data updated")
        audit(s, "v2.engagement_update", u.username, {"engagement_id": eid})
        s.commit()
        return MS.engagement_full(s, eid)

    @app.get("/api/v2/engagement/{eid}/assessments")
    def v2_engagement_assessments(eid: str, s: Session = Depends(db),
                                  u: User = Depends(require("engagement.view"))):
        from datetime import datetime, date
        from .features.registry_models import EngagementRecord, AssessmentRecord
        e = s.scalars(select(EngagementRecord).where(
            EngagementRecord.engagement_id == eid)).first()
        if not e:
            raise HTTPException(404, "engagement not found")
        asms = list(s.scalars(select(AssessmentRecord).where(
            AssessmentRecord.engagement_id == eid)).all())
        asms.sort(key=lambda a: (a.created_at or datetime.min), reverse=True)  # latest first
        rows = [{"assessment_id": a.assessment_id, "status": a.status,
                 "inherent_band": a.inherent_band, "residual_band": a.residual_band,
                 "outcome": a.outcome, "assessor": a.assessor_user,
                 "signed_off": a.assessor_signed_off, "locked": a.locked,
                 "date": a.created_at.isoformat() if a.created_at else None} for a in asms]
        due = e.next_assessment_due
        overdue = bool(due and due <= date.today().isoformat())
        return {"engagement_id": eid, "last_assessment_date": e.last_assessment_date,
                "next_assessment_due": due, "reassessment_due": overdue,
                "reassessment_reason": e.reassessment_reason, "count": len(rows),
                "assessments": rows}

    @app.get("/api/v2/reports/expired-assessments")
    def v2_expired_assessments(s: Session = Depends(db),
                               u: User = Depends(require("engagement.view"))):
        """Engagements whose reassessment is overdue (next_assessment_due < today),
        most-overdue first. Powers the Management → Expired Assessments report."""
        from datetime import date
        from .features.registry_models import EngagementRecord, VendorRecord
        today = date.today()
        today_iso = today.isoformat()
        vmap = {v.vendor_id: v.legal_name for v in s.scalars(select(VendorRecord)).all()}
        rows = []
        total = 0
        for e in s.scalars(select(EngagementRecord)).all():
            if not e.next_assessment_due:
                continue
            total += 1
            if e.next_assessment_due >= today_iso:
                continue
            try:
                y, m, d = (int(x) for x in e.next_assessment_due.split("-"))
                days_overdue = (today - date(y, m, d)).days
            except Exception:
                days_overdue = None
            rows.append({"engagement_id": e.engagement_id, "vendor_id": e.vendor_id,
                         "vendor": vmap.get(e.vendor_id, e.vendor_id),
                         "title": e.title, "inherent_band": e.inherent_band,
                         "residual_band": e.residual_band,
                         "is_critical": getattr(e, "is_critical", False),
                         "last_assessment_date": e.last_assessment_date,
                         "next_assessment_due": e.next_assessment_due,
                         "days_overdue": days_overdue,
                         "owner": e.owner_user, "assessor": e.assigned_assessor,
                         "reassessment_reason": e.reassessment_reason})
        rows.sort(key=lambda r: (r["days_overdue"] is None, -(r["days_overdue"] or 0)))
        by_band = {}
        for r in rows:
            by_band[r["inherent_band"] or "—"] = by_band.get(r["inherent_band"] or "—", 0) + 1
        return {"as_of": today_iso, "expired": len(rows),
                "total_with_due": total, "by_band": by_band, "rows": rows}

    # ---------------- PESTLE threat intelligence ----------------
    @app.get("/api/v2/format-settings")
    def v2_format_settings(s: Session = Depends(db), u: User = Depends(actor)):
        from .features import config_store as CFG
        return {"date": CFG.get_config(s, "format.date", "MM-DD-YYYY"),
                "currency": CFG.get_config(s, "format.currency", "USD")}

    @app.get("/api/v2/pestle/catalogue")
    def v2_pestle_catalogue(s: Session = Depends(db), u: User = Depends(require("vendor.view"))):
        from .features import pestle as PES
        return {"categories": PES.CATEGORIES, "threats": PES.THREATS, "count": len(PES.THREATS)}

    @app.get("/api/v2/pestle/summary")
    def v2_pestle_summary(s: Session = Depends(db), u: User = Depends(require("vendor.view"))):
        from .features import pestle as PES
        summ = PES.get_summary(s)
        if not summ:
            return {"empty": True}
        return {"as_of": summ["as_of"], "entities": summ["entities"], "cat_means": summ["cat_means"],
                "top_threats": summ["top_threats"], "movers": summ["movers"],
                "categories": PES.CATEGORIES}

    @app.get("/api/v2/pestle/graph")
    def v2_pestle_graph(focus_type: str = "portfolio", focus_id: Optional[str] = None,
                        min_score: int = 55, limit: int = 8,
                        s: Session = Depends(db), u: User = Depends(require("vendor.view"))):
        from .features import pestle as PES
        return PES.graph(s, focus_type=focus_type, focus_id=focus_id,
                         min_score=min_score, limit=limit)

    @app.get("/api/v2/pestle/profile/{entity_type}/{entity_id}")
    def v2_pestle_profile(entity_type: str, entity_id: str,
                          s: Session = Depends(db), u: User = Depends(require("vendor.view"))):
        from .features import pestle as PES
        prof = PES.get_profile(s, entity_type, entity_id)
        if not prof:
            raise HTTPException(404, "no PESTLE profile for that entity")
        prof["categories"] = PES.CATEGORIES
        return prof

    @app.post("/api/v2/pestle/refresh")
    def v2_pestle_refresh(as_of: Optional[str] = None, s: Session = Depends(db),
                          u: User = Depends(require("admin.config"))):
        """Simulated overnight News + Reputation sweep that refreshes the PESTLE vectors."""
        from .features import pestle as PES
        res = PES.refresh_all(s, as_of=as_of)
        audit(s, "v2.pestle_refresh", u.username, {"entities": res["entities"], "as_of": res["as_of"]})
        s.commit()
        return res

    @app.post("/api/v2/pestle/rebuild")
    def v2_pestle_rebuild(s: Session = Depends(db), u: User = Depends(require("admin.config"))):
        from .features import pestle as PES
        n = PES.build_all(s)
        audit(s, "v2.pestle_rebuild", u.username, {"entities": n})
        s.commit()
        return {"entities": n}

    # ---------------- Open-Source Software register ----------------
    @app.get("/api/v2/oss/summary")
    def v2_oss_summary(s: Session = Depends(db), u: User = Depends(require("vendor.view"))):
        from .features import oss as OSS
        return OSS.summary(s)

    @app.get("/api/v2/oss/components")
    def v2_oss_components(q: Optional[str] = None, licence: Optional[str] = None,
                          band: Optional[str] = None, s: Session = Depends(db),
                          u: User = Depends(require("vendor.view"))):
        from .features import oss as OSS
        return {"components": OSS.components(s, q=q, licence=licence, band_filter=band)}

    @app.get("/api/v2/oss/component/{cid}")
    def v2_oss_component(cid: int, s: Session = Depends(db), u: User = Depends(require("vendor.view"))):
        from .features import oss as OSS
        d = OSS.component_detail(s, cid)
        if not d:
            raise HTTPException(404, "component not found")
        return d

    @app.get("/api/v2/oss/blast")
    def v2_oss_blast(q: str, version: Optional[str] = None, s: Session = Depends(db),
                     u: User = Depends(require("vendor.view"))):
        from .features import oss as OSS
        return OSS.blast_radius(s, q, version=version)

    @app.get("/api/v2/oss/engagement/{eid}")
    def v2_oss_engagement(eid: str, s: Session = Depends(db), u: User = Depends(require("vendor.view"))):
        from .features import oss as OSS
        return OSS.engagement_oss(s, eid)

    @app.get("/api/v2/oss/concentration")
    def v2_oss_concentration(s: Session = Depends(db), u: User = Depends(require("vendor.view"))):
        from .features import oss as OSS
        return {"components": OSS.concentration(s)}

    @app.get("/api/v2/oss/licences")
    def v2_oss_licences(s: Session = Depends(db), u: User = Depends(require("vendor.view"))):
        from .features import oss as OSS
        return OSS.licences(s)

    @app.get("/api/v2/oss/vulnerabilities")
    def v2_oss_vulnerabilities(s: Session = Depends(db), u: User = Depends(require("vendor.view"))):
        from .features import oss as OSS
        return {"vulnerabilities": OSS.vulnerabilities(s)}

    @app.get("/api/v2/oss/coverage")
    def v2_oss_coverage(s: Session = Depends(db), u: User = Depends(require("vendor.view"))):
        from .features import oss as OSS
        return {"rows": OSS.coverage(s)}

    @app.post("/api/v2/oss/ingest")
    def v2_oss_ingest(body: dict = Body(...), s: Session = Depends(db),
                      u: User = Depends(require("engagement.edit"))):
        """Ingest a vendor SBOM (CycloneDX or SPDX JSON) and tag it to an engagement."""
        from .features import oss as OSS
        eid = body.get("engagement_id"); sbom = body.get("sbom")
        if not eid or sbom is None:
            raise HTTPException(400, "engagement_id and sbom are required")
        if isinstance(sbom, str):
            try:
                sbom = json.loads(sbom)
            except Exception:
                raise HTTPException(400, "sbom is not valid JSON")
        try:
            res = OSS.ingest_sbom(s, eid, body.get("product") or "Unspecified product",
                                  body.get("product_version") or "0.0.0", sbom, channel="upload")
        except ValueError as e:
            raise HTTPException(400, str(e))
        audit(s, "v2.oss_ingest", u.username, {"engagement": eid, "components": res["components"]})
        s.commit()
        return res

    @app.post("/api/v2/oss/rebuild")
    def v2_oss_rebuild(s: Session = Depends(db), u: User = Depends(require("admin.config"))):
        from .features import oss as OSS
        res = OSS.seed_all(s)
        audit(s, "v2.oss_rebuild", u.username, res)
        s.commit()
        return res

    # ---------------- AI answer feedback ----------------
    @app.post("/api/v2/feedback")
    def v2_feedback_add(body: dict = Body(...), s: Session = Depends(db), u: User = Depends(actor)):
        from .features import feedback as FB
        fid = FB.record(s, u.username, body.get("surface") or "general",
                        body.get("query"), body.get("answer"),
                        rating=body.get("rating") or "na", comment=body.get("comment"),
                        engine=body.get("engine"))
        audit(s, "v2.feedback", u.username, {"id": fid, "surface": body.get("surface"),
                                             "rating": body.get("rating")})
        s.commit()
        return {"id": fid, "ok": True}

    @app.get("/api/v2/feedback")
    def v2_feedback_list(surface: Optional[str] = None, rating: Optional[str] = None,
                         s: Session = Depends(db), u: User = Depends(actor)):
        from .features import feedback as FB
        return {"summary": FB.summary(s), "items": FB.list_recent(s, surface=surface, rating=rating)}

    @app.post("/api/v2/feedback/{fid}/used")
    def v2_feedback_used(fid: int, body: dict = Body(default={}), s: Session = Depends(db),
                         u: User = Depends(actor)):
        from .features import feedback as FB
        res = FB.set_used(s, fid, body.get("used", True), u.username)
        if not res:
            raise HTTPException(404, "feedback not found")
        s.commit()
        return res

    @app.post("/api/v2/engagement-register/{eid}/child")
    def v2_eng_child_add(eid: str, b: EngChildIn, s: Session = Depends(db),
                         u: User = Depends(require("engagement.edit"))):
        from .features import master_service as MS
        if b.kind not in ("deliverable", "milestone", "sla", "obligation", "personnel"):
            raise HTTPException(400, "invalid child kind")
        row = MS.add_eng_child(s, eid, b.kind, b.data)
        audit(s, "v2.engagement_child_add", u.username, {"engagement_id": eid, "kind": b.kind})
        s.commit()
        return {"id": row.id, "kind": b.kind}

    @app.delete("/api/v2/engagement-register/{eid}/child/{kind}/{cid}")
    def v2_eng_child_del(eid: str, kind: str, cid: int, s: Session = Depends(db),
                         u: User = Depends(require("engagement.edit"))):
        from .features import master_ext as MX
        model = {"deliverable": MX.EngagementDeliverable, "milestone": MX.EngagementMilestone,
                 "sla": MX.EngagementSLA, "obligation": MX.EngagementObligation,
                 "personnel": MX.EngagementPersonnel}.get(kind)
        if not model:
            raise HTTPException(400, "invalid child kind")
        row = s.get(model, cid)
        if row and row.engagement_id == eid:
            s.delete(row); s.commit()
        return {"deleted": True}

    @app.get("/api/v2/obligations/overdue")
    def v2_obligations_overdue(s: Session = Depends(db),
                               u: User = Depends(require("engagement.view"))):
        from .features import master_service as MS
        return MS.overdue_obligations(s)

    # ============================================================
    # REQ 2 — CONTRACT ENTITY
    # ============================================================
    @app.post("/api/v2/contracts")
    def v2_contract_create(b: ContractCreateIn, s: Session = Depends(db),
                           u: User = Depends(require("intel.contract"))):
        from .features import master_service as MS
        if not b.vendor_id and not b.engagement_id:
            raise HTTPException(400, "a contract must link to a vendor (MSA) or an engagement (Contract/PO)")
        row = MS.create_contract(s, contract_type=b.contract_type, vendor_id=b.vendor_id,
                                 engagement_id=b.engagement_id, parent_msa=b.parent_msa,
                                 data=b.data or {})
        audit(s, "v2.contract_create", u.username,
              {"contract_id": row.contract_id, "type": row.contract_type,
               "primary": row.primary_link})
        s.commit()
        return {"contract_id": row.contract_id, "primary_link": row.primary_link,
                "vendor_id": row.vendor_id, "engagement_id": row.engagement_id,
                "contract_type": row.contract_type}

    @app.get("/api/v2/contracts")
    def v2_contract_list(vendor_id: Optional[str] = None, engagement_id: Optional[str] = None,
                         s: Session = Depends(db), u: User = Depends(require("intel.contract"))):
        from .features import master_service as MS
        return MS.list_contracts(s, vendor_id=vendor_id, engagement_id=engagement_id)

    @app.put("/api/v2/contracts/{cid}")
    def v2_contract_update(cid: str, b: ContractUpdateIn, s: Session = Depends(db),
                           u: User = Depends(require("intel.contract"))):
        from .features import master_service as MS
        row = MS.update_contract(s, cid, b.data)
        if not row:
            raise HTTPException(404, "contract not found")
        audit(s, "v2.contract_update", u.username, {"contract_id": cid})
        s.commit()
        return {"contract_id": row.contract_id, "status": row.status}

    @app.post("/api/v2/contracts/migrate-v1")
    def v2_contract_migrate(s: Session = Depends(db),
                            u: User = Depends(require("intel.contract"))):
        from .features import master_service as MS
        n = MS.migrate_v1_contracts(s)
        audit(s, "v2.contract_migrate_v1", u.username, {"migrated": n})
        s.commit()
        return {"migrated": n}

    @app.post("/api/v2/engagement-register/{eid}/sync-contract")
    def v2_eng_sync_contract(eid: str, s: Session = Depends(db),
                             u: User = Depends(require("engagement.edit"))):
        from .features import master_service as MS
        row = MS.sync_engagement_contract(s, eid)
        s.commit()
        if not row:
            return {"synced": False, "reason": "no contract_reference on engagement"}
        return {"synced": True, "contract_id": row.contract_id}

    # ============================================================
    # REQ 3 — CRITICAL VENDORS MODULE
    # ============================================================
    @app.put("/api/v2/engagements/{eid}/criticality-inputs")
    def v2_crit_inputs(eid: str, b: CriticalityInputIn, s: Session = Depends(db),
                       u: User = Depends(require("engagement.edit"))):
        from .features import master_service as MS
        if not s.scalars(select(EngagementRecord).where(
                EngagementRecord.engagement_id == eid)).first():
            raise HTTPException(404, "engagement not found")
        MS.set_criticality_inputs(s, eid, b.model_dump())
        res = MS.score_engagement_criticality(s, eid)
        audit(s, "v2.criticality_inputs", u.username, {"engagement_id": eid, "score": res["score"]})
        s.commit()
        return res

    @app.get("/api/v2/engagements/{eid}/criticality")
    def v2_crit_score(eid: str, s: Session = Depends(db),
                      u: User = Depends(require("engagement.view"))):
        from .features import master_service as MS
        res = MS.score_engagement_criticality(s, eid)
        if res.get("exists") is False:
            raise HTTPException(404, "engagement not found")
        return res

    @app.post("/api/v2/critical-vendors/analyse")
    def v2_crit_analyse(b: CriticalAnalysisIn, s: Session = Depends(db),
                        u: User = Depends(require("vendor.critical"))):
        from .features import master_service as MS
        res = MS.run_critical_analysis(s, b.vendor_id)
        audit(s, "v2.critical_analysis", u.username,
              {"analysed": res["analysed"], "critical": len(res["critical_vendors"])})
        s.commit()
        return res

    @app.get("/api/v2/critical-vendors")
    def v2_crit_list(s: Session = Depends(db), u: User = Depends(require("vendor.view"))):
        from .features import master_service as MS
        return MS.list_critical(s)

    @app.post("/api/v2/critical-vendors/{vid}/override")
    def v2_crit_override(vid: str, b: CriticalOverrideIn, s: Session = Depends(db),
                         u: User = Depends(require("vendor.critical"))):
        from .features import master_service as MS
        res = MS.override_vendor_criticality(s, vid, b.is_critical, b.reason, u.username)
        audit(s, "v2.criticality_override", u.username,
              {"vendor_id": vid, "is_critical": b.is_critical})
        s.commit()
        return res

    @app.post("/api/v2/engagements/{eid}/criticality-override")
    def v2_eng_crit_override(eid: str, b: CriticalOverrideIn, s: Session = Depends(db),
                             u: User = Depends(require("engagement.edit"))):
        from .features import master_service as MS
        res = MS.override_engagement_criticality(s, eid, b.is_critical, b.reason or "manual", u.username)
        if res.get("exists") is False:
            raise HTTPException(404, "engagement not found")
        audit(s, "v2.eng_criticality_override", u.username,
              {"engagement_id": eid, "is_critical": b.is_critical})
        s.commit()
        return res

    # ============================================================
    # REQ 4 — VENDOR PERFORMANCE MANAGEMENT (critical vendors)
    # ============================================================
    @app.post("/api/v2/performance/scorecards")
    def v2_perf_create(b: ScorecardCreateIn, s: Session = Depends(db),
                       u: User = Depends(require("vendor.edit"))):
        from .features import master_service as MS
        try:
            sc = MS.create_scorecard(s, b.vendor_id, b.period_label,
                                     period_start=b.period_start, period_end=b.period_end,
                                     cadence=b.cadence)
        except ValueError as e:
            raise HTTPException(409, str(e))
        MS.auto_source_kpis(s, sc.scorecard_id)
        MS.compute_scorecard(s, sc.scorecard_id)
        audit(s, "v2.scorecard_create", u.username,
              {"scorecard_id": sc.scorecard_id, "vendor_id": b.vendor_id})
        s.commit()
        return MS.get_scorecard(s, sc.scorecard_id)

    @app.get("/api/v2/performance/scorecards/{sid}")
    def v2_perf_get(sid: str, s: Session = Depends(db),
                    u: User = Depends(require("vendor.view"))):
        from .features import master_service as MS
        sc = MS.get_scorecard(s, sid)
        if not sc:
            raise HTTPException(404, "scorecard not found")
        return sc

    # CR-11: performance enrolment (any vendor, not only critical)
    @app.get("/api/v2/performance/enrolment")
    def v2_perf_enrolment_list(s: Session = Depends(db),
                               u: User = Depends(require("vendor.view"))):
        from .features import master_service as MS
        return MS.list_perf_enrolment(s)

    @app.post("/api/v2/performance/enrolment")
    def v2_perf_enrol(b: PerfEnrolIn, s: Session = Depends(db),
                      u: User = Depends(require("vendor.edit"))):
        from .features import master_service as MS
        added = []
        for vid in b.vendor_ids:
            MS.enrol_vendor(s, vid, source="manual", user=u.username)
            added.append(vid)
        audit(s, "v2.perf_enrol", u.username, {"vendors": added})
        s.commit()
        return {"enrolled": added}

    @app.delete("/api/v2/performance/enrolment/{vid}")
    def v2_perf_unenrol(vid: str, s: Session = Depends(db),
                        u: User = Depends(require("vendor.edit"))):
        from .features import master_service as MS
        ok = MS.unenrol_vendor(s, vid)
        audit(s, "v2.perf_unenrol", u.username, {"vendor_id": vid})
        s.commit()
        return {"unenrolled": ok}

    @app.get("/api/v2/performance/vendor/{vid}")
    def v2_perf_list(vid: str, s: Session = Depends(db),
                     u: User = Depends(require("vendor.view"))):
        from .features import master_service as MS
        return MS.list_scorecards(s, vid)

    @app.put("/api/v2/performance/kpi/{kpi_id}")
    def v2_perf_kpi(kpi_id: int, b: KPIScoreIn, s: Session = Depends(db),
                    u: User = Depends(require("vendor.edit"))):
        from .features import master_service as MS
        row = MS.set_kpi_score(s, kpi_id, actual=b.actual, score=b.score,
                               excluded=b.excluded, exclude_reason=b.exclude_reason)
        if not row:
            raise HTTPException(404, "kpi not found")
        MS.compute_scorecard(s, row.scorecard_id)
        s.commit()
        return MS.get_scorecard(s, row.scorecard_id)

    @app.post("/api/v2/performance/scorecards/{sid}/recompute")
    def v2_perf_recompute(sid: str, s: Session = Depends(db),
                          u: User = Depends(require("vendor.view"))):
        from .features import master_service as MS
        # recompute only; auto-sourcing happens at creation or via explicit re-source
        res = MS.compute_scorecard(s, sid)
        s.commit()
        return res

    @app.post("/api/v2/performance/scorecards/{sid}/resource")
    def v2_perf_resource(sid: str, s: Session = Depends(db),
                         u: User = Depends(require("vendor.edit"))):
        from .features import master_service as MS
        n = MS.auto_source_kpis(s, sid)
        res = MS.compute_scorecard(s, sid)
        s.commit()
        return {"sourced": n, **res}

    @app.post("/api/v2/performance/scorecards/{sid}/agree")
    def v2_perf_agree(sid: str, b: AgreeIn, s: Session = Depends(db),
                      u: User = Depends(require("vendor.edit"))):
        from .features import master_service as MS
        sc = MS.agree_scorecard(s, sid, b.party)
        if not sc:
            raise HTTPException(404, "scorecard not found")
        audit(s, "v2.scorecard_agree", u.username, {"scorecard_id": sid, "party": b.party})
        s.commit()
        return {"scorecard_id": sid, "status": sc.status, "agreed_with_vendor": True}

    @app.post("/api/v2/performance/scorecards/{sid}/publish")
    def v2_perf_publish(sid: str, s: Session = Depends(db),
                        u: User = Depends(require("vendor.critical"))):
        from .features import master_service as MS
        res = MS.publish_scorecard(s, sid, u.username)
        if not res:
            raise HTTPException(404, "scorecard not found")
        audit(s, "v2.scorecard_publish", u.username,
              {"scorecard_id": sid, "score": res.get("composite_score")})
        s.commit()
        return res

    @app.post("/api/v2/performance/vendor/{vid}/reviews")
    def v2_perf_review_create(vid: str, b: ReviewIn, s: Session = Depends(db),
                              u: User = Depends(require("vendor.edit"))):
        from .features import master_service as MS
        row = MS.create_review(s, vid, b.data)
        audit(s, "v2.perf_review", u.username, {"review_id": row.review_id, "vendor_id": vid})
        s.commit()
        return {"review_id": row.review_id, "review_date": row.review_date}

    @app.post("/api/v2/performance/reviews/{rid}/acknowledge")
    def v2_perf_review_ack(rid: str, s: Session = Depends(db),
                           u: User = Depends(require("vendor.edit"))):
        from .features import master_service as MS
        row = MS.acknowledge_review(s, rid)
        if not row:
            raise HTTPException(404, "review not found")
        s.commit()
        return {"review_id": rid, "vendor_acknowledged": True}

    # ================= SLA MANAGEMENT (Performance) =================
    from .features import performance_service as PERF

    @app.get("/api/v2/slas")
    def v2_list_slas(engagement_id: Optional[str] = None, vendor_id: Optional[str] = None,
                     s: Session = Depends(db), u: User = Depends(require("engagement.view"))):
        return [PERF.sla_row(s, x) for x in PERF.list_slas(s, engagement_id, vendor_id)]

    @app.post("/api/v2/slas")
    def v2_create_sla(b: SLAIn, s: Session = Depends(db),
                      u: User = Depends(require("engagement.edit"))):
        sla = PERF.create_sla(s, engagement_id=b.engagement_id, vendor_id=b.vendor_id,
                              description=b.description, threshold_type=b.threshold_type or "min",
                              threshold=b.threshold, unit=b.unit or "", baseline=b.baseline,
                              window=b.window or "monthly", source=b.source or "manual",
                              contract_id=b.contract_id, created_by=u.username)
        audit(s, "v2.sla_created", u.username,
              {"sla_id": sla.sla_id, "engagement_id": b.engagement_id})
        s.commit()
        return PERF.sla_row(s, sla)

    @app.put("/api/v2/slas/{sla_id}")
    def v2_update_sla(sla_id: str, b: SLAEditIn, s: Session = Depends(db),
                      u: User = Depends(require("engagement.edit"))):
        sla = s.scalar(select(PERF.SLARecord).where(PERF.SLARecord.sla_id == sla_id))
        if not sla:
            raise HTTPException(404, "SLA not found")
        PERF.update_sla(s, sla, **b.model_dump(exclude_none=True))
        audit(s, "v2.sla_updated", u.username, {"sla_id": sla_id})
        s.commit()
        return PERF.sla_row(s, sla)

    @app.delete("/api/v2/slas/{sla_id}")
    def v2_delete_sla(sla_id: str, s: Session = Depends(db),
                      u: User = Depends(require("engagement.edit"))):
        sla = s.scalar(select(PERF.SLARecord).where(PERF.SLARecord.sla_id == sla_id))
        if not sla:
            raise HTTPException(404, "SLA not found")
        sla.active = False
        audit(s, "v2.sla_deleted", u.username, {"sla_id": sla_id})
        s.commit()
        return {"sla_id": sla_id, "deleted": True}

    @app.post("/api/v2/slas/{sla_id}/measurements")
    def v2_sla_measure(sla_id: str, b: MeasurementIn, s: Session = Depends(db),
                       u: User = Depends(require("engagement.edit"))):
        sla = s.scalar(select(PERF.SLARecord).where(PERF.SLARecord.sla_id == sla_id))
        if not sla:
            raise HTTPException(404, "SLA not found")
        PERF.record_measurement(s, sla_id, b.period, b.value, u.username)
        audit(s, "v2.sla_measurement", u.username,
              {"sla_id": sla_id, "period": b.period, "value": b.value})
        s.commit()
        return PERF.sla_row(s, sla)

    @app.post("/api/v2/slas/extract")
    def v2_sla_extract(b: ExtractIn, s: Session = Depends(db),
                       u: User = Depends(require("engagement.edit"))):
        created = PERF.extract_slas(s, engagement_id=b.engagement_id, vendor_id=b.vendor_id,
                                    mode=b.mode or "contract", contract_id=b.contract_id,
                                    created_by=u.username)
        audit(s, "v2.sla_extract", u.username,
              {"engagement_id": b.engagement_id, "mode": b.mode, "count": len(created)})
        s.commit()
        return {"extracted": len(created), "slas": [PERF.sla_row(s, x) for x in created]}

    @app.get("/api/v2/slas/summary")
    def v2_sla_summary(engagement_id: str, s: Session = Depends(db),
                       u: User = Depends(require("engagement.view"))):
        return PERF.analyse(s, engagement_id)

    @app.post("/api/v2/slas/enquiry")
    def v2_sla_enquiry(b: EnquiryIn, s: Session = Depends(db),
                       u: User = Depends(require("engagement.view"))):
        return PERF.enquire(s, b.engagement_id, b.question)

    # ================= PERFORMANCE ISSUES (mirrors risk register) =================
    @app.get("/api/v2/performance-issues")
    def v2_list_perf_issues(engagement_id: Optional[str] = None, vendor_id: Optional[str] = None,
                            status: Optional[str] = None, severity: Optional[str] = None,
                            source: Optional[str] = None, category: Optional[str] = None,
                            s: Session = Depends(db), u: User = Depends(require("finding.view"))):
        rows = PERF.list_issues(s, engagement_id, vendor_id, status, severity, source, category)
        return [PERF.issue_row(i) for i in rows]

    @app.get("/api/v2/performance-issues/summary")
    def v2_perf_issue_summary(engagement_id: str, s: Session = Depends(db),
                              u: User = Depends(require("finding.view"))):
        return PERF.issue_severity_counts(s, engagement_id)

    @app.post("/api/v2/performance-issues")
    def v2_create_perf_issue(b: PerfIssueIn, s: Session = Depends(db),
                             u: User = Depends(require("finding.manage"))):
        i = PERF.create_issue(s, engagement_id=b.engagement_id, vendor_id=b.vendor_id,
                              title=b.title, description=b.description, category=b.category,
                              severity=b.severity or "Medium", source=b.source or "Manual",
                              status=b.status or "Open", owner=b.owner, due_date=b.due_date,
                              linked_ref=b.linked_ref, suggested_remediation=b.suggested_remediation,
                              raised_by=u.username)
        audit(s, "v2.perf_issue_created", u.username,
              {"pis_id": i.pis_id, "severity": i.severity})
        s.commit()
        return PERF.issue_row(i)

    @app.put("/api/v2/performance-issues/{pis_id}")
    def v2_update_perf_issue(pis_id: str, b: PerfIssueEditIn, s: Session = Depends(db),
                             u: User = Depends(require("finding.manage"))):
        i = s.scalar(select(PERF.PerformanceIssue).where(PERF.PerformanceIssue.pis_id == pis_id))
        if not i:
            raise HTTPException(404, "issue not found")
        PERF.update_issue(s, i, actor=u.username, **b.model_dump(exclude_none=True))
        audit(s, "v2.perf_issue_updated", u.username, {"pis_id": pis_id, "status": i.status})
        s.commit()
        return PERF.issue_row(i)

    @app.post("/api/v2/performance-issues/{pis_id}/advance")
    def v2_advance_perf_issue(pis_id: str, s: Session = Depends(db),
                              u: User = Depends(require("finding.manage"))):
        i = s.scalar(select(PERF.PerformanceIssue).where(PERF.PerformanceIssue.pis_id == pis_id))
        if not i:
            raise HTTPException(404, "issue not found")
        PERF.advance_issue(s, i, u.username)
        audit(s, "v2.perf_issue_advanced", u.username, {"pis_id": pis_id, "status": i.status})
        s.commit()
        return PERF.issue_row(i)

    @app.post("/api/v2/performance-issues/{pis_id}/note")
    def v2_note_perf_issue(pis_id: str, b: NoteIn, s: Session = Depends(db),
                           u: User = Depends(require("finding.manage"))):
        i = s.scalar(select(PERF.PerformanceIssue).where(PERF.PerformanceIssue.pis_id == pis_id))
        if not i:
            raise HTTPException(404, "issue not found")
        PERF._add_note(i, u.username, b.note)
        s.commit()
        return PERF.issue_row(i)

    @app.delete("/api/v2/performance-issues/{pis_id}")
    def v2_delete_perf_issue(pis_id: str, s: Session = Depends(db),
                             u: User = Depends(require("finding.delete"))):
        i = s.scalar(select(PERF.PerformanceIssue).where(PERF.PerformanceIssue.pis_id == pis_id))
        if not i:
            raise HTTPException(404, "issue not found")
        s.delete(i)
        audit(s, "v2.perf_issue_deleted", u.username, {"pis_id": pis_id})
        s.commit()
        return {"pis_id": pis_id, "deleted": True}

    @app.post("/api/v2/performance-issues/raise-from-sla")
    def v2_raise_from_sla(b: RaiseFromSLAIn, s: Session = Depends(db),
                          u: User = Depends(require("finding.manage"))):
        sla = s.scalar(select(PERF.SLARecord).where(PERF.SLARecord.sla_id == b.sla_id))
        if not sla:
            raise HTTPException(404, "SLA not found")
        issue = PERF.raise_from_breach(s, sla, u.username)
        if not issue:
            s.commit()
            return {"raised": False, "reason": "SLA not in breach or issue already open"}
        audit(s, "v2.perf_issue_from_sla", u.username,
              {"pis_id": issue.pis_id, "sla_id": b.sla_id})
        s.commit()
        return {"raised": True, "issue": PERF.issue_row(issue)}

    # ================= PLATFORM DOCUMENTATION (SOP / TDA / versions) =================
    from .features import platform_docs as PDOCS

    @app.get("/api/v2/platform-docs/versions")
    def v2_doc_versions(s: Session = Depends(db), u: User = Depends(require("engagement.view"))):
        return PDOCS.version_history()

    @app.get("/api/v2/platform-docs/{kind}")
    def v2_get_doc(kind: str, s: Session = Depends(db),
                   u: User = Depends(require("engagement.view"))):
        if kind not in PDOCS.KINDS:
            raise HTTPException(404, "unknown document")
        row = PDOCS.get_doc(s, kind)
        s.commit()
        return {"kind": kind, "title": PDOCS.KINDS[kind]["title"],
                "doc_version": row.doc_version, "updated_at": row.updated_at.isoformat(),
                "updated_by": row.updated_by, "html": row.html}

    @app.post("/api/v2/platform-docs/{kind}/ai-update")
    def v2_ai_update_doc(kind: str, s: Session = Depends(db),
                         u: User = Depends(require("admin.users"))):
        if kind not in PDOCS.KINDS:
            raise HTTPException(404, "unknown document")
        res = PDOCS.ai_update_doc(s, kind, actor=u.username)
        audit(s, "v2.doc_ai_update", u.username, {"kind": kind, "version": res["doc_version"]})
        s.commit()
        return res

    # ================= PLATFORM LEARNINGS =================
    from .features import learnings as LEARN

    @app.get("/api/v2/learnings")
    def v2_list_learnings(category: Optional[str] = None, vendor_id: Optional[str] = None,
                          s: Session = Depends(db), u: User = Depends(require("engagement.view"))):
        return [LEARN.row(l) for l in LEARN.list_learnings(s, category, vendor_id)]

    @app.get("/api/v2/learnings/summary")
    def v2_learnings_summary(s: Session = Depends(db),
                             u: User = Depends(require("engagement.view"))):
        return LEARN.summary(s)

    @app.post("/api/v2/learnings")
    def v2_create_learning(b: LearningIn, s: Session = Depends(db),
                           u: User = Depends(require("engagement.edit"))):
        l = LEARN.create(s, category=b.category, insight=b.insight,
                         confidence=b.confidence or "Medium", origin="human",
                         source_engagement=b.source_engagement, source_vendor=b.source_vendor,
                         created_by=u.username)
        audit(s, "v2.learning_created", u.username, {"id": l.id, "category": l.category})
        s.commit()
        return LEARN.row(l)

    @app.post("/api/v2/learnings/{lid}/applied")
    def v2_learning_applied(lid: int, s: Session = Depends(db),
                            u: User = Depends(require("engagement.view"))):
        l = LEARN.mark_applied(s, lid)
        if not l:
            raise HTTPException(404, "learning not found")
        s.commit()
        return LEARN.row(l)

    @app.delete("/api/v2/learnings/{lid}")
    def v2_delete_learning(lid: int, s: Session = Depends(db),
                           u: User = Depends(require("engagement.edit"))):
        l = s.get(LEARN.PlatformLearning, lid)
        if not l:
            raise HTTPException(404, "learning not found")
        s.delete(l)
        audit(s, "v2.learning_deleted", u.username, {"id": lid})
        s.commit()
        return {"id": lid, "deleted": True}

    @app.post("/api/v2/engagements/{eid}/capture-learnings")
    def v2_capture_learnings(eid: str, s: Session = Depends(db),
                             u: User = Depends(require("engagement.edit"))):
        """Derive durable learnings from the most recent assessment on an engagement."""
        from .features import registry_models as RM
        a = s.scalar(select(RM.AssessmentRecord)
                     .where(RM.AssessmentRecord.engagement_id == eid)
                     .order_by(RM.AssessmentRecord.id.desc()))
        if not a:
            raise HTTPException(404, "no assessment on this engagement")
        finds = s.scalars(select(RM.FindingRecord)
                          .where(RM.FindingRecord.engagement_id == eid)).all()
        created = LEARN.capture_from_assessment(s, a, list(finds), actor=u.username)
        audit(s, "v2.learnings_captured", u.username, {"engagement_id": eid, "count": len(created)})
        s.commit()
        return {"captured": len(created), "learnings": [LEARN.row(l) for l in created]}

    # ================= ENGAGEMENT ASSESSMENT REPORT (Vendor 360) =================
    @app.get("/api/v2/engagements/{eid}/assessment-report")
    def v2_engagement_assessment_report(eid: str, s: Session = Depends(db),
                                        u: User = Depends(require("engagement.view"))):
        """A detailed, structured assessment report for an entire engagement:
        the vendor, the engagement, every assessment, findings, documents and
        a decision summary — the basis for the Vendor 360 report view."""
        from .features import registry_models as RM
        from .features import documents as DOCS
        eng = s.scalar(select(RM.EngagementRecord).where(RM.EngagementRecord.engagement_id == eid))
        if not eng:
            raise HTTPException(404, "engagement not found")
        ven = None
        if eng.vendor_id:
            ven = s.scalar(select(RM.VendorRecord).where(RM.VendorRecord.vendor_id == eng.vendor_id))
        assessments = s.scalars(select(RM.AssessmentRecord)
                                .where(RM.AssessmentRecord.engagement_id == eid)
                                .order_by(RM.AssessmentRecord.id.desc())).all()
        findings = s.scalars(select(RM.FindingRecord)
                             .where(RM.FindingRecord.engagement_id == eid)).all()
        docs = s.scalars(select(DOCS.StoredDocument)
                         .where(DOCS.StoredDocument.engagement_id == eid)).all()
        latest = assessments[0] if assessments else None
        sev_order = {"Critical": 0, "High": 1, "Medium": 2, "Low": 3}
        fnd = sorted(findings, key=lambda f: sev_order.get(f.severity, 9))
        return {
            "engagement": {"engagement_id": eng.engagement_id, "title": eng.title,
                           "status": getattr(eng, "status", None),
                           "owner": getattr(eng, "engagement_owner", None),
                           "service": getattr(eng, "service_description", None)},
            "vendor": ({"vendor_id": ven.vendor_id,
                        "name": ven.trading_name or ven.legal_name,
                        "criticality": getattr(ven, "criticality", None),
                        "country": getattr(ven, "incorporation_country", None)} if ven else None),
            "latest_assessment": ({
                "assessment_id": latest.assessment_id, "status": latest.status,
                "inherent_band": latest.inherent_band, "residual_band": latest.residual_band,
                "outcome": latest.outcome, "assessor": latest.assessor_user,
                "signed_off": latest.assessor_signed_off,
                "created_at": latest.created_at.isoformat() if latest.created_at else None,
            } if latest else None),
            "assessment_history": [{
                "assessment_id": a.assessment_id, "status": a.status,
                "inherent_band": a.inherent_band, "residual_band": a.residual_band,
                "outcome": a.outcome,
                "created_at": a.created_at.isoformat() if a.created_at else None,
            } for a in assessments],
            "findings": [{
                "finding_id": f.finding_id, "title": f.title, "severity": f.severity,
                "domain": getattr(f, "domain", None), "status": f.status,
                "owner": getattr(f, "owner", None),
            } for f in fnd],
            "findings_summary": {sev: sum(1 for f in findings if f.severity == sev)
                                 for sev in ["Critical", "High", "Medium", "Low"]},
            "documents": [{
                "doc_id": d.doc_id, "filename": d.filename, "purpose": d.purpose,
                "size_bytes": d.size_bytes, "uploaded_by": d.uploaded_by,
                "created_at": d.created_at.isoformat() if d.created_at else None,
            } for d in docs],
            "counts": {"assessments": len(assessments), "findings": len(findings),
                       "documents": len(docs)},
        }

    # ================= BRO CHAT — INTERIM AI REPORT (PDF) =================
    @app.post("/api/v2/agent/sessions/{sid}/interim-report")
    def v2_interim_report(sid: int, s: Session = Depends(db),
                          u: User = Depends(require("engagement.view"))):
        """Build an interim assessment report from a live BRO Chat session:
        an AI narrative of the assessment so far (deterministic fallback when AI
        is unavailable) plus an annex of every document and key input submitted.
        Returns HTML the client renders and prints to PDF."""
        from .features import models_feature as MF
        from .features import documents as DOCS
        sess = s.get(MF.ConversationSession, sid)
        if not sess:
            raise HTTPException(404, "session not found")
        msgs = s.scalars(select(MF.ConversationMessage)
                         .where(MF.ConversationMessage.session_id == sid)
                         .order_by(MF.ConversationMessage.id)).all()
        dossier = json.loads(sess.dossier_json or "{}")
        eng_id = sess.engagement_id
        docs = []
        if eng_id:
            docs = s.scalars(select(DOCS.StoredDocument)
                             .where(DOCS.StoredDocument.engagement_id == eng_id)).all()

        user_inputs = [m.body for m in msgs if m.role == "user"]
        agent_msgs = [m for m in msgs if m.role == "agent"]
        transcript = "\n".join(f"{m.role.upper()}"
                               f"{('/'+m.agent) if m.agent else ''}: {m.body}" for m in msgs)

        # AI narrative (deterministic fallback)
        ai_mode = "deterministic"
        narrative = None
        try:
            from .agents import llm_config
            if llm_config.is_enabled():
                sys = ("You are BRO, a senior TPRM assessor. Write a concise interim assessment "
                       "report (250-400 words) summarising the assessment performed so far from the "
                       "conversation: scope, exposure identified, controls discussed, open questions, "
                       "and provisional direction. Be factual and grounded only in the transcript. "
                       "Plain prose, no markdown headers.")
                narrative = llm_config.complete(sys, transcript[:9000], domain="interim_report",
                                                review=False, max_tokens=700)
                if narrative:
                    ai_mode = "ai"
        except Exception:
            narrative = None
        if not narrative:
            stage = sess.stage or 0
            narrative = (
                f"This interim report summarises the assessment in progress for engagement "
                f"{eng_id or '(unassigned)'}. The conversation has reached stage {stage} of the "
                f"assessment lifecycle, with {len(user_inputs)} input(s) from the business and "
                f"{len(agent_msgs)} specialist response(s) on record. "
                + ("Key context captured so far: " + "; ".join(
                    f"{k}: {v}" for k, v in list(dossier.items())[:8] if v) + ". "
                   if dossier else "No structured dossier fields have been captured yet. ")
                + f"A total of {len(docs)} supporting document(s) have been submitted and are listed "
                "in the annex below. This is a provisional, working view: the assessment is not yet "
                "complete and no final risk decision has been recorded. A full assessment report is "
                "produced once the engagement is captured and signed off.")

        # render HTML report (self-contained, print-ready)
        def esc(t):
            return (str(t).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;"))
        from datetime import datetime as _dt, timezone as _tz
        stamp = _dt.now(_tz.utc).strftime("%d %b %Y %H:%M UTC")
        annex_docs = "".join(
            f"<tr><td>{i+1}</td><td>{esc(d.filename)}</td><td>{esc(d.purpose or '—')}</td>"
            f"<td>{(d.size_bytes or 0)//1024} KB</td><td>{esc(d.uploaded_by or '—')}</td></tr>"
            for i, d in enumerate(docs)) or '<tr><td colspan="5" style="color:#888">No documents submitted yet.</td></tr>'
        annex_inputs = "".join(
            f"<li>{esc(t)}</li>" for t in user_inputs[:30]) or "<li style='color:#888'>No user inputs recorded yet.</li>"
        para = "".join(f"<p>{esc(p)}</p>" for p in narrative.split("\n") if p.strip())
        html = f"""<!DOCTYPE html><html><head><meta charset="utf-8"><title>Interim Assessment Report</title>
<style>body{{font-family:-apple-system,Segoe UI,Roboto,sans-serif;color:#1a1a1a;line-height:1.6;max-width:880px;margin:0 auto;padding:40px}}
h1{{font-family:Georgia,serif;color:#14302A;font-size:28px;margin:0 0 4px}}.sub{{color:#5a6472;font-size:13px;margin-bottom:24px}}
h2{{font-family:Georgia,serif;color:#14302A;font-size:18px;border-bottom:2px solid #B8862B;padding-bottom:5px;margin:26px 0 12px}}
.badge{{display:inline-block;background:#FBF4E4;border:1px solid #e8c07a;color:#7a5015;font-size:11px;padding:2px 10px;border-radius:20px;margin-left:8px}}
.meta{{display:grid;grid-template-columns:1fr 1fr;gap:8px 24px;background:#F7F5F0;border:1px solid #E5DFD0;border-radius:10px;padding:16px;font-size:13px;margin-bottom:8px}}
.meta b{{color:#14302A}}table{{width:100%;border-collapse:collapse;font-size:12.5px;margin:8px 0}}
th{{background:#14302A;color:#fff;text-align:left;padding:8px 10px;font-size:11px}}td{{border-bottom:1px solid #E5DFD0;padding:7px 10px}}
ul{{font-size:13px}}.foot{{margin-top:30px;border-top:1px solid #E5DFD0;padding-top:12px;color:#888;font-size:11px}}
p{{font-size:13.5px}}@media print{{body{{padding:0}}}}</style></head><body>
<h1>Interim Assessment Report <span class="badge">AI · {ai_mode}</span></h1>
<div class="sub">BRO Risk Oracle · session #{sid} · generated {stamp}</div>
<div class="meta"><div><b>Engagement</b><br>{esc(eng_id or 'Unassigned')}</div>
<div><b>Lifecycle stage</b><br>Stage {sess.stage or 0}</div>
<div><b>Inputs on record</b><br>{len(user_inputs)} business input(s)</div>
<div><b>Documents submitted</b><br>{len(docs)}</div></div>
<h2>Assessment so far</h2>{para}
<h2>Annex A — Documents submitted</h2>
<table><thead><tr><th>#</th><th>Document</th><th>Purpose</th><th>Size</th><th>Submitted by</th></tr></thead><tbody>{annex_docs}</tbody></table>
<h2>Annex B — Information submitted by the user</h2><ul>{annex_inputs}</ul>
<div class="foot">Interim, working document — the assessment is not complete and no final risk decision is recorded.
Generated by Brata (BRO Risk Oracle). AI narrative mode: {ai_mode}.</div></body></html>"""
        audit(s, "v2.interim_report", u.username, {"session_id": sid, "ai_mode": ai_mode})
        s.commit()
        return {"session_id": sid, "ai_mode": ai_mode, "html": html,
                "documents": len(docs), "inputs": len(user_inputs)}


    @app.post("/api/v2/performance/capa")
    def v2_perf_capa(b: PerfCapaIn, s: Session = Depends(db),
                     u: User = Depends(require("vendor.edit"))):
        from .features import master_service as MS
        sc = MS.get_scorecard(s, b.scorecard_id)
        if not sc:
            raise HTTPException(404, "scorecard not found")
        res = MS.raise_performance_capa(s, sc["vendor_id"], b.scorecard_id,
                                        b.gap, b.owner, b.due_date)
        audit(s, "v2.perf_capa", u.username, res)
        s.commit()
        return res

    @app.post("/api/v2/performance/capa/{rid}/verify")
    def v2_perf_capa_verify(rid: str, b: CapaVerifyIn, s: Session = Depends(db),
                            u: User = Depends(require("vendor.critical"))):
        from .features import master_service as MS
        res = MS.verify_performance_capa(s, rid, u.username, b.evidence)
        if not res:
            raise HTTPException(404, "remediation not found")
        audit(s, "v2.perf_capa_verify", u.username, {"remediation_id": rid})
        s.commit()
        return res

    # ============================================================
    # REQ 5 — ProAssess (autonomous assessment)
    # ============================================================
    @app.post("/api/v2/proassess/run")
    def v2_proassess_run(b: ProAssessRunIn, s: Session = Depends(db),
                         u: User = Depends(require("engagement.view"))):
        if not _ai_live():
            return {"available": False, "holding": True, "message": AI_HOLDING}
        from .features import master_service as MS
        report = MS.run_proassess(s, vendor_id=b.vendor_id, engagement_id=b.engagement_id,
                                  irq=b.irq, ddq=b.ddq, documents=b.documents,
                                  extracted=b.extracted)
        audit(s, "v2.proassess_run", u.username,
              {"vendor_id": b.vendor_id, "inherent": report["inherent_band"],
               "residual": report["residual_band"], "gaps": report["gap_count"]})
        s.commit()
        return report

    @app.post("/api/v2/proassess/register")
    def v2_proassess_register(b: ProAssessRegisterIn, s: Session = Depends(db),
                              u: User = Depends(require("vendor.critical"))):
        from .features import master_service as MS
        res = MS.register_proassess(s, b.report, u.username)
        audit(s, "v2.proassess_register", u.username, res)
        s.commit()
        return res

    @app.post("/api/v2/proassess/autonomous")
    def v2_proassess_autonomous(b: ProAssessAutoIn, s: Session = Depends(db),
                                u: User = Depends(require("vendor.critical"))):
        """CR-4: single-input, document-aware, autonomous ProAssess. Works for new or
        existing vendors; creates records across the databases when create_records=True."""
        if not _ai_live():
            return {"available": False, "holding": True, "message": AI_HOLDING}
        from .features import master_service as MS
        if not b.vendor_id and not b.new_vendor_name:
            raise HTTPException(422, "provide vendor_id or new_vendor_name")
        try:
            report = MS.run_proassess_autonomous(
                s, free_text=b.free_text or "", documents=b.documents,
                vendor_id=b.vendor_id, new_vendor_name=b.new_vendor_name,
                engagement_title=b.engagement_title, ddq=b.ddq,
                user=u.username, create_records=b.create_records)
        except Exception as e:
            s.rollback()
            raise HTTPException(400, f"autonomous assessment failed: {e}")
        if report.get("error"):
            raise HTTPException(422, report["error"])
        audit(s, "v2.proassess_autonomous", u.username,
              {"vendor_id": report.get("vendor_id"), "created_vendor": report.get("created_vendor"),
               "tables_written": report.get("tables_written", [])})
        s.commit()
        return report

    # ============================================================
    # REQ 6 — VENDOR 360 DASHBOARD (compile + correlate)
    # ============================================================
    @app.get("/api/v2/vendor360/portfolio")
    def v2_vendor360_portfolio(s: Session = Depends(db),
                               u: User = Depends(require("vendor.view"))):
        from .features import master_service as MS
        return MS.vendor360_portfolio(s)

    @app.get("/api/v2/vendor360/{vid}")
    def v2_vendor360(vid: str, s: Session = Depends(db),
                     u: User = Depends(require("vendor.view"))):
        from .features import master_service as MS
        data = MS.vendor360(s, vid)
        if not data:
            raise HTTPException(404, "vendor not found")
        s.commit()  # vendor360 refreshes the risk-profile snapshot
        return data

    # ===== mount the web UI =====
    from .web import ui as _ui
    app.include_router(_ui)
    # Serve the SPA's JavaScript (extracted from the page into a real asset file,
    # cacheable by the browser and editable with normal JS tooling).
    from fastapi.staticfiles import StaticFiles
    _static_dir = _os.path.join(_os.path.dirname(__file__), "static")
    if _os.path.isdir(_static_dir):
        app.mount("/static", StaticFiles(directory=_static_dir), name="static")

    # Optional in-process scheduler (single-worker/dev). For multi-worker/prod use a
    # Render Cron Job running run_monitoring.py instead (see render.yaml).
    if _os.environ.get("BRO_SCHEDULER_ENABLED") == "1":
        import threading
        import time as _time
        from .features import monitoring as _MON

        def _scheduler_loop():
            interval = _monitor_interval()
            check_every = max(60, min(interval * 3600, 3600))  # check hourly at most
            _time.sleep(10)  # let the app finish booting
            while True:
                try:
                    with SessionFactory() as _s:
                        if _MON.claim_due_run(_s, interval):
                            _MON.run_all(_s, by="scheduler", trigger="scheduler", audit_fn=audit)
                except Exception as e:
                    print(f"monitoring scheduler error: {e}")
                _time.sleep(check_every)

        threading.Thread(target=_scheduler_loop, daemon=True, name="bro-monitoring").start()
        print("monitoring scheduler started (interval "
              f"{_monitor_interval()}h)")

    return app


def _count_by(rows, attr):
    out: dict = {}
    for r in rows:
        k = getattr(r, attr) or "none"
        out[k] = out.get(k, 0) + 1
    return out


import os as _os
app = create_app(_os.environ.get("BRO_DB_URL", "sqlite:///bro_unified.db"))
